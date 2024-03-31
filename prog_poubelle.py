import openpyxl

"""
writebook = openpyxl.load_workbook('fichiers_xls/pour tony_30032024(1).xlsx')

sheet = writebook['Feuil1']

L = [i for i in range(1,1336)]

for i in range(1,sheet.max_row + 1):
    if sheet.cell(i,1).value in L:
        L.remove(sheet.cell(i,1).value)
print(L, len(L))
"""

writebook = openpyxl.load_workbook('fichiers_xls/vers1415(1).xlsx')
#writebook2 = openpyxl.load_workbook('fichiers_xls/liste_n=1335.xlsx')

sheet = writebook['Feuil1']
sheet2 = writebook['Feuil2']

L1 = []
L2 = [] 

for j in range(2, sheet2.max_row + 1):
    L2.append(sheet2.cell(j,3).value)

print(L2)

for i in range(2, sheet.max_row+1):
    if sheet.cell(i,3).value not in L2:
        L1.append(sheet.cell(i,1).value + sheet.cell(i,2).value)
print(len(L1),L1)



