from unittest import TestCase, main
from module_pour_excel import * 
import openpyxl
import os
 

class TestPath(TestCase):

    """ def test_delete_other_columns(self):
        Pb de ce test : ne fonctionne qu'une fois car une fois les colonnes supprimées il en supprimera d'autres et le test sera faux.
        path = Path('fichiers_xls/gathertests/')
        path.delete_other_columns('A-F,H-J', 'test_keep_only_columns.xlsx', 'sheet1')
        directories = [f for f in os.listdir(path.pathname) if os.path.isdir(os.path.join(path.pathname, f))]

        for directory in directories:
            verify_sheets_identical(Sheet('test_keep_only_columns.xlsx', 'sheet1', path.pathname + '/').sheet,
                                    Sheet('test_keep_only_columns.xlsx', 'Feuille2', path.pathname + '/').sheet) """
    
    def test_create_one_onglet_by_participant(self):
        path = Path('fichiers_xls/gathertests/')
        path.create_one_onglet_by_participant('test_cmd_ongletbyparticipant.xlsx',
                                               'test', 'A', 'divided_test_cmd_ongletbyparticipant.xlsx')
        file = File('divided_test_cmd_ongletbyparticipant_before.xlsx', path.pathname)
        file2 = File('divided_test_cmd_ongletbyparticipant.xlsx', path.pathname)
        
        verify_files_identical(file, file2)
        os.remove(path.pathname + 'divided_test_cmd_ongletbyparticipant.xlsx') 
        

class TestFile(TestCase):
    
    def test_open_and_copy(self): 
        file = File('test_copie.xlsx') 
        self.assertNotEqual(file.writebook,None)  
        print(datetime.now().strftime("%Y-%m-%d %Hh%M"))
        del file

    def test_files_identical(self):
        """On prend deux fichiers excel, on vérifie qu'ils ont les mêmes onglets et que dans chaque onglet on a les mêmes cellules."""
        file1 = File('test_date_2023-05-20.xlsx')
        file2 = File('test_copie.xlsx')
        verify_files_identical(file1, file2)
        del file1, file2

    def test_create_one_onglet_by_participant(self): 
        file = File('test_create_one_onglet_by_participant.xlsx') 
        file.create_one_onglet_by_participant('Stroops_test (7)', 'A', "divided_test_create_one_onglet_by_participant.xlsx", 'fichiers_xls/')  
        file2 = File('divided_test_create_one_onglet_by_participant.xlsx')
        verify_files_identical(File('test_create_one_onglet_by_participant_before.xlsx'),
                               file2)
        
        os.remove("fichiers_xls/divided_test_create_one_onglet_by_participant.xlsx")


    def test_extract_column_from_all_sheets(self):
        file = File('test_extract_column.xlsx')
        file.extract_column_from_all_sheets('B') 
        verify_files_identical(File('test_extract_column_ref.xlsx'),file)

        del file.writebook[file.sheets_name[-1]]
        file.writebook.save(file.path + 'test_extract_column.xlsx') 


    """
    def test_apply_column_formula_on_all_sheets(self):
        file = File('dataset.xlsx', dataonly = False)

        #sauvegarde du fichier initial
        file2 = file.sauvegarde()
        # test sur trois colonnes réparties au hasard
        file.apply_column_formula_on_all_sheets(2,5,10)
        #vérifier sur tous les onglets que ces trois colonnes sont identiques au fichier de réf.
        # on supprime file et on renomme file2 avec le nom de file pour retrouver le fichier initial.
        
        # test sur toutes les colonnes à partir de la colonne 2 
        file.apply_column_formula_on_all_sheets(*[i for i in range(2,6)]) 
        #on refait ce qui a été fait ci-dessus.
    """

    def test_gather_columns_in_one(self):
        file = File("test_gather_columns_in_one.xlsx")
        file.gather_columns_in_one("test", ['C','D','E'], ['G','H','I'])

        verify_files_identical(File("test_gather_columns_in_one - ref.xlsx"), File("test_gather_columns_in_one.xlsx"))

        del file.writebook['onglet 1']
        del file.writebook['onglet 2']
        file.writebook.save(file.path + 'test_gather_columns_in_one.xlsx')
        del file


    def test_one_file_by_tab_sendmail(self):
        file = File("test_onefile_sendmail.xlsx")
        file.one_file_by_tab_sendmail()
 
        sheet1 = Sheet("tony fevrier.xlsx","Sheet", path = "multifiles/")
        sheet2 = Sheet("Marine Moyon.xlsx","Sheet", path = "multifiles/")

        sheet1o = Sheet("test_onefile_sendmail.xlsx","tony fevrier")
        sheet2o = Sheet("test_onefile_sendmail.xlsx","Marine Moyon")

        verify_sheets_identical(sheet1, sheet1o)
        verify_sheets_identical(sheet2, sheet2o) 

    def test_merge_cells_on_all_tabs(self): 
        file1 = File("test_merging.xlsx") 
        file1.merge_cells_on_all_tabs('C', 'D', 5, 7)

        #voir comment tester le fait qu'une cellule est mergée : comprendre l'objet mergedcells
        """ for tab in file1.sheets_name:
            sheet = file1.writebook[tab]
            mergedcells = sheet.merged_cells
            print(mergedcells.ranges, type(mergedcells))
            self.assertEqual('C5' in mergedcells.ranges,True)
            self.assertIn(sheet['C6'],mergedcells)
            self.assertIn(sheet['C7'],mergedcells)
            self.assertIn(sheet['D5'],mergedcells)
            self.assertIn(sheet['D6'],mergedcells) """
        
    def test_apply_cell_formula_on_all_sheets(self):
        file = File("test_merging.xlsx")
        file.apply_cells_formula_on_all_sheets('A10','B10','C10')

        for tab in file.sheets_name[1:]:
            sheet = file.writebook[tab]
            self.assertEqual(sheet['A10'].value, file.writebook[file.sheets_name[0]]['A10'].value)
            self.assertEqual(sheet['B10'].value, file.writebook[file.sheets_name[0]]['B10'].value)
            self.assertEqual(sheet['C10'].value, file.writebook[file.sheets_name[0]]['C10'].value)
    
    def test_check_linenumber_of_tabs(self):
        file = File('test.xlsx')
        tabs = file.check_linenumber_of_tabs(14)
        self.assertListEqual(tabs, ['cutinparts', 'cutinpartsbis', 'delete_lines', 'delete_lines_bis'])

    def test_extract_cells_from_all_sheets(self):
        file = File('test_extract_cells_from_all_sheets.xlsx')
        file.extract_cells_from_all_sheets('C7','D7','C8','D8') 
        verify_files_identical(file, File('test_extract_cells_from_all_sheets - after.xlsx'))
        del file.writebook['gathered_data']
        file.writebook.save(file.path + 'test_extract_cells_from_all_sheets.xlsx')


class TestSheet(TestCase, Other):
    def test_sheet_correctly_opened(self):
        """Ici je teste que l'attribut sheet de la classe sheet contient bien la bonne page correspondant à l'onglet.
        Pour cela, je génère la feuille via mes classes et par la préocédure habituelle et je regarde si la première colonne des deux fichiers se correspondent.""" 
        feuille = Sheet('test.xlsx','sheet1')

        readbook = openpyxl.load_workbook('fichiers_xls/test.xlsx', data_only=True)
        feuille2 = readbook.worksheets[0] 
        for i in range(1,feuille2.max_row):
            self.assertEqual(feuille.sheet.cell(i,1).value,feuille2.cell(i,1).value)
         
    def column_identical(self,name_file1, name_file2, index_onglet1, index_onglet2, column1,column2):
        """
        Méthode qui prend deux fichiers et regarde si à une colonne donnée les valeurs sont les mêmes
        """
        file1 = File(name_file1) 
        file2 = File(name_file2)  
        sheet1 = file1.writebook.worksheets[index_onglet1] 
        sheet2 = file2.writebook.worksheets[index_onglet2] 
        self.assertEqual(sheet1.max_row,sheet2.max_row) 
        
        for i in range(2,sheet1.max_row+1 ): 
            self.assertEqual(sheet1.cell(i,column1).value,sheet2.cell(i,column2).value)

    def test_column_transform_string_in_binary(self):
        sheet = Sheet('test.xlsx','sheet1')  

        sheet.column_transform_string_in_binary(12,13,'partie 1 : Vrai',insert=False,label = False)
        self.column_identical('test.xlsx','test_generated.xlsx',0,0, 13, 13) 
        sheet.column_transform_string_in_binary('L','M','partie 1 : Vrai',insert=False)
        self.column_identical('test.xlsx','test_generated.xlsx',0,0, 13, 13) 
        sheet.column_transform_string_in_binary(14,15,'partie 2 : Vrai',insert=False,label = False)
        self.column_identical('test.xlsx','test_generated.xlsx',0,0, 15, 15) 
        sheet.column_transform_string_in_binary(16,17,'partie 3 : Vrai',insert=False,label = False)
        self.column_identical('test.xlsx','test_generated.xlsx',0,0, 17, 17)
        sheet.column_transform_string_in_binary(41,42,'Laser Interferometer Gravitational-Wave Observatory(LIGO)','virgo','Virgo',insert=False,label = False)
        self.column_identical('test.xlsx','test_generated.xlsx',0,0, 42, 42)

        sheet2 = Sheet('test.xlsx', 'Feuille2')
         
        sheet2.column_transform_string_in_binary(6,7,'partie 12 : Faux',1,label = False)
        self.column_identical('test.xlsx','test.xlsx', 1, 1, 7,8)
        sheet2.sheet.delete_cols(7) #sinon à chaque lancement de test.py il insère une colonne en plus.
        sheet2.writebook.save(sheet2.path + 'test.xlsx') 

    def test_column_set_answer_in_group(self):
        sheet = Sheet('test_column_set_answer.xlsx','sheet1')  
        
        groups_of_response = {"group1":['2','5','6'], "group2":['7','8','9'], "group3":['1','3','4'], "group4":['10']}  

        sheet.column_set_answer_in_group('B','C',groups_of_response)
 
        self.column_identical('test_column_set_answer.xlsx','test_column_set_answer.xlsx',0,1,3,3)
        self.column_identical('test_column_set_answer.xlsx','test_column_set_answer.xlsx',0,1,4,4)

        sheet.sheet.delete_cols(3)
        sheet.updateCellFormulas(sheet.sheet,False,'column',['C'])
        sheet.writebook.save(sheet.path + 'test_column_set_answer.xlsx') 
        

    def test_column_security(self):
        sheet = Sheet('test.xlsx','sheet1')
        
        self.assertEqual(sheet.column_security(1), False)
        self.assertEqual(sheet.column_security(123), True)
    
    """
    def test_color_special_cases_in_sheet(self):
        sheet = Sheet('test.xlsx','Feuille3')
        sheet2 = Sheet('test.xlsx','Feuille4')

        sheet.color_special_cases_in_sheet({"partie 6 : Faux":'0000a933',0:'00ffff00',"partie 7 : Vrai":'00ff0000','Les électrons sont plus petits que les atomes':'002a6099','accuracy_Q6':'00bf0041'})
        for i in range(1,sheet.sheet.max_row+1):
            for j in range(1,sheet2.sheet.max_column+1): 
                print(i,j)
                self.assertEqual(sheet.sheet.cell(i,j).fill.fgColor.rgb,sheet2.sheet.cell(i,j).fill.fgColor.rgb)
    """
    
    def test_add_column_in_sheet_differently_sorted(self):
        sheet1 = Sheet('test.xlsx','Feuille5') 
        sheet1.add_column_in_sheet_differently_sorted('C','E',['test.xlsx','sheet1','C',['B','F']]) 
        self.column_identical('test.xlsx','test.xlsx',4,5,5,5)
        self.column_identical('test.xlsx','test.xlsx',4,5,6,6)
        sheet1.sheet.delete_cols(5,2)

        sheet1.add_column_in_sheet_differently_sorted(3,5,['test.xlsx','sheet1',3,[2,6]],label = False) 
        self.column_identical('test.xlsx','test.xlsx',4,5,5,5)
        self.column_identical('test.xlsx','test.xlsx',4,5,6,6)
        sheet1.sheet.delete_cols(5,2)
        sheet1.writebook.save(sheet1.path + 'test.xlsx')
        
    def test_color_line_containing_chaines(self):
        sheet = Sheet('test.xlsx','color_line')
        sheet.color_lines_containing_chaines('0000a933','-','+')
        
    def test_column_cut_string_in_parts(self):
        sheet = Sheet('test.xlsx','cutinparts')
        sheet.column_cut_string_in_parts('B','C',';') 
        self.column_identical('test.xlsx','test.xlsx',7,8, 3, 3)
        self.column_identical('test.xlsx','test.xlsx',7,8, 4, 4)
        self.column_identical('test.xlsx','test.xlsx',7,8, 5, 5)
        self.column_identical('test.xlsx','test.xlsx',7,8, 6, 6)
        sheet.sheet.delete_cols(3,3)
        sheet.writebook.save(sheet.path + 'test.xlsx') 

    def test_delete_lines(self):
        sheet = Sheet('test.xlsx','delete_lines') 
        sheet.delete_lines_containing_str('D', '0')
        sheet.delete_lines_containing_str('D','p a')
        self.column_identical('test.xlsx','test.xlsx',9,10, 1, 1)
        self.column_identical('test.xlsx','test.xlsx',9,10, 2, 2)
        self.column_identical('test.xlsx','test.xlsx',9,10, 3, 3)
        self.column_identical('test.xlsx','test.xlsx',9,10, 4, 4)
        self.column_identical('test.xlsx','test.xlsx',9,10, 5, 5)
        self.column_identical('test.xlsx','test.xlsx',9,10, 6, 6)

    def test_delete_lines_with_formulas(self):
        sheet = Sheet('listing_par_etape - Copie.xlsx','Feuil1') 
        sheet.delete_lines_containing_str('B', 'pas consenti') 
        self.column_identical('listing_par_etape - Copie.xlsx','listing_par_etape - Copie.xlsx',0, 1, 2, 2)
        self.column_identical('listing_par_etape - Copie.xlsx','listing_par_etape - Copie.xlsx',0, 1, 10, 10) 

    def test_delete_doublons(self): 
        sheet1 = Sheet('test_doublons.xlsx','sheet1')
        sheet2 = Sheet('test_doublons.xlsx','Feuille2')
        sheet1.delete_doublons('C', color = True)
        verify_sheets_identical(sheet1,sheet2)

    def test_create_one_column_by_QCM_answer(self):
        sheet = Sheet('test_create_one_column.xlsx','sheet1')  

        sheet.create_one_column_by_QCM_answer('D','E',['OUI', 'NON'], 'Alain', 'Henri', 'Tony', 'Dulcinée') 
        
        self.column_identical('test_create_one_column.xlsx','test_create_one_column.xlsx',0, 1, 5, 5)
        self.column_identical('test_create_one_column.xlsx','test_create_one_column.xlsx',0, 1, 6, 6) 
        self.column_identical('test_create_one_column.xlsx','test_create_one_column.xlsx',0, 1, 7, 7) 
        self.column_identical('test_create_one_column.xlsx','test_create_one_column.xlsx',0, 1, 8, 8) 

        sheet.sheet.delete_cols(4,4)

    def test_gather_multiple_answers(self):
        sheet = Sheet('testongletbyparticipant.xlsx','test')  
        sheet.gather_multiple_answers('A','B')
        del sheet

        sheet1, sheet2 = Sheet('testongletbyparticipant.xlsx','severalAnswers'),Sheet('testongletbyparticipant-result.xlsx','Feuille2') 
        verify_sheets_identical(sheet1, sheet2)
        del sheet1, sheet2
        
        file = File('testongletbyparticipant.xlsx')
        
        del file.writebook[file.sheets_name[-1]]
        file.writebook.save(file.path + 'testongletbyparticipant.xlsx')
        del file

    def test_give_names_of_maximum(self):
        sheet = Sheet('test_give_names.xlsx','sheet1')
        sheet.give_names_of_maximum('D', 'A', 'B', 'C') 

        verify_sheets_identical(sheet, Sheet('test_give_names.xlsx','Feuille2'))

        sheet.sheet.delete_cols(4)
        sheet.writebook.save(sheet.path + 'test_give_names.xlsx') 
        
    """ def test_delete_other_columns(self):
        # Fonctionnel une fois
        sheet = Sheet('test_keep_only_columns.xlsx','sheet1')
        sheet.delete_other_columns('B-H,L,M,AI-AJ')

        verify_sheets_identical(sheet, Sheet('test_keep_only_columns.xlsx','Feuille2')) """

    def test_column_get_begin_of_str(self):
        sheet = Sheet('test_colgetpartofstr.xlsx','Feuille2')
        sheet.column_get_begin_of_str('C','D','_')
        sheet.column_get_begin_of_str('F','G',';')
        verify_sheets_identical(sheet, Sheet('test_colgetpartofstr.xlsx','expected'))
        sheet.sheet.delete_cols(7)
        sheet.sheet.delete_cols(4)
        sheet.writebook.save(sheet.path + 'test_colgetpartofstr.xlsx') 

    def test_map_two_columns_to_a_third_column(self):
        sheet = Sheet('test_maptwocolumns.xlsx','Feuille2')
        sheet.map_two_columns_to_a_third_column(['B', 'C'], 'D', {'cat1':['prime','1'], 'cat2':['probe','2']})
        verify_sheets_identical(sheet, Sheet('test_maptwocolumns.xlsx','expected'))
        sheet.sheet.delete_cols(4)
        sheet.writebook.save(sheet.path + 'test_maptwocolumns.xlsx')        


class TestStr(TestCase, Other):
    def test_transform_string_in_binary(self):
        chaine = Str('prout') 
        
        self.assertEqual(chaine.transform_string_in_binary('prout','rr'),1)
        self.assertEqual(chaine.transform_string_in_binary('rr'),0)
        self.assertEqual(chaine.transform_string_in_binary(''),0)

    def test_set_answer_in_group(self):
        chaine = Str(1)
        chaine2 = Str(9)
        
        groups_of_response = {"group1":['2','5','6'], "group2":['7','8','9'], "group3":['1','3','4'], "group4":['10']}
        reversed_group = self.reverse_dico_for_set_answer_in_group(groups_of_response)
 
        """ 
        groups_of_response = {}
        for elt in ['2','5','6']:
            groups_of_response[elt] = "group1"
        
        for elt in ['7','8','9']:
            groups_of_response[elt] = "group2"

        for elt in ['1','3','4']:
            groups_of_response[elt] = "group3"
        groups_of_response['10'] = "group4" """
        
        self.assertEqual(chaine.set_answer_in_group(reversed_group),"group3")
        self.assertEqual(chaine2.set_answer_in_group(reversed_group), "group2")


    def test_clean_string(self):
        chaine1 = Str('prout').clean_string()
        chaine2 = Str(' prout').clean_string()
        chaine3 = Str('prout ').clean_string()
        chaine4 = Str(' prout ').clean_string()
        chaine5 = Str('prout  ').clean_string()
        chaine6 = Str('  prout').clean_string()
        self.assertEqual(chaine1.chaine,'prout')
        self.assertEqual(chaine2.chaine,'prout') 
        self.assertEqual(chaine3.chaine,'prout') 
        self.assertEqual(chaine4.chaine,'prout') 
        self.assertEqual(chaine5.chaine,'prout') 
        self.assertEqual(chaine6.chaine,'prout') 

    def test_cut_string_in_parts(self):
        chaine = Str("partie 1 : Vrai; partie 2 : Faux; partie 3 : Vrai; partie 4 : Vrai; partie 5 : Vrai")
        tuple_of_str = chaine.cut_string_in_parts(";")
        
        self.assertEqual(tuple_of_str,("partie 1 : Vrai"," partie 2 : Faux"," partie 3 : Vrai"," partie 4 : Vrai"," partie 5 : Vrai"))

    def test_convert_time_in_minutes(self):
        duration1 = Str("2 jours 2 heures")
        duration2 = Str("1 heure 25 min")
        duration3 = Str("16 min 35 s")
        
        self.assertEqual(duration1.convert_time_in_minutes(), '3000,0')
        self.assertEqual(duration2.convert_time_in_minutes(), '85,0')
        self.assertEqual(duration3.convert_time_in_minutes(), '16,58')

    def test_columns_from_string(self):
        self.assertListEqual(Str.columns_from_strings("C-E,H,J-L"), ['C','D','E','H','J','K','L'])

    def test_listFromColumnsStrings(self):
        self.assertListEqual(Str.listFromColumnsStrings("C-E,H,J-L", "D,G","H-K"),[['C','D','E','H','J','K','L'],['D','G'],['H','I','J','K']])

    def test_range_Letter(self):
        self.assertListEqual(Str.rangeLetter('D-H'), ['D','E','F','G','H'])

    def testUpdateOneFormulaForOneInsertion(self):
        formula = Str.updateOneFormulaForOneInsertion("SI(J10+K$1+L$3)",True,'row','2')
        self.assertEqual(formula, "SI(J11+K$1+L$4)")

        formula = Str.updateOneFormulaForOneInsertion("SI(J12+K$1+L$3)",False,'row','11')
        self.assertEqual(formula, "SI(J11+K$1+L$3)")

        formula = Str.updateOneFormulaForOneInsertion("SI(J12+K$1+L$3)",False,'row','13')
        self.assertEqual(formula, "SI(J12+K$1+L$3)")

        formula = Str.updateOneFormulaForOneInsertion("SI(J10+K$1+L$3)",True,'column','C')
        self.assertEqual(formula, "SI(K10+L$1+M$3)")

        formula = Str.updateOneFormulaForOneInsertion("SI(J10+K$1+L$3)",False,'column','C')
        self.assertEqual(formula, "SI(I10+J$1+K$3)")

    def testUpdateOneFormula(self):
        formula = Str.updateOneFormula("SI(J10+K$1+L$3)",True,'row',['2','5'])
        self.assertEqual(formula, "SI(J12+K$1+L$4)")
 
        
def verify_files_identical(file1, file2):
    testcase = TestCase()
    testcase.assertEqual(file1.sheets_name,file2.sheets_name)

    for onglet in file1.sheets_name: 
        sheet1 = file1.writebook[onglet]
        sheet2 = file2.writebook[onglet]
        for i in range(1,sheet1.max_row+1):
            for j in range(1,sheet1.max_column+1):
                testcase.assertEqual(sheet1.cell(i,j).value,sheet2.cell(i,j).value)

def verify_sheets_identical(sheet1, sheet2):  
    testcase = TestCase()
    testcase.assertEqual(sheet1.sheet.max_row,sheet2.sheet.max_row)
    testcase.assertEqual(sheet1.sheet.max_column,sheet2.sheet.max_column)

    for i in range(1,sheet1.sheet.max_row+1):
        for j in range(1,sheet1.sheet.max_column+1):
            testcase.assertEqual(sheet1.sheet.cell(i,j).value,sheet2.sheet.cell(i,j).value)

if __name__== "__main__":
    main()