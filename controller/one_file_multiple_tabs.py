from utils.utils import GetIndex, TabsCopy, DisplayRunningInfos
from openpyxl.utils import column_index_from_string  
from model.model_factorise import Cell
from time import time 


class MultipleSameTabController():

    def __init__(self, file_object, tab_controller, file_options=None):
        """
        Attributs: 
            - file_object (object of class File) 
            - file_options (FileOptions object)
            - first_line (optional int)  
            - tabs_copy (TabsCopy object): object to apply copy method from a tab to a new tab
            - display (DisplayRunningInfos object): to display the current state of the run
        """
        self.file_object = file_object
        self.tab_controller = tab_controller 
        self.file_options = file_options  
        self.display = DisplayRunningInfos() 

    def reinitialize_tab_controller(self, tab_name):
        self.tab_controller.tab = self.file_object.get_tab_by_name(tab_name)
        self.tab_controller.reinitialize_storing_attributes()

    def apply_method_on_some_tabs(self, method_name, *args, **kwargs):
        """ 
        Vous avez un fichier contenant plusieurs onglets et vous souhaitez appliquer une même méthode de la 
        classe Sheet sur une liste de ces onglets du fichier. On s'attend à ce que tous les onglets aient une structure identique.

        Inputs:
            - method_name (str): the name of the method to execute 
            - *args, **kwargs : arguments of the method associated with method_name
        """  
        self.display.start_time = time()
        for tab_name in self.file_options.names_of_tabs_to_modify:    
            # Get the method from its name and apply it  
            self.reinitialize_tab_controller(tab_name)
            method = getattr(self.tab_controller, method_name)
            method(*args, **kwargs) 

            self._update_display_infos(method_name, tab_name, self.file_options.names_of_tabs_to_modify) 
            self.display.display_running_infos() 

        self.file_object.save_file() 
    
    def _update_display_infos(self, method_name, current_running_part, list_of_running_parts):
        self.display.method_name = method_name
        self.display.current_running_part = current_running_part
        self.display.list_of_running_parts = list_of_running_parts


class OneFileMultipleTabsController(GetIndex):
    """
    Handle methods involving multiple tabs of a file.
    """
    def __init__(self, file_object, file_options=None, first_line=2):
        """
        Attributs: 
            - file_object (object of class File) 
            - file_options (FileOptions object)
            - first_line (optional int)  
            - tabs_copy (TabsCopy object): object to apply copy method from a tab to a new tab
            - display (DisplayRunningInfos object): to display the current state of the run
        """
        self.file_object = file_object
        self.file_options = file_options  
        self.first_line = first_line 
        self.tabs_copy = TabsCopy()
        self.display = DisplayRunningInfos() 

    def _update_display_infos(self, method_name, current_running_part, list_of_running_parts):
        self.display.method_name = method_name
        self.display.current_running_part = current_running_part
        self.display.list_of_running_parts = list_of_running_parts

    def extract_a_column_from_all_tabs(self):
        """
        Fonction qui récupère une colonne dans chaque onglet pour former une nouvelle feuille
        contenant toutes les colonnes. La première cellule de chaque colonne correspond alors 
        au nom de l'onglet. Attention, en l'état, il faut que tous les onglets aient la même structure.
        """ 
        self.file_options.column_to_read = column_index_from_string(self.file_options.column_to_read)
        new_tab = self.file_object.create_and_return_new_tab(f"gather_{self.file_options.column_to_read}")
        self.tabs_copy._choose_the_tab_to_write_in(new_tab)
        self.file_options.column_to_write = 1

        self.display.start_time = time() 
        for tab_name in self.file_object.sheets_name: 
            self.tabs_copy._choose_the_tab_to_read(self.file_object.get_tab_by_name(tab_name))
            self._copy_column_from_a_tab_in_the_next_new_tab_column(tab_name)

            self._update_display_infos('extract_column_from_all_sheets', tab_name, self.file_object.sheets_name)
            self.display.display_running_infos()

        self.file_object.save_file()  
        self.file_object.update_sheet_names()
    
    def _copy_column_from_a_tab_in_the_next_new_tab_column(self, tab_name): 
        self.tabs_copy.copy_paste_column(self.file_options.column_to_read, self.file_options.column_to_write) 
        self._choose_the_new_column_title(tab_name)  
        self.file_options.column_to_write = self.tabs_copy.tab_to.max_column + 1

    def _choose_the_new_column_title(self, title):
        self.tabs_copy.tab_to.cell(1, self.tabs_copy.tab_to.max_column).value = title    

    def apply_columns_formula_on_all_tabs(self):
        """
        Fonction qui reproduit les formules d'une colonne ou plusieurs colonnes
          du premier onglet sur toutes les colonnes situées à la même position dans les 
          autres onglets.

        Input : 
            -column_list : int. les positions des colonnes où récupérer et coller
        """
        columns_int_list = self.get_list_of_columns_indexes(self.file_options.columns_to_read)

        self.display.start_time = time()
        self.tabs_copy._choose_the_tab_to_read(self.file_object.get_tab_by_name(self.file_object.sheets_name[0]))

        # on applique les copies des formules dans tous les onglets sauf le premier duquel viennent ces formules
        for tab_name in self.file_object.sheets_name[1:]:
            self.tabs_copy._choose_the_tab_to_write_in(self.file_object.get_tab_by_name(tab_name))
            self.tabs_copy.copy_paste_multiple_columns(columns_int_list) 

            self._update_display_infos('apply_column_formula_on_all_sheets', tab_name, self.file_object.sheets_name[1:])
            self.display.display_running_infos()
            
        self.file_object.save_file() 

    def apply_cells_formula_on_all_tabs(self, *cells):
        """
        Fonction qui reproduit les formules d'une cellule ou plusieurs cellules
          du premier onglet sur toutes les cellules situées à la même position dans les 
          autres onglets.

        Input : 
            -cells : string. les positions des cellule où récupérer et coller 
        """
        cells_list = GetIndex.get_list_of_cells_coordinates(cells) 

        self.display.start_time = time()
        tab_to_read = self.file_object.get_tab_by_name(self.file_object.sheets_name[0])
        self.tabs_copy._choose_the_tab_to_read(tab_to_read)

        # on applique les copies de formules dans tous les onglets sauf le premier duquel proviennent les formules
        for tab_name in self.file_object.sheets_name[1:]: 
            self.tabs_copy._choose_the_tab_to_write_in(self.file_object.get_tab_by_name(tab_name))  
            self.tabs_copy.deep_copy_multiple_cells(cells_list)  
            
            self._update_display_infos('apply_cells_formula_on_all_sheets', tab_name, self.file_object.sheets_name[1:])
            self.display.display_running_infos()

        self.file_object.save_file() 

    def gather_groups_of_multiple_columns_in_tabs_of_two_columns_containing_tags_and_values(self, *lists_of_columns):
        """
        Vous avez des groupes de colonnes de valeurs avec une étiquette en première cellule. Pour chaque groupe, vous souhaitez former deux colonnes de valeurs : l'une qui contient
        les valeurs rassemblées en une colonne, l'autre, à sa gauche, qui indique l'étiquette de la colonne dans laquelle elle a été prise.

        Inputs : 
            - onglet (str) : nom de l'onglet d'où on importe les colonnes.
            - column_lists (list[list[str]]) : liste de groupes de colonnes. Chaque groupe est une liste de colonnes.
        """ 
        tab_to_read = self.file_object.get_tab_by_name(self.file_options.name_of_tab_to_read)
        self.tabs_copy._choose_the_tab_to_read(tab_to_read)

        for list_of_columns in lists_of_columns: 
            self.file_options.columns_to_read = list_of_columns
            tab_to = self._create_a_tab_for_a_list_of_columns()
            self.tabs_copy._choose_the_tab_to_write_in(tab_to) 
            self.copy_tags_and_values_of_a_list_of_columns()

        self.file_object.save_file() 
    
    def _create_a_tab_for_a_list_of_columns(self):
        string_of_columns = ''.join(self.file_options.columns_to_read)
        tab_name = f"tab_column_gathered_{string_of_columns}"
        return self.file_object.writebook.create_sheet(tab_name) 

    def copy_tags_and_values_of_a_list_of_columns(self):
        for column in self.file_options.columns_to_read: 
            self.tabs_copy.copy_tag_and_values_of_a_column_at_tab_bottom(column_index_from_string(column))         

    def merge_cells_on_all_tabs(self, merged_cells_range):
        """
        Fonction qui merge les mêmes cellules sur tous les onglets d'un fichier 
        """
        
        merged_cells_range.start_column = column_index_from_string(merged_cells_range.start_column)
        merged_cells_range.end_column = column_index_from_string(merged_cells_range.end_column)

        self.display.start_time = time()

        for tab_name in self.file_object.sheets_name: 
            self.tabs_copy._choose_the_tab_to_write_in(self.file_object.get_tab_by_name(tab_name)) 
            self.tabs_copy.tab_to.merge_cells(start_row=merged_cells_range.start_line, 
                                              start_column=merged_cells_range.start_column, 
                                              end_row=merged_cells_range.end_line, 
                                              end_column=merged_cells_range.end_column)
            self._update_display_infos('merge_cells_on_all_tabs', tab_name, self.file_object.sheets_name)
            self.display.display_running_infos()

        self.file_object.save_file() 

    def list_tabs_with_different_number_of_lines(self, number_of_lines):
        """
        Fonction qui prend un fichier et qui contrôle si tous les onglets ont un nombre de lignes égal à l'argument
        """
        list_of_tab_names = []
        for tab_name in self.file_object.sheets_name:
            list_of_tab_names = self.add_tab_to_list_if_different_number_of_lines(tab_name, list_of_tab_names, number_of_lines)
        return list_of_tab_names
    
    def add_tab_to_list_if_different_number_of_lines(self, tab_name, list_of_tab_names, number_of_lines):
        tab = self.file_object.get_tab_by_name(tab_name)
        if tab.max_row != number_of_lines:
            list_of_tab_names.append(tab_name)
        return list_of_tab_names
    
    def gather_multiple_answers(self, column_identifier, column_data):
        """
        Dans un onglet, nous avons les réponses de participants qui ont pu répondre plusieurs fois à un questionnaire.
        Cette fonction parcourt les noms et écrit dans un autre onglet ceux qui ont répondu plusieurs fois.
        La ligne du participant est alors constituée des différentes valeurs d'une même donnée récupérée.
        
        Inputs :
            - column_read (str) : la colonne avec les identifiants des participants.
            - column_store (str) : lettre de la colonne contenant la donnée qu'on veut stocker.
            - line_beggining (int) : ligne où débute la recherche. 
        """ 
        tab_to_read = self.file_object.get_tab_by_name(self.file_options.name_of_tab_to_read)
        self.tabs_copy._choose_the_tab_to_read(tab_to_read)
        column_identifier = column_index_from_string(column_identifier) 
        column_data = column_index_from_string(column_data) 

        map_participants_to_data = self._store_multiple_answers_participants_data(column_identifier, column_data)
        self._create_tab_storing_multiple_answers(map_participants_to_data)
        self.file_object.save_file()

    def _store_multiple_answers_participants_data(self, column_identifier, column_data):
        """
        Fonction qui crée un dictionnaire de clés les identifiants et de valeur la ou les valeurs
        d'une même donnée associée à chaque identifiant. 
        """
        map_participant_to_data = {}
        for line_index in range(self.first_line, self.tabs_copy.tab_from.max_row + 1):
            self._store_one_participant_data(map_participant_to_data,
                                                           Cell(line_index, column_identifier), 
                                                           Cell(line_index, column_data))
        return map_participant_to_data
    
    def _store_one_participant_data(self, map_participant_to_data, cell_identifier, cell_data):
        tab_from = self.tabs_copy.tab_from
        identifier = self.file_object.get_compiled_cell_value(tab_from, cell_identifier)
        value_to_store = self.file_object.get_compiled_cell_value(tab_from, cell_data)
        
        if identifier in map_participant_to_data.keys(): 
            map_participant_to_data[identifier].append(value_to_store)
        else: 
            map_participant_to_data[identifier] = [value_to_store]
    
    def _create_tab_storing_multiple_answers(self, map_participant_to_data):
        new_tab = self.file_object.writebook.create_sheet('multiple_answers')
        new_tab.cell(1, 1).value = 'Identifiers'
        
        for participant_item in map_participant_to_data.items(): 
            values_to_store = participant_item[1]
            if len(values_to_store) >= 2:
                self._write_data_of_a_multiple_answers_participant(new_tab, participant_item) 
    
    def _write_data_of_a_multiple_answers_participant(self, new_tab, participant_item):
        identifier = participant_item[0]
        values_to_store = participant_item[1] 
        new_tab.cell(new_tab.max_row + 1, 1).value = identifier
        for index in range(len(values_to_store)):
            new_tab.cell(new_tab.max_row + 1, index + 2).value = values_to_store[index]
            
