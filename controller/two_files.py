import openpyxl
import os

from utils.utils import GetIndex, TabsCopy, DisplayRunningInfos, TabUpdate, ColumnInsert, Str
from model.model_factorise import Cell 
from controller.one_file_one_tab import create_empty_workbook
from openpyxl.utils import column_index_from_string
from datetime import datetime 
from time import time


class TwoFilesController(GetIndex):
    """Handle methods linking two existing files"""
    def __init__(self, file_object_from, file_object_to, tab_name_from, tab_name_to,
                 column_with_identifiers_from=None, column_with_identifiers_to=None, first_line=2):
        self.file_object_from = file_object_from
        self.file_object_to = file_object_to
        self.first_line = first_line
        self.tabs_copy = TabsCopy(file_object_from.get_tab_by_name(tab_name_from),
                                  file_object_to.get_tab_by_name(tab_name_to))
        self.tab_update = TabUpdate()
        self._get_columns_identifiers_indexes(column_with_identifiers_from, column_with_identifiers_to)

    def _get_columns_identifiers_indexes(self, column_with_identifiers_from, column_with_identifiers_to):
        try:
            self.column_with_identifiers_to = column_index_from_string(column_with_identifiers_to) 
            self.column_with_identifiers_from = column_index_from_string(column_with_identifiers_from)
        except TypeError:
            pass
        
    def copy_columns_in_a_tab_differently_sorted(self, columns_to_copy, column_insertion):
        """
        Fonction qui insère dans un onglet des colonnes d'un autre onglet de référence. 
        Les deux feuilles ont chacun une colonne d'identifiants (exemple : des mails) avec des identifiants communs.
        Ces identifiants sont a priori triés dans des ordres différents.
        La fonction récupère un ou plusieurs éléments d'une ligne déterminée par un identifiant.
        Elle recherche l'identifiant dans la seconde feuille et insère les éléments dans la ligne correspondante.
        """ 
        column_insertion_index = column_index_from_string(column_insertion) 
        columns_to_copy_indexes = self.get_list_of_columns_indexes(columns_to_copy) 
        cells_to_copy_by_identifier = self._create_a_dictionary_with_identifiers_and_indexes_of_cells_to_copy(columns_to_copy_indexes)
        self.tabs_copy.tab_to.insert_cols(column_insertion_index, len(columns_to_copy)) 
        self._copy_cells_values_in_the_new_tab(cells_to_copy_by_identifier, column_insertion_index)

        modification_object = ColumnInsert(self.get_list_of_consecutive_column_letters(column_insertion_index, len(columns_to_copy))) 
        self.update_cell_formulas(modification_object)   
        self.file_object_to.save_file()

    def update_cell_formulas(self, modification_object): 
        self.tab_update.choose_modifications_to_apply(modification_object) 
        self.tab_update.update_cells_formulas(self.tabs_copy.tab_to) 

    def _create_a_dictionary_with_identifiers_and_indexes_of_cells_to_copy(self, columns_to_copy_indexes):
        tab_from = self.tabs_copy.tab_from
        dico = {}
        for line_index in range(1, tab_from.max_row + 1):
            identifier = self.file_object_from.get_compiled_cell_value(tab_from, Cell(line_index, self.column_with_identifiers_from))  
            dico[identifier] = GetIndex.get_cells_indexes_of_one_line_and_some_columns(line_index, columns_to_copy_indexes)
        return dico

    def _copy_cells_values_in_the_new_tab(self, cells_to_copy_indexes_by_identifier, column_insertion_index):
        tab_to = self.tabs_copy.tab_to
        for line_index in range(1, tab_to.max_row + 1):

            # The identifiers of tab_to may not be present in tab_from so that it may not be a key of the dictionary created from tab_from
            try:
                identifier = self.file_object_to.get_compiled_cell_value(tab_to, Cell(line_index, self.column_with_identifiers_to))
                cells_to_copy_indexes = cells_to_copy_indexes_by_identifier[identifier]
                self._copy_cells_values_for_a_line_in_the_new_tab(line_index, cells_to_copy_indexes, column_insertion_index)
            except KeyError:
                continue
        
    def _copy_cells_values_for_a_line_in_the_new_tab(self, line_index, cells_to_copy_indexes, column_insertion_index): 
        for column_index in range(len(cells_to_copy_indexes)):
            self.tabs_copy.deep_copy_of_a_cell(Cell(*cells_to_copy_indexes[column_index]), Cell(line_index, column_insertion_index + column_index))
 

class OneFileCreatedController(GetIndex):
    """Handle methods creating a new file from an existing file"""

    def __init__(self, file_object, file_options=None, first_line=2):
        """
        Attributs: 
            - file_object (object of class File)
            - file_options (FileOptions object)
            - first_line (optional int)
            - current_line (int): line likely to evolve in methods
            - new_writebook (openpyxl.WorkBook) : eventual workbook to complete
            - tabs_copy (TabsCopy object): object to apply copy method from a tab to a new tab
            - display (DisplayRunningInfos object): to display the current state of the run
        """
        self.file_object = file_object
        self.file_options = file_options 
        self.first_line = first_line
        self.current_line = 2
        self.new_writebook = None 
        self.tabs_copy = TabsCopy()
        self.display = DisplayRunningInfos()

    def make_horodated_copy_of_a_file(self):
        self.new_writebook = create_empty_workbook()
        self._copy_tabs_in_new_workbook()
        self._save_horodated_file()            
                     
    def _copy_tabs_in_new_workbook(self): 
        self.display.start_time = time() 

        for tab_name in self.file_object.sheets_name:            
            self.new_writebook.create_sheet(tab_name)
            self._update_old_file_tab_and_new_file_tab(tab_name) 
            self.tabs_copy.copy_old_file_tab_in_new_file_tab()

            self._update_display_infos('make_horodated_copy_of_a_file', tab_name, self.file_object.sheets_name)
            self.display.display_running_infos() 

    def _update_old_file_tab_and_new_file_tab(self, tab_name): 
        self.tabs_copy._choose_the_tab_to_read(self.file_object.get_tab_by_name(tab_name))
        self.tabs_copy._choose_the_tab_to_write_in(self.new_writebook[tab_name])

    def _save_horodated_file(self):
        name_file_without_extension = Str(self.file_object.name_file).del_extension() 
        file_to_save_name = self.file_object.path  + name_file_without_extension + '_date_' + datetime.now().strftime("%Y-%m-%d_%Hh%M") + '.xlsx'
        self.new_writebook.save(file_to_save_name) 

    def _update_display_infos(self, method_name, current_running_part, list_of_running_parts):
        self.display.method_name = method_name
        self.display.current_running_part = current_running_part
        self.display.list_of_running_parts = list_of_running_parts

    def split_one_tab_in_multiple_tabs(self):
        """
        Fonction qui prend un onglet dont une colonne contient des chaînes de caractères comme par exemple un nom.
        Chaque chaîne de caractères peut apparaître plusieurs fois dans cette colonne (exe : quand un participant répond plusieurs fois)
        La fonction retourne un fichier contenant un onglet par chaîne de caractères.
        Chaque onglet contient toutes les lignes correspondant à cette chaîne de caractères.
        """ 
        
        new_file_name = f'divided_{self.file_object.name_file}'
        self._create_or_load_workbook(new_file_name)
        
        tab_to_read_name = self.file_options.name_of_tab_to_read
        self.tabs_copy._choose_the_tab_to_read(self.file_object.get_tab_by_name(tab_to_read_name)) 

        last_line = self.tabs_copy.tab_from.max_row + 1
        for line_index in range(self.first_line, last_line):
            self.current_line = line_index
            self._create_or_complete_a_tab_by_identifier()

        # The first tab automatically created when opening the new workbook is useless
        self._delete_first_tab_of_new_workbook(new_file_name)
        self.new_writebook.save(self.file_object.path + new_file_name)

    def _create_or_load_workbook(self, new_file_name): 
        try:
            self.new_writebook = openpyxl.load_workbook(self.file_object.path + new_file_name)
        except OSError:
            self.new_writebook = openpyxl.Workbook()
    
    def _create_or_complete_a_tab_by_identifier(self): 
        tab_names = self.new_writebook.sheetnames 
        column_to_read_by_index = column_index_from_string(self.file_options.column_to_read)
        identifier = self.file_object.get_compiled_cell_value(self.tabs_copy.tab_from, Cell(self.current_line, column_to_read_by_index))

        if identifier not in tab_names:
            self._create_tab_called_identifier_and_fill_first_line(identifier)
            tab_names.append(identifier) 
 
        self.tabs_copy._choose_the_tab_to_write_in(self.new_writebook[identifier])
        self.tabs_copy.add_line_at_bottom(self.current_line)

    def _create_tab_called_identifier_and_fill_first_line(self, identifier):
        self.new_writebook.create_sheet(identifier)
        self.tabs_copy._choose_the_tab_to_write_in(self.new_writebook[identifier])
        self.tabs_copy.copy_paste_line(1, 1)
         
    def _delete_first_tab_of_new_workbook(self, new_file_name):
        if new_file_name not in os.listdir(self.file_object.path):
            first_tab_name = self.new_writebook.sheetnames[0]
            del self.new_writebook[first_tab_name] 

    def extract_cells_from_all_tabs(self, *cells):
        """
        Vous avez un fichier avec des onglets de structure identique avec un onglet par participant. Vous souhaitez
        récupérer des cellules identiques dans tous les onglets et créer un fichier à un onglet avec une ligne par participant,
        qui contient les valeurs de ces cellules. Fonction analogue à gather_multiple_answers mais ne portant pas sur une
        seule feuille.

        Inputs:
            - cells (list[str])
        """  
        cells_list = self.get_list_of_cells_coordinates(cells) 
        
        self._create_workbook_and_choose_first_tab_to_write_in()
 
        self.display.start_time = time()
  
        for tab_name in self.file_object.sheets_name: 
            self.tabs_copy._choose_the_tab_to_read(self.file_object.get_tab_by_name(tab_name))   
            self._fill_the_line_corresponding_to_one_tab(cells_list, tab_name)
            
            self._update_display_infos('extract_cells_from_all_sheets', tab_name, self.file_object.sheets_name)
            self.display.display_running_infos()
 
        self.new_writebook.save(self.file_object.path + 'gathered_data_' + self.file_object.name_file) 

    def _create_workbook_and_choose_first_tab_to_write_in(self):
        self.new_writebook = openpyxl.Workbook()
        first_tab_name = self.new_writebook.sheetnames[0]
        return self.tabs_copy._choose_the_tab_to_write_in(self.new_writebook[first_tab_name])
    
    def _fill_the_line_corresponding_to_one_tab(self, cells_list, tab_name):
        self._fill_the_first_cell_of_the_line_with_tab_name(tab_name)
        self._fill_the_line_with_extracted_cells(cells_list)
        
    def _fill_the_first_cell_of_the_line_with_tab_name(self, tab_name):
        self.tabs_copy.tab_to.cell(self.current_line, 1).value = tab_name

    def _fill_the_line_with_extracted_cells(self, cells_list):
        current_column = 2 
        for cell in cells_list:   
            self.tabs_copy.copy_of_a_cell(Cell(cell[0],cell[1]), Cell(self.current_line, current_column)) 
            current_column += 1
        self.current_line += 1
            
    def create_one_file_by_tab(self):
        """
        Vous souhaitez fabriquer un fichier par onglet. Chaque fichier aura le nom de l'onglet. 
        """
        self.display.start_time = time()

        for tab_name in self.file_object.sheets_name: 
            self._create_a_file_from_a_tab(tab_name)
            self._update_display_infos('one_file_by_tab_sendmail', tab_name, self.file_object.sheets_name)
            self.display.display_running_infos()

    def _create_a_file_from_a_tab(self, tab_name):
        """
        Fonction qui prend un nom d'onglet dans un fichier et qui crée un fichier associé. 
        """
        self.tabs_copy._choose_the_tab_to_read(self.file_object.get_tab_by_name(tab_name))
        self._create_workbook_and_choose_first_tab_to_write_in() 
        self.tabs_copy.deep_copy_of_a_tab() 
        self.new_writebook.save('multifiles/' + tab_name + '.xlsx') 