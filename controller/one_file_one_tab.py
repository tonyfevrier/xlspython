"""Handle methods reading and modifying a unique tab of a file."""

from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter 
from utils.utils_factorise import String, MapIndexLetter, TabUpdateFormula, ColumnDelete, ColumnInsert, LineDelete, RegularExpression, Dictionary
from model.model_factorise import Cell 


class ColorTabController(String):
    """Handle methods coloring a unique tab of a file."""

    def __init__(self, file_object=None, tab_name=None, tab_options=None, color=None, first_line=2, save=False):
        """
        Attributs:
            - file_object (file object)
            - tab_name (str)
            - tab (openpyxl.workbook.tab)            
            - tab_options (TabOptions object)
            - first_line (optional int)
        """ 
        self.file_object = file_object 
        if file_object is not None and tab_name is not None:
            self.tab = self.file_object.get_tab_by_name(tab_name)
        self.tab_options = tab_options 
        self.first_line = first_line 
        self.color = color
        self.save = save

    def save_file(fonction):
        def wrapper(self, *args, **kwargs):
            fonction(self, *args, **kwargs)
            if self.save:
                self.file_object.save_file()
        return wrapper

    def reinitialize_storing_attributes(self):
        pass

    @save_file
    def color_cases_in_column(self, map_string_to_color):
        """
        Fonction qui pour une colonne donnée colore les cases correspondant à certaines chaînes de caractères.
        """  
        column_index = column_index_from_string(self.tab_options.column_to_read) 
        
        for i in range(self.first_line, self.tab.max_row + 1):
            cell = self.tab.cell(i, column_index) 
            self._color_cell_if_contains_string(cell, map_string_to_color) 
        

    def _color_cell_if_contains_string(self, cell, map_string_to_color):
        cleaned_cell_value = self._get_cleaned_cell_value(cell)
        if cleaned_cell_value in map_string_to_color.keys():
            cell.fill = PatternFill(fill_type = 'solid', start_color = map_string_to_color[cleaned_cell_value])

    def _get_cleaned_cell_value(self, cell): 
        if cell.value is str:
            cleaned_cell_value = self.clean_string_from_spaces(cell.value) 
        else: 
            cleaned_cell_value = cell.value
        return cleaned_cell_value

    @save_file
    def color_cases_in_sheet(self, map_string_to_color): 
        """
        Fonction qui colore les cases contenant à certaines chaînes de caractères d'une feuille 
        """  

        for j in range(1, self.tab.max_column + 1):
            self.tab_options.column_to_read = get_column_letter(j) 
            self.color_cases_in_column(map_string_to_color)

    @save_file
    def color_lines_containing_strings(self, *strings):
        """
        Fonction qui colore les lignes dont une des cases contient une str particulière.
        """ 

        lines_indexes = self._list_lines_containing_strings(*strings)
        self._color_lines(lines_indexes) 

    def _list_lines_containing_strings(self, *strings):
        lines_indexes = []
        for line_index in range(self.first_line, self.tab.max_row + 1):
            lines_indexes = self._add_line_containing_strings_to_list(strings, line_index, lines_indexes)
        return lines_indexes
    
    def _color_line(self, line_index): 
        for column_index in range(1, self.tab.max_column + 1):
            self.tab.cell(line_index, column_index).fill = PatternFill(fill_type = 'solid', start_color = self.color)

    def _add_line_containing_strings_to_list(self, strings, line_index, lines_indexes):
        for column_index in range(1, self.tab.max_column + 1):
            if self.file_object.get_compiled_cell_value(self.tab, Cell(line_index, column_index)) in strings:
                lines_indexes.append(line_index)
                break
        return lines_indexes
    
    def _color_lines(self, list_of_lines):
        for line_index in list_of_lines:
            self._color_line(line_index) 
    

class DeleteController(String):
    """Handle methods deleting lines or columns of a tab"""

    def __init__(self, file_object=None, tab_name=None, first_line=2, save=False):
        self.file_object = file_object
        self.tab_name = tab_name
        self.tab = None
        if file_object is not None and tab_name is not None:
            self.tab = self.file_object.get_tab_by_name(tab_name)  
        self.tab_update = TabUpdateFormula()
        self.columns_to_delete = []
        self.lines_to_delete = []
        self.first_line = first_line  
        self.save = save

    def save_file(fonction):
        def wrapper(self, *args, **kwargs):
            fonction(self, *args, **kwargs)
            if self.save:
                self.file_object.save_file()
        return wrapper
    
    def reinitialize_storing_attributes(self):
        self.columns_to_delete = []
        self.lines_to_delete = []

    def _update_cell_formulas(self, modification_object): 
        self.tab_update.choose_modifications_to_apply(modification_object)  
        self.tab_update.update_cells_formulas(self.tab) 
    
    @save_file
    def delete_columns(self, string_of_columns):
        """
        Prend une séquence de colonnes sous la forme 'C-J,K,L-N,Z' qu'on souhaite supprimer. 
        """  

        # Réordonner par les lettres les plus grandes pour supprimer de la droite vers la gauche dans l'excel  
        self.columns_to_delete = self.get_columns_from(string_of_columns)
        self.columns_to_delete.sort(reverse = True)  

        for column_letter in self.columns_to_delete:  
            self.tab.delete_cols(column_index_from_string(column_letter)) 

        self._update_cell_formulas(ColumnDelete(self.columns_to_delete)) 
    
    @save_file
    def delete_other_columns(self, string_of_columns):
        """
        Prend une séquence de colonnes sous la forme 'C-J,K,L-N,Z' et supprime les autres 
        """  
        columns_to_keep = self.get_columns_from(string_of_columns)
        list_all_columns = self._get_list_of_columns()

        #Réordonner par les lettres les plus grandes pour supprimer de la droite vers la gauche
        list_all_columns.sort(reverse=True) 

        for column_letter in list_all_columns: 
            self._delete_column_not_to_keep(column_letter, columns_to_keep)
       
        self._update_cell_formulas(ColumnDelete(self.columns_to_delete))
        self.file_object.save_file() 

    def _delete_column_not_to_keep(self, column_letter, columns_to_keep):
        if column_letter not in columns_to_keep:
            self.columns_to_delete.append(column_letter)
            self.tab.delete_cols(column_index_from_string(column_letter)) 

    def _get_list_of_columns(self):
        return [get_column_letter(column_index) for column_index in range(1, self.tab.max_column + 1)]

    @save_file
    def delete_lines_containing_strings_in_given_column(self, column_letter, *strings):
        """
        Fonction qui parcourt une colonne et qui supprime la ligne si celle-ci contient une chaîne particulière.

        """ 

        column_index = column_index_from_string(column_letter)  

        # On part de la plus grande ligne pour éviter qu'une suppression ne change la position d'une ligne à supprimer après
        for line_index in range(self.tab.max_row, 0, -1):
            cell_value = self.file_object.get_compiled_cell_value(self.tab, Cell(line_index,column_index)) 
            self._delete_line_containing_strings(line_index, cell_value, *strings)
 
        self._update_cell_formulas(LineDelete(self.lines_to_delete))   

    def _delete_line_containing_strings(self, line_index, cell_value, *strings):
        if str(cell_value) in strings:  
            self.tab.delete_rows(line_index) 
            self.lines_to_delete.append(str(line_index))     

    @save_file
    def delete_twins_lines_and_color_last_twin(self, column_identifier, color = 'FFFFFF00'):
        """
        Certains participants répondent plusieurs fois à une étude. Cette fonction supprime les premières lignes réponses
        des participants dans ce cas. Elle ne garde que leur dernière réponse. On repère les participants
        par leur identifiant unique donné dans colum_identifiant.
        """

        column_identifier = column_index_from_string(column_identifier) 

        map_identifier_to_line = {}  

        #On parcourt dans le sens inverse afin d'éviter que la suppression progressive impacte la position des lignes étudiées ensuite. 
        for line_index in range(self.tab.max_row, 0, -1):
            cell_identifier = Cell(line_index, column_identifier)
            self._delete_line_and_color_last_twin(cell_identifier, map_identifier_to_line, color)
             
        self._update_cell_formulas(LineDelete(self.lines_to_delete))       

    def _delete_line_and_color_last_twin(self, cell_identifier, map_identifier_to_line, color):
        """
        map_identifier_to_line stocke tous les identifiants et la ligne associée. Si un identifiant I1
        est déjà dans le dictionnaire, on va supprimer sa ligne et donc décaler toutes les lignes en-dessous 
        de -1, il faut donc baisser la ligne de I1 de 1 car si I1 vient une troisième fois, on ne colorerait pas 
        la bonne ligne 
        """
        cell_value = self.file_object.get_compiled_cell_value(self.tab, cell_identifier)
        identifier = self.clean_string_from_spaces(cell_value) 

        if identifier in map_identifier_to_line.keys(): 
            self._color_line(map_identifier_to_line[identifier], color)
            self.tab.delete_rows(cell_identifier.line_index)
            self.lines_to_delete.append(str(cell_identifier.line_index))
            map_identifier_to_line[identifier] -= 1    
        else:
            map_identifier_to_line[identifier] = cell_identifier.line_index 

        return map_identifier_to_line

    def _color_line(self, line_index, color):  
        for column_index in range(1, self.tab.max_column + 1):
            self.tab.cell(line_index, column_index).fill = PatternFill(fill_type = 'solid', start_color = color)  
        

class InsertController(MapIndexLetter, RegularExpression, String):
    """Handle methods inserting columns in a tab""" 

    def __init__(self, file_object=None, tab_name=None, tab_options=None, first_line=2, save=False):
        """
        Attributs:
            - file_object (file object)
            - tab_name (str)
            - tab (openpyxl.workbook.tab)            
            - tab_options (TabOptions object)
            - first_line (optional int)
        """ 
        self.file_object = file_object
        self.tab_name = tab_name
        if file_object is not None and tab_name is not None:
            self.tab = self.file_object.get_tab_by_name(tab_name)
        self.tab_update = TabUpdateFormula()
        self.tab_options = tab_options 
        self.first_line = first_line 
        self.save = save

    def save_file(fonction):
        def wrapper(self, *args, **kwargs):
            fonction(self, *args, **kwargs)
            if self.save:
                self.file_object.save_file()
        return wrapper

    def reinitialize_storing_attributes(self):
        pass

    def _update_cell_formulas(self, modification_object): 
        self.tab_update.choose_modifications_to_apply(modification_object)  
        self.tab_update.update_cells_formulas(self.tab) 
 
    @save_file
    def insert_splitted_strings_of(self, column_to_split, separator):
        """
        Fonction qui prend une colonne dont chaque cellule contient une grande chaîne de
          caractères. Toutes les chaînes sont composés du nombre de morceaux délimités par un séparateur,
        La fonction insère autant de colonnes que de morceaux et place un morceau par colonne dans l'ordre des morceaux.
        """ 
        column_to_split = column_index_from_string(column_to_split) 
        self.tab_options.column_to_write = column_index_from_string(self.tab_options.column_to_write)
        
        for line_index in range(self.first_line, self.tab.max_row + 1): 
            parts = self._get_string_and_split_it(Cell(line_index, column_to_split), separator)  
            self._insert_splitted_string(line_index, parts)

        modifications = [get_column_letter(column_to_split + index) for index in range(len(parts))]
        self._update_cell_formulas(ColumnInsert(modifications))   

    def _get_string_and_split_it(self, cell, separator):
        cell_value = self.file_object.get_compiled_cell_value(self.tab, cell) 
        return cell_value.split(separator)

    def _insert_splitted_string(self, line_index, parts):
        if line_index == self.first_line:
            self.tab.insert_cols(self.tab_options.column_to_write, len(parts))

        for part_index in range(len(parts)):
            self.tab.cell(line_index, self.tab_options.column_to_write + part_index).value = parts[part_index]
    
    @save_file
    def fill_one_column_by_QCM_answer(self, *answers):
        """
        Fonction qui recoit des réponses et crée une colonne par réponse. Elle regarde ensuite dans une cellule si la réponse 
        y est contenue. Si oui elle l'indique dans la colonne correspondante.
        """ 
        self.tab_options.column_to_read = column_index_from_string(self.tab_options.column_to_read) 
        self.tab_options.column_to_write = column_index_from_string(self.tab_options.column_to_write) 
        self._insert_answers_columns(answers)

        for line_index in range(self.first_line, self.tab.max_row + 1):
            self._check_columns_of_answers_contained(Cell(line_index, self.tab_options.column_to_read), answers)    

        modifications = [get_column_letter(self.tab_options.column_to_write + index) for index in range(len(answers))]
        self._update_cell_formulas(ColumnInsert(modifications))  

    def _insert_answers_columns(self, answers):
        self.tab.insert_cols(self.tab_options.column_to_write, len(answers))
        for index in range(0, len(answers)):
            self.tab.cell(1, index + self.tab_options.column_to_write).value = answers[index]

    def _check_columns_of_answers_contained(self, cell, answers):
        try:
            for index in range(0, len(answers)):  
                self._check_column_if_answer_contained(cell, (index, answers[index]))
        except TypeError:
            pass

    def _check_column_if_answer_contained(self, cell, answer):
        if answer[1] in self.tab.cell(cell.line_index, cell.column_index).value:
            self.tab.cell(cell.line_index, answer[0] + self.tab_options.column_to_write).value = 'X'
    
    def act_on_columns(function):
        """
        Décorateur qui en plus d'appliquer la fonction, transforme les lettres de colonnes en index, met à jour les formules 
        après l'insertion de la colonne et sauvegarde le fichier.
        """
        def wrapper(self, *args, **kwargs):
            self._get_indexes_of_columns_to_read()
            column_insertion = column_index_from_string(self.tab_options.column_to_write) 
            self.tab.insert_cols(column_insertion)
            function(self, *args, **kwargs)

            modifications = self.tab_options.column_to_write
            self._update_cell_formulas(ColumnInsert(modifications)) 
        return wrapper

    def _get_indexes_of_columns_to_read(self): 
        if self.tab_options.columns_to_read is not None:
            self.tab_options.columns_to_read = self.get_list_of_columns_indexes(self.tab_options.columns_to_read)
        if self.tab_options.column_to_read is not None:
            self.tab_options.column_to_read = column_index_from_string(self.tab_options.column_to_read)
    
    @save_file
    @act_on_columns
    def map_two_columns_to_a_third_column(self, mapping):
        """
        Vous avez deux colonnes de lecture, si les valeurs des deux cellules correspondent à une clé du dictionnaire mapping,
        on écrit la valeur associée dans une nouvelle colonne.
        """ 
        for line_index in range(self.first_line, self.tab.max_row + 1):
            self._fill_one_third_column_cell(line_index, mapping)

    def _fill_one_third_column_cell(self, line_index, mapping):
        cells_values = self._get_cells_values(line_index)
        column_to_write = column_index_from_string(self.tab_options.column_to_write)

        for key, value in mapping.items(): 
            if cells_values == value:
                self.tab.cell(line_index, column_to_write).value = key
                break

    def _get_cells_values(self, line_index):
        columns_to_read = self.tab_options.columns_to_read
        cell_value1 = self.file_object.get_compiled_cell_value(self.tab, Cell(line_index, columns_to_read[0])) 
        cell_value2 = self.file_object.get_compiled_cell_value(self.tab, Cell(line_index, columns_to_read[1]))
        return [cell_value1, cell_value2]

    @save_file
    @act_on_columns
    def write_piece_of_string_in_column(self, separator, piece_index):
        """
        Vous avez une colonne qui contient une chaîne que vous couper en morceaux et en sélectionner un. 
        """  
        column_to_read = self.tab_options.column_to_read
        column_to_write = column_index_from_string(self.tab_options.column_to_write)

        for line_index in range(self.first_line, self.tab.max_row + 1): 
            self._write_piece_of_string_in_cell(Cell(line_index, column_to_read), Cell(line_index, column_to_write),
                                             separator, piece_index)

    def _write_piece_of_string_in_cell(self, cell_read, cell_write, separator, piece_index):
        cell_value = self.file_object.get_compiled_cell_value(self.tab, cell_read)
        if cell_value is not None: 
            piece_of_string = cell_value.split(separator)[piece_index]
            self.tab.cell(cell_write.line_index, cell_write.column_index).value = piece_of_string  
             
    @save_file
    @act_on_columns
    def insert_column_for_prime_probe_congruence(self):
        """
        Vous avez trois colonnes l'une contient des chaines de caractères particulières qui sont prime, probe, croix de fixation ...
          Les deux autres contiennent des chaines de la forme MOTnb_.jpg où MOT peut 
        être congruent, neutre, incongruent et nb est un nombre. Vous souhaitez insérer une colonne contenant soit rien, soit prime
        suivi du MOT de la deuxième colonne si la chaîne de la première colonne est prime, soit probe suivi du MOT de la troisième 
        colonne si la chaîne de la première colonne est probe. 
        """  

        mapping = self._map_prime_probe_to_column()

        for line_index in range(self.first_line, self.tab.max_row + 1):
            self._insert_cell_for_prime_probe_congruence(line_index, mapping)

    def _map_prime_probe_to_column(self):
        """Permet de choisir quelle colonne choisir suivant le mot présent dans la première colonne"""
        return {'prime': self.tab_options.columns_to_read[1],
                'Prime': self.tab_options.columns_to_read[1],
                'probe': self.tab_options.columns_to_read[2], 
                'Probe': self.tab_options.columns_to_read[2]}
    
    def _insert_cell_for_prime_probe_congruence(self, line_index, mapping):
        try:
            first_cell_value = self.file_object.get_compiled_cell_value(self.tab, Cell(line_index, self.tab_options.columns_to_read[0]))
            second_cell_value = self.file_object.get_compiled_cell_value(self.tab, Cell(line_index, mapping[first_cell_value]))
            word = self.get_word_jpg_name_file(second_cell_value) 
            column_to_write = column_index_from_string(self.tab_options.column_to_write)
            self.tab.cell(line_index, column_to_write).value = first_cell_value + "_" + word 
        except KeyError:
            pass

    @save_file
    @act_on_columns
    def insert_tags_of_maximum_of_column_list(self):
        """
        Vous avez une liste de colonnes avec des chiffres, chaque colonne a un nom dans sa première cellule. 
        Cette fonction crée une colonne dans laquelle on entre pour chaque ligne le nom de la colonne ou des colonnes qui contient le max.
        """  
        
        self._write_new_column_title()
        map_columns_to_tags = self._map_columns_to_tags() 
 
        for line_index in range(self.first_line, self.tab.max_row + 1): 
            self._insert_tags_of_maximum_in_cell(line_index, map_columns_to_tags) 
    
    def _write_new_column_title(self):
        column_to_write = column_index_from_string(self.tab_options.column_to_write)
        self.tab.cell(1, column_to_write).value = "Colonne de(s) maximum(s)"

    def _map_columns_to_tags(self):
        #Les tags sont situés dans la première cellule de chaque colonne.
        map_columns_to_tags = {}

        for column_index in self.tab_options.columns_to_read:
            map_columns_to_tags[column_index] = self.tab.cell(1, column_index).value
        return map_columns_to_tags
    
    def _insert_tags_of_maximum_in_cell(self, line_index, map_columns_to_tags):
        maximum = -1
        tags = "" 
        
        for column_index in self.tab_options.columns_to_read: 
            cell_value = self.tab.cell(line_index, column_index).value 
            if cell_value > maximum:
                maximum = cell_value
                tags = map_columns_to_tags[column_index]
            elif cell_value == maximum: # On adjoint le tag au tag déjà construit en cas d'égalité
                tags += "_" + map_columns_to_tags[column_index]

        column_to_write = column_index_from_string(self.tab_options.column_to_write)
        self.tab.cell(line_index, column_to_write).value = tags 

    @save_file
    @act_on_columns 
    def transform_string_in_binary_in_column(self, *good_answers):
        """
        Fonction qui prend une colonne de chaîne de caractères et qui renvoie une colonne de 0 ou de 1. On renvoit 1
        si la cellule contient un élément de good_answers.
        """  
        column_to_read = self.tab_options.column_to_read
        column_to_write = column_index_from_string(self.tab_options.column_to_write)

        for line_index in range(self.first_line, self.tab.max_row + 1):
            cell_read = Cell(line_index, column_to_read)
            cell_value = self.file_object.get_compiled_cell_value(self.tab, cell_read)  
            cleaned_cell_value = self.clean_string_from_spaces(cell_value)
            binary = self.transform_string_in_binary(cleaned_cell_value, *good_answers) 
            self.tab.cell(line_index, column_to_write).value = binary

    @save_file
    @act_on_columns 
    def convert_time_in_minutes_in_columns(self):
        """
        Fonction qui prend une colonne de chaines de caractères de la forme "10 jours 5 heures" 
        ou "5 heures 10 min" ou "10 min 5s" ou "5s" et qui renvoie le temps en minutes. 
        """   
        for line_index in range(self.first_line, self.tab.max_row + 1):
            self._convert_filled_cell_in_minutes(line_index)

    def _convert_filled_cell_in_minutes(self, line_index):
        column_to_read = self.tab_options.column_to_read
        column_to_write = column_index_from_string(self.tab_options.column_to_write)

        cell_read = Cell(line_index, column_to_read)
        cell_value = self.file_object.get_compiled_cell_value(self.tab, cell_read) 

        if cell_value != "None":
            cleaned_cell_value = self.clean_string_from_spaces(cell_value)
            time_in_min = self.convert_time_in_minutes(cleaned_cell_value) 
            self.tab.cell(line_index, column_to_write).value = time_in_min

    @save_file
    @act_on_columns 
    def insert_group_associated_with_answer(self, map_groups_to_answers):
        """
        Cette fonction qui prend une colonne de chaîne de caractères 
        et qui renvoie une colonne contenant pour chaque cellule le groupe associé à la chaîne de caractères.
        """  
        column_to_read = self.tab_options.column_to_read
        column_to_write = column_index_from_string(self.tab_options.column_to_write)
        
        map_answers_to_groups = Dictionary.reverse_dictionary(map_groups_to_answers)

        for line_index in range(self.first_line, self.tab.max_row + 1):
            cell_read = Cell(line_index, column_to_read)
            cell_value = self.file_object.get_compiled_cell_value(self.tab, cell_read)  
            answer = self.clean_string_from_spaces(cell_value)
            group = self.set_answer_in_group(answer, map_answers_to_groups) 
            self.tab.cell(line_index, column_to_write).value = group 