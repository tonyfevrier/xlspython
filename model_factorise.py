import openpyxl
import os   

""" def display_run():
    def wrapper(method, *args, **kwargs):
        start = time()
        method(*args, **kwargs)
        Other.display_running_infos('sauvegarde', tab_name, self.sheets_name, start)
    return wrapper """

class Path():
    def __init__(self, pathname = 'fichiers_xls/'):
        self.pathname = pathname
        self.directories = [f for f in os.listdir(self.pathname) if os.path.isdir(os.path.join(self.pathname, f))]
        

class File(): 
    def __init__(self, name_file, path = 'fichiers_xls/', dataonly = False):
        """
        Handle methods modifying a tab
        """ 
        self.name_file = name_file  
        self.path = path
        self.dataonly = dataonly 
        self.writebook = openpyxl.load_workbook(self.path + self.name_file, data_only = dataonly)
        self.sheets_name = self.writebook.sheetnames 
        
    def create_and_return_new_tab(self, tab_name):
        return self.writebook.create_sheet(tab_name)

    def get_tab_by_name(self, tab_name):
        return self.writebook[tab_name]
    
    def get_cell_value_from_a_tab(self, tab, cell):
        return str(tab.cell(cell.line_index, cell.column_index).value)
    
    def update_sheet_names(self):
        self.sheets_name = self.writebook.sheetnames 

    def save_file(self):
        self.writebook.save(self.path + self.name_file) 


class Tab():
    """Handle methods modifying a tab"""
    def __init__(self, tab_name):   
        self.tab_name = tab_name 


class Cell():
    """Handle methods modifying a cell"""
    def __init__(self, line_index, column_index):  
        self.line_index = line_index
        self.column_index = column_index


class OptionalNamesOfTab():
    """Get all names of tabs, columns, lines which will be read or modified in methods """
    def __init__(self, column_to_read=None, columns_to_read=None,
                 column_to_write=None, columns_to_write=None): 
        
        self.column_to_read = column_to_read
        self.columns_to_read = columns_to_read
        self.column_to_write = column_to_write
        self.columns_to_write = columns_to_write


class OptionalNamesOfFile(OptionalNamesOfTab):
    """Get all names of tabs, columns, lines which will be read or modified in methods """
    def __init__(self, name_of_tab_to_read=None, names_of_tabs_to_read=None,
                 name_of_tab_to_modify=None, names_of_tabs_to_modify=None, column_to_read=None, columns_to_read=None,
                 column_to_write=None, columns_to_write=None):
        
        OptionalNamesOfTab.__init__(self, column_to_read, columns_to_read, column_to_write, columns_to_write)
        self.name_of_tab_to_read = name_of_tab_to_read
        self.names_of_tabs_to_read = names_of_tabs_to_read
        self.name_of_tab_to_modify = name_of_tab_to_modify 
        self.names_of_tabs_to_modify = names_of_tabs_to_modify 
 





