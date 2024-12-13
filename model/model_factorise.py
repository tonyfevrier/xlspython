import openpyxl
import os   
import yagmail

from pycel import ExcelCompiler

""" def display_run():
    def wrapper(method, *args, **kwargs):
        start = time()
        method(*args, **kwargs)
        Other.display_running_infos('sauvegarde', tab_name, self.sheets_name, start)
    return wrapper """


class Path():
    def __init__(self, pathname = 'fichiers_xls/'):
        self.pathname = pathname
        self.directories = [f for f in os.listdir(self.pathname) if os.path.isdir(os.path.join(self.pathname, f))]
        

class File(): 
    """Classe permettant de récupérer et mettre à jour les données d'un fichier"""
    def __init__(self, name_file, path = 'fichiers_xls/', dataonly = False): 
        self.name_file = name_file  
        self.path = path
        self.dataonly = dataonly 
        self.writebook = openpyxl.load_workbook(self.path + self.name_file, data_only = dataonly)
        self.sheets_name = self.writebook.sheetnames 
        self.compiler = None
    
    def _create_excel_compiler(self):
        return ExcelCompiler(self.path + self.name_file) 
        
    def create_and_return_new_tab(self, tab_name):
        return self.writebook.create_sheet(tab_name)

    def get_tab_by_name(self, tab_name):
        return self.writebook[tab_name]
    
    def get_uncompiled_cell_value(self, tab, cell):
        return tab.cell(cell.line_index, cell.column_index).value
    
    def evaluate_cell_formula(self, tab, cell):
        """
        Fonction qui calcule une valeur numérique liée à une formule
        """  
        self.compiler = self._create_excel_compiler()
        return self.compiler.evaluate(tab.title + '!' + tab.cell(cell.line_index, cell.column_index).coordinate) 

    def get_compiled_cell_value(self, tab, cell):
        """
        Fonction qui prend la valeur d'une cellule et qui, si c'est une formule, retourne sa valeur numérique
        """  

        cell_value = self.get_uncompiled_cell_value(tab, cell)
        if self._is_cell_value_a_formula(cell_value):
            return self.evaluate_cell_formula(tab, cell)
        else:
            return str(cell_value)
    
    @staticmethod
    def _is_cell_value_a_formula(cell_value):
        return isinstance(cell_value, str) and cell_value.startswith('=')
    
    def update_sheet_names(self):
        self.sheets_name = self.writebook.sheetnames 

    def save_file(self):
        self.writebook.save(self.path + self.name_file) 


class Cell():
    def __init__(self, line_index, column_index):  
        self.line_index = line_index
        self.column_index = column_index
        

class MergedCellsRange():
    def __init__(self, start_column, end_column, start_line, end_line):
        self.start_column = start_column
        self.end_column = end_column
        self.start_line = start_line
        self.end_line = end_line
        

class TabOptions():
    """Get all names of tabs, columns, lines which will be read or modified in methods """
    def __init__(self, column_to_read=None, columns_to_read=None,
                 column_to_write=None, columns_to_write=None): 
        
        self.column_to_read = column_to_read
        self.columns_to_read = columns_to_read
        self.column_to_write = column_to_write
        self.columns_to_write = columns_to_write

    def __repr__(self):
        return f'TabOptions(column_to_read={self.column_to_read}, columns_to_read={self.columns_to_read}, column_to_write={self.column_to_write}, columns_to_write={self.columns_to_write})'


class FileOptions(TabOptions):
    """Get all names of tabs, columns, lines which will be read or modified in methods """
    def __init__(self, name_of_tab_to_read=None, names_of_tabs_to_read=None,
                 name_of_tab_to_modify=None, names_of_tabs_to_modify=None, column_to_read=None, columns_to_read=None,
                 column_to_write=None, columns_to_write=None):
        
        TabOptions.__init__(self, column_to_read, columns_to_read, column_to_write, columns_to_write)
        self.name_of_tab_to_read = name_of_tab_to_read
        self.names_of_tabs_to_read = names_of_tabs_to_read
        self.name_of_tab_to_modify = name_of_tab_to_modify 
        self.names_of_tabs_to_modify = names_of_tabs_to_modify 
 

class Mail():
    def __init__(self, sender_mail, receiver_mail, joint_file, subject, message, password):
        self.sender_mail = sender_mail
        self.receiver_mail = receiver_mail
        self.joint_file = joint_file
        self.subject = subject
        self.message = message
        self.password = password
        
    def send(self): 
        """
        Fonction qui envoie un mail avec une pièce jointe. 
        """
        yag = yagmail.SMTP(self.sender_mail, self.password)  
        yag.send(to=self.receiver_mail, subject=self.subject, contents=self.message, attachments = self.joint_file)





