from unittest import TestCase, main 
from model.model import File, FileOptions, TabOptions, MergedCellsRange, Path
from controller.one_file_one_tab import ColorTabController, DeleteController, InsertController 
from controller.one_file_multiple_tabs import OneTabCreatedController, MultipleSameTabController, EvenTabsController
from controller.two_files import OneFileCreatedController, TwoFilesController
from controller.path import SeveralFoldersOneFileController
from utils.utils import ColumnDelete, ColumnInsert, LineDelete, LineInsert, TabUpdateFormula, MapIndexLetter, String, Dictionary
from openpyxl.utils import column_index_from_string

import os
 

class TestPath(TestCase):

    def setUp(self):
        self.path = None
        self.method_data = None
        self.file_controller = None
        self.file_data_compare1 = None
        self.file_data_compare2 = None
    
    def test_split_one_tab_in_multiple_tabs(self):
        self._build_split_one_tab_in_multiple_tabs()
        self._apply_method_on_all_files()
        self._compare_created_file_to_expected_file()
        self._delete_created_file()

    def _build_split_one_tab_in_multiple_tabs(self):
        self.path = Path('fichiers_xls/gathertests/')
        self.method_data = MethodData('split_one_tab_in_multiple_tabs')
        file_options = FileOptions(name_of_tab_to_read='test', column_to_read='A')
        self.file_controller = OneFileCreatedController(file_options=file_options, new_path=self.path.pathname)
        self.test_file_name = 'test_cmd_ongletbyparticipant.xlsx'
        self.expected_file_name = 'divided_test_cmd_ongletbyparticipant_before.xlsx' 
        self.created_file_name = 'divided_test_cmd_ongletbyparticipant.xlsx'

    def _apply_method_on_all_files(self, args=()):
        controller = SeveralFoldersOneFileController(self.path, self.test_file_name, self.file_controller)
        controller.apply_method_on_homononymous_files(self.method_data.method_name, *args) 

    def _compare_created_file_to_expected_file(self):
        self.file_data_compare1 = FileData(self.expected_file_name, path=self.path.pathname)
        self.file_data_compare2 = FileData(self.created_file_name, path=self.path.pathname)
        assert_object = AssertIdentical(self.file_data_compare1, self.file_data_compare2)
        assert_object.verify_files_identical()
    
    def _delete_created_file(self):
        os.remove(self.path.pathname + self.file_data_compare2.name_file) 

    def test_copy_a_tab_at_tab_bottom(self):
        self._build_copy_a_tab_at_tab_bottom()
        self._apply_method_on_all_files(self.method_data.args)
        self._compare_created_file_to_expected_file()
        self._delete_created_file() 

    def _build_copy_a_tab_at_tab_bottom(self):
        self.path = Path('gatherfiles/')
        self.method_data = MethodData('copy_a_tab_at_tab_bottom', 'Sheet')
        self.file_controller = OneFileCreatedController(new_path=self.path.pathname)
        self.test_file_name = 'test.xlsx'
        self.expected_file_name = 'gathered_test_ref.xlsx'
        self.created_file_name = 'gathered_test.xlsx' 

    def test_transform_string_in_binary_in_column(self):
        self._build_transform_string_in_binary_in_column()
        self._apply_method_on_all_tabs()
        self._compare_transform_string_in_binary_in_column()     

    def _build_transform_string_in_binary_in_column(self):
        self.path = Path('fichiers_xls/gathertests/')
        tab_options = TabOptions(column_to_read='G', column_to_write='H')
        self.file_controller = InsertController(tab_name='sheet1' , tab_options = tab_options, save=True)
        self.method_data = MethodData('transform_string_in_binary_in_column', 'partie 1 : Vrai')
        self.test_file_name = 'test_string_in_binary.xlsx'

    def _apply_method_on_all_tabs(self):
        controller = SeveralFoldersOneFileController(self.path, self.test_file_name, self.file_controller)
        controller.apply_method_on_homononymous_tabs(self.method_data.method_name, *self.method_data.args) 

    def _compare_transform_string_in_binary_in_column(self):
        folders_names = ['Nouveau dossier/', 'Nouveau dossier - Copie/']
        for name in folders_names:
            self._compare_modified_tab_to_expected(name)
            self._delete_created_columns_and_save()

    def _compare_modified_tab_to_expected(self, name):
        self.file_data_compare = FileData(self.test_file_name, 'sheet1', path=self.path.pathname + name) 
        self.file_data_compare_ref = FileData(self.test_file_name, 'expected', path=self.path.pathname + name)
        self.assert_object = AssertIdentical(self.file_data_compare, self.file_data_compare_ref)
        self.assert_object.verify_tabs_identical()

    def _delete_created_columns_and_save(self):
        tab_name = self.file_data_compare.tab_name
        tab = self.file_data_compare.file_object.writebook[tab_name] 
        tab.delete_cols(8)
        self.file_data_compare.file_object.save_file()
        

class TestOneTabCreatedController(TestCase):

    def setUp(self):
        self.file_object = None
        self.controller = None
        self.file_data_compare_1 = None
        self.file_data_compare_2 = None
        self.file_options = None

    def test_extract_a_column_from_all_tabs(self):
        self._build_files_to_compare('test_extract_column.xlsx', 'test_extract_column_ref.xlsx')
        self.file_options = FileOptions(column_to_read='B')
        
        self.controller = OneTabCreatedController(self.file_data_compare_1.file_object, self.file_options)
        self.controller.extract_a_column_from_all_tabs()
 
        self._compare_files() 
        tab_to_delete = [self.file_data_compare_1.file_object.sheets_name[-1]]
        self._delete_created_tabs(tab_to_delete)
    
    def _build_files_to_compare(self, file_name, file_name_reference): 
        self.file_data_compare_1 = FileData(file_name)
        self.file_data_compare_1.create_file_object()
        self.file_data_compare_2 = FileData(file_name_reference)
        self.file_data_compare_2.create_file_object() 

    def _compare_files(self):
        self.assert_object = AssertIdentical(self.file_data_compare_1, self.file_data_compare_2) 
        self.assert_object.verify_files_identical()

    def _compare_tabs(self):
        self.assert_object = AssertIdentical(self.file_data_compare_1, self.file_data_compare_2)  
        self.assert_object.verify_tabs_identical()

    def _delete_created_tabs(self, tabs_to_delete):
        self.file_object = self.file_data_compare_1.file_object
        for tab in tabs_to_delete: 
            del self.file_object.writebook[tab]
        self.file_object.save_file()

    def test_gather_groups_of_multiple_columns_in_tabs_of_two_columns_containing_tags_and_values(self):
        self._build_files_to_compare('test_gather_columns_in_one.xlsx', 'test_gather_columns_in_one - ref.xlsx')
        self.file_options = FileOptions(name_of_tab_to_read='test')

        self.controller = OneTabCreatedController(self.file_data_compare_1.file_object, self.file_options)
        self.controller.gather_groups_of_multiple_columns_in_tabs_of_two_columns_containing_tags_and_values(['C','D','E'], ['G','H','I'])
 
        self._compare_files()
        tabs_to_delete = ['tab_column_gathered_CDE', 'tab_column_gathered_GHI']
        self._delete_created_tabs(tabs_to_delete)

    def test_gather_multiple_answers(self):
        self._build_gather_multiple_answers()

        self.controller = OneTabCreatedController(self.file_data_compare_1.file_object, self.file_options)
        self.controller.gather_multiple_answers('A', 'B')

        self._compare_tabs()
        tab_to_delete = ['multiple_answers']
        self._delete_created_tabs(tab_to_delete)

    def _build_gather_multiple_answers(self): 
        self.file_data_compare_1 = FileData('testongletbyparticipant.xlsx', 'multiple_answers')
        self.file_data_compare_1.create_file_object()
        self.file_data_compare_2 = FileData('testongletbyparticipant-result.xlsx', 'Feuille2')
        self.file_data_compare_2.create_file_object() 
        self.file_options = FileOptions(name_of_tab_to_read='test')


class TestEvenTabsController(TestCase):
    def setUp(self):
        self.file_object = None
        self.file_options = None
        self.tab_names = []
        self.tab_name_to_compare = None

    def test_apply_column_formula_on_all_tabs(self):
        self._build_apply_column_formula_on_all_tabs()

        controller = EvenTabsController(self.file_object, self.file_options)        
        controller.apply_columns_formula_on_all_tabs()

        self._compare_multiple_tabs()

    def _build_apply_column_formula_on_all_tabs(self):
        self.file_object = File('dataset.xlsx', dataonly = False)
        self.file_options = FileOptions(columns_to_read=['B','C'])
        self.tab_names = ['Feuille2', 'Feuille3', 'Feuille4']
        self.tab_name_to_compare = 'ref'

    def _compare_multiple_tabs(self):
        file_name = self.file_object.name_file
        file_data_compare = FileData(file_name, self.tab_name_to_compare)
        file_data_list = [FileData(file_name, tab_name) for tab_name in self.tab_names]

        for file_data in file_data_list:
            assert_object = AssertIdentical(file_data, file_data_compare)
            assert_object.verify_tabs_identical()

    def test_apply_cell_formula_on_all_tabs(self):
        self._build_apply_cell_formula_on_all_tabs()

        controller = EvenTabsController(self.file_object)
        controller.apply_cells_formula_on_all_tabs(*self.cells)

        self._compare_cells_of_multiple_tabs()

    def _build_apply_cell_formula_on_all_tabs(self): 
        self.file_object = File("test_merging.xlsx") 
        self.tab_names = self.file_object.sheets_name[1:]
        self.tab_name_to_compare = self.file_object.sheets_name[0]
        self.cells = ['A10','B10','C10']

    def _compare_cells_of_multiple_tabs(self):
        file_name = self.file_object.name_file
        file_data_compare = FileData(file_name, self.tab_name_to_compare)

        for tab_name in self.tab_names:
            file_data = FileData(file_name, tab_name)
            assert_object = AssertIdentical(file_data, file_data_compare)
            assert_object.verify_cells_identical(*self.cells)
    
    def test_merge_cells_on_all_tabs(self): 
        self.file_object = File("test_merging.xlsx")
        
        controller = EvenTabsController(self.file_object)
        merged_cells_range = MergedCellsRange('E', 'G', 12, 15)
        controller.merge_cells_on_all_tabs(merged_cells_range)        

    def test_list_tabs_with_different_number_of_lines(self):
        self.file_object = File('test.xlsx')
        expected_tabs = ['cutinparts', 'cutinpartsbis', 'delete_lines', 'delete_lines_bis', 'time_min', 'time_min_expected']

        controller = EvenTabsController(self.file_object)
        tabs = controller.list_tabs_with_different_number_of_lines(14)

        self.assertListEqual(tabs, expected_tabs)


class TestOneFileCreatedController(TestCase):

    def setUp(self): 
        self.file_object = None
        self.controller = None
        self.file_data_compare_1 = None
        self.file_data_compare_2 = None
        self.file_options = None

    def _compare_tabs(self):
        self.file_data_compare_2.create_file_object()
        self.assert_object = AssertIdentical(self.file_data_compare_1, self.file_data_compare_2) 
        self.assert_object.verify_tabs_identical()

    def _compare_files(self):
        self.file_data_compare_2.create_file_object()
        self.assert_object = AssertIdentical(self.file_data_compare_1, self.file_data_compare_2) 
        self.assert_object.verify_files_identical()

    def _delete_created_file(self):
        os.remove(self.file_data_compare_2.file_object.path + self.file_data_compare_2.name_file)

    def test_extract_cells_from_all_tabs(self): 
        self._build_extract_cells_from_all_tabs_data()

        self.controller = OneFileCreatedController(self.file_object)
        self.controller.extract_cells_from_all_tabs('C7','D7','C8','D8') 

        self._compare_tabs()
        self._delete_created_file() 
    
    def _build_extract_cells_from_all_tabs_data(self):
        self.file_object = File('test_extract_cells_from_all_sheets.xlsx')  
        self.file_data_compare_1 = FileData('test_extract_cells_from_all_sheets - after.xlsx', 'gathered_data') 
        self.file_data_compare_2 = FileData('gathered_data_test_extract_cells_from_all_sheets.xlsx', 'Sheet') 

    def test_split_one_tab_in_multiple_tabs(self): 
        self._build_split_one_tab_in_multiple_tabs_data() 

        self.controller = OneFileCreatedController(self.file_object, file_options=self.file_options)
        self.controller.split_one_tab_in_multiple_tabs() 

        self._compare_files()
        self._delete_created_file()
    
    def _build_split_one_tab_in_multiple_tabs_data(self):
        self.file_object = File('test_create_one_onglet_by_participant.xlsx')  
        self.file_data_compare_1 = FileData('test_create_one_onglet_by_participant_before.xlsx')
        self.file_options = FileOptions(name_of_tab_to_read='Stroops_test (7)', column_to_read='A')
        self.file_data_compare_2 = FileData('divided_test_create_one_onglet_by_participant.xlsx')

    def test_one_file_by_tab(self):
        self._build_one_file_by_tab_for_first_tab()

        self.controller = OneFileCreatedController(self.file_object)
        self.controller.create_one_file_by_tab()

        self._compare_tabs()
        self._delete_created_file()

        self._build_one_file_by_tab_for_second_tab()

        self._compare_tabs()
        self._delete_created_file()
    
    def _build_one_file_by_tab_for_first_tab(self):
        self.file_object = File('test_onefile_sendmail.xlsx')  
        self.file_data_compare_1 = FileData('test_onefile_sendmail.xlsx', 'tony fevrier') 
        self.file_data_compare_2 = FileData('tony fevrier.xlsx', 'Sheet', path="multifiles/")

    def _build_one_file_by_tab_for_second_tab(self):
        self.file_data_compare_1 = FileData('test_onefile_sendmail.xlsx', 'Marine Moyon') 
        self.file_data_compare_2 = FileData('Marine Moyon.xlsx', 'Sheet', path="multifiles/")


class TestColumnInsertion(TestCase):

    def setUp(self): 
        self.controller = None
        self.file_data = None
        self.file_data_compare = None
        self.method_data = None 
        self.tab_options = None
        self.columns_to_compare = []
        self.columns_to_delete = []

    def apply_compare_columns_restore(fonction):
        def wrapper(self):
            fonction(self) 
            self._apply_the_method_to_test() 
            self._compare_new_columns()
            self._restore_file_state_before_modification()
        return wrapper 
    
    def apply_compare_tabs_restore(fonction):
        def wrapper(self):
            fonction(self) 
            self._apply_the_method_to_test() 
            self._compare_tabs()
            self._restore_file_state_before_modification()
        return wrapper 
    
    def _apply_the_method_to_test(self):
        tab_controller = InsertController(self.file_data.file_object, tab_options=self.tab_options)  
        file_options = FileOptions(names_of_tabs_to_modify=[self.file_data.tab_name])
        self.controller = MultipleSameTabController(self.file_data.file_object, tab_controller, file_options)
        self.controller.apply_method_on_some_tabs(self.method_data.method_name, *self.method_data.args, **self.method_data.kwargs) 

    def _compare_new_columns(self):
        self.assert_object = AssertIdentical(self.file_data, self.file_data_compare) 
        columns_to_compare = MapIndexLetter.get_list_of_columns_indexes(self.columns_to_compare)
        for column in columns_to_compare:
            self.assert_object.verify_columns_identical(column, column)

    def _compare_tabs(self):
        self.assert_object = AssertIdentical(self.file_data, self.file_data_compare) 
        self.assert_object.verify_tabs_identical()
    
    def _restore_file_state_before_modification(self):
        tab = self.file_data.file_object.writebook[self.file_data.tab_name]  
        tab.delete_cols(column_index_from_string(self.columns_to_delete[0]), len(self.columns_to_delete))
        modification_object = ColumnDelete(self.columns_to_delete)
        self.controller.tab_controller._update_cell_formulas(modification_object) 
        self.file_data.file_object.save_file()

    @apply_compare_columns_restore
    def test_transform_string_in_binary_in_column(self): 
        self.file_data = FileData('test.xlsx', 'Feuille2')
        self.file_data_compare = FileData('test.xlsx', 'Feuille2')
        self.tab_options = TabOptions(column_to_read='F', column_to_write='G')
        self.method_data = MethodData('transform_string_in_binary_in_column', 'partie 12 : Faux', 1) 
        self.columns_to_compare = ['G']
        self.columns_to_delete = self.columns_to_compare

    @apply_compare_columns_restore
    def test_convert_time_in_minutes_in_columns(self): 
        self.file_data = FileData('test.xlsx', 'time_min')
        self.file_data_compare = FileData('test.xlsx', 'time_min_expected')
        self.tab_options = TabOptions(column_to_read='E', column_to_write='F')
        self.method_data = MethodData('convert_time_in_minutes_in_columns') 
        self.columns_to_compare = ['F']
        self.columns_to_delete = self.columns_to_compare

    @apply_compare_columns_restore
    def test_insert_group_associated_with_answer(self):
        self.file_data = FileData('test_column_set_answer.xlsx', 'sheet1')
        self.file_data_compare = FileData('test_column_set_answer.xlsx', 'Feuille2')
        self.tab_options = TabOptions(column_to_read='B', column_to_write='C')
        map_groups_to_answers = {"group1":['2','5','6'], "group2":['7','8','9'], "group3":['1','3','4'], "group4":['10']}  
        self.method_data = MethodData('insert_group_associated_with_answer', map_groups_to_answers)
        self.columns_to_compare = ['C']
        self.columns_to_delete = self.columns_to_compare

    @apply_compare_columns_restore
    def test_insert_splitted_strings_of(self):
        self.file_data = FileData('test.xlsx', 'cutinparts')
        self.file_data_compare = FileData('test.xlsx', 'cutinpartsbis') 
        self.tab_options = TabOptions(column_to_read='B', column_to_write='C')
        self.method_data = MethodData('insert_splitted_strings_of', ';')
        self.columns_to_compare = ['C', 'D', 'E']
        self.columns_to_delete = self.columns_to_compare

    @apply_compare_columns_restore
    def test_fill_one_column_by_QCM_answer(self):
        self.file_data = FileData('test_create_one_column.xlsx', 'sheet1')
        self.file_data_compare = FileData('test_create_one_column.xlsx', 'Feuille2') 
        self.tab_options = TabOptions(column_to_read='D', column_to_write='E')
        self.method_data = MethodData('fill_one_column_by_QCM_answer', 'Alain', 'Henri', 'Tony', 'Dulcinée')
        self.columns_to_compare = ['E', 'F', 'G', 'H']
        self.columns_to_delete = self.columns_to_compare

    @apply_compare_columns_restore
    def test_write_piece_of_string_in_column_1(self):
        self.file_data = FileData('test_colgetpartofstr.xlsx', 'Feuille2')
        self.file_data_compare = FileData('test_colgetpartofstr.xlsx', 'expected') 
        self.tab_options = TabOptions(column_to_read='C', column_to_write='D')
        self.method_data = MethodData('write_piece_of_string_in_column', '_', 0)
        self.columns_to_compare = ['D']
        self.columns_to_delete = self.columns_to_compare

    @apply_compare_columns_restore
    def test_write_piece_of_string_in_column_2(self):
        self.file_data = FileData('test_colgetpartofstr.xlsx', 'Feuille2')
        self.file_data_compare = FileData('test_colgetpartofstr.xlsx', 'expected2') 
        self.tab_options = TabOptions(column_to_read='E', column_to_write='F')
        self.method_data = MethodData('write_piece_of_string_in_column', ';', 1)
        self.columns_to_compare = ['F']
        self.columns_to_delete = self.columns_to_compare

    @apply_compare_columns_restore
    def test_map_two_columns_to_a_third_column(self):
        self.file_data = FileData('test_maptwocolumns.xlsx', 'Feuille2')
        self.file_data_compare = FileData('test_maptwocolumns.xlsx', 'expected') 
        self.tab_options = TabOptions(columns_to_read=['B', 'C'], column_to_write='D')
        self.method_data = MethodData('map_two_columns_to_a_third_column', {'cat1':['prime','1'], 'cat2':['probe','2']})
        self.columns_to_compare = ['D']
        self.columns_to_delete = self.columns_to_compare

    @apply_compare_tabs_restore
    def test_verify_tabs_when_map_two_columns_to_a_third_column(self):
        self.file_data = FileData('test_maptwocolumns.xlsx', 'Feuille2')
        self.file_data_compare = FileData('test_maptwocolumns.xlsx', 'expected') 
        self.tab_options = TabOptions(columns_to_read=['B', 'C'], column_to_write='D')
        self.method_data = MethodData('map_two_columns_to_a_third_column', {'cat1':['prime','1'], 'cat2':['probe','2']})
        self.columns_to_delete = ['D']

    @apply_compare_tabs_restore
    def test_insert_column_for_prime_probe_congruence(self):   
        self.file_data = FileData('test_prime_probe.xlsx', 'Feuille2')
        self.file_data_compare = FileData('test_prime_probe.xlsx', 'expected') 
        self.tab_options = TabOptions(columns_to_read=['B', 'C', 'D'], column_to_write='E')
        self.method_data = MethodData('insert_column_for_prime_probe_congruence')
        self.columns_to_delete = ['E']

    @apply_compare_tabs_restore
    def test_insert_tags_of_maximum_of_column_list(self):
        self.file_data = FileData('test_give_names.xlsx', 'sheet1')
        self.file_data_compare = FileData('test_give_names.xlsx', 'Feuille2') 
        self.tab_options = TabOptions(columns_to_read=['A', 'B', 'C'], column_to_write='D')
        self.method_data = MethodData('insert_tags_of_maximum_of_column_list')
        self.columns_to_delete = ['D']


class TestDeleteItems(TestCase):
    def setUp(self):
        self.file_data = None
        self.file_data_compare = None
        self.tab_controller = None
        self.file_options = None
        self.columns_to_compare = []

    def test_delete_lines_containing_strings_in_given_column(self):
        self._build_delete_lines_containing_strings_in_given_column()
        
        self.tab_controller = DeleteController(self.file_data.file_object)
        controller = MultipleSameTabController(self.file_data.file_object, self.tab_controller, self.file_options)
        controller.apply_method_on_some_tabs('delete_lines_containing_strings_in_given_column', 'D', '0', 'p a') 

        self._compare_new_columns() 
    
    def _build_delete_lines_containing_strings_in_given_column(self):
        self.file_data = FileData('test.xlsx', 'delete_lines')
        self.file_data_compare = FileData('test.xlsx', 'delete_lines_bis') 
        self.file_options = FileOptions(names_of_tabs_to_modify=['delete_lines'])
        self.columns_to_compare = [i for i in range(1, 7)]

    def _compare_new_columns(self):
        self.assert_object = AssertIdentical(self.file_data, self.file_data_compare)  
        for column in self.columns_to_compare:
            self.assert_object.verify_columns_identical(column, column)

    def test_delete_lines_containing_strings_in_given_column_with_formulas(self):
        self._build_delete_lines_containing_strings_in_given_column_with_formulas()

        self.tab_controller = DeleteController(self.file_data.file_object)
        controller = MultipleSameTabController(self.file_data.file_object, self.tab_controller, self.file_options)
        controller.apply_method_on_some_tabs('delete_lines_containing_strings_in_given_column', 'B', 'pas consenti')         
        
        self._compare_new_columns()

    def _build_delete_lines_containing_strings_in_given_column_with_formulas(self):
        self.file_data = FileData('listing_par_etape - Copie.xlsx', 'Feuil1')
        self.file_data_compare = FileData('listing_par_etape - Copie.xlsx', 'Feuille2') 
        self.file_options = FileOptions(names_of_tabs_to_modify=['Feuil1'])
        self.columns_to_compare = [2, 10]

    def test_delete_twins_lines_and_color_last_twin(self): 
        self._build_delete_twins_lines_and_color_last_twin()

        self.tab_controller = DeleteController(self.file_object)
        controller = MultipleSameTabController(self.file_object, self.tab_controller, self.file_options)
        controller.apply_method_on_some_tabs('delete_twins_lines_and_color_last_twin', 'C', color = 'FFFFFF00')

        self._compare_multiple_tabs()        

    def _build_delete_twins_lines_and_color_last_twin(self):
        self.file_object = File('test_doublons.xlsx')
        self.file_data_compare = FileData('test_doublons.xlsx', 'result')
        self.tab_names = ['sheet2', 'sheet1', 'sheet3']
        self.file_options = FileOptions(names_of_tabs_to_modify=self.tab_names)

    def _compare_multiple_tabs(self):
        for tab_name in self.tab_names:
            self.file_data = FileData(self.file_object.name_file, tab_name)
            self.assert_object = AssertIdentical(self.file_data, self.file_data_compare) 
            self.assert_object.verify_tabs_identical()
        
    def test_delete_other_columns(self):
        self._build_delete_columns()

        self.controller = MultipleSameTabController(self.file_data.file_object, self.tab_controller, self.file_options)
        self.controller.apply_method_on_some_tabs('delete_other_columns', 'A-C,D-K')

        self._compare_tab()

    def test_delete_columns(self):
        self._build_delete_columns()

        self.controller = MultipleSameTabController(self.file_data.file_object, self.tab_controller, self.file_options)
        self.controller.apply_method_on_some_tabs('delete_columns', 'L,M,N-V')

        self._compare_tab()

    def _build_delete_columns(self):
        tab_name = 'sheet1'
        self.file_data = FileData('test_keep_only_columns.xlsx', tab_name)
        self.file_data_compare = FileData('test_keep_only_columns.xlsx', 'Feuille2')
        self.tab_controller = DeleteController(self.file_data.file_object, tab_name) 
        self.file_options = FileOptions(names_of_tabs_to_modify=[tab_name])

    def _compare_tab(self):
        self.assert_object = AssertIdentical(self.file_data, self.file_data_compare)
        self.assert_object.verify_tabs_identical()


class TestColorItems(TestCase):
    def setUp(self):
        self.file_object = None
        self.tab_controller = None
        self.file_options = None
        self.method_data = None

    def apply_method(fonction):
        def wrapper(self): 
            fonction(self)
            self._apply_the_method_to_test() 
        return wrapper
    
    def _apply_the_method_to_test(self): 
        controller = MultipleSameTabController(self.file_object,
                                              self.tab_controller, 
                                              file_options=self.file_options)   
        controller.apply_method_on_some_tabs(self.method_data.method_name, *self.method_data.args)
    
    @apply_method
    def test_color_cases_in_column(self):
        self.file_object = File('test.xlsx')
        tab_options = TabOptions(column_to_read='D')
        self.file_options = FileOptions(names_of_tabs_to_modify=['cutinpartsbis'])
        self.method_data = MethodData('color_cases_in_column', {' partie 2 : Vrai':'0000a933'})
        self.tab_controller = ColorTabController(self.file_object, tab_options=tab_options)
    
    @apply_method
    def test_color_cases_in_sheet(self):
        self.file_object = File('test.xlsx')
        tab_options = TabOptions()
        self.file_options = FileOptions(names_of_tabs_to_modify=['cutinpartsbis'])
        self.method_data = MethodData('color_cases_in_tab', {'partie 1 : Vrai':'0000a933', 'Abbas':'0000a933'})
        self.tab_controller = ColorTabController(self.file_object, tab_options=tab_options)
    
    @apply_method    
    def test_color_line_containing_chaines(self):
        self.file_object = File('test.xlsx')
        self.file_options = FileOptions(names_of_tabs_to_modify=['color_line'])
        self.method_data = MethodData('color_lines_containing_strings', '-', '+')
        self.tab_controller = ColorTabController(self.file_object, color='0000a933')
 

class TestTwoFilesController(TestCase):
    def setUp(self): 
        self.file_object = None
        self.controller = None
        self.file_data_compare_1 = None
        self.file_data_compare_2 = None
        self.column_to_read_1 = None
        self.column_to_read_2 = None
        self.columns_to_compare = []
        self.columns_to_copy = []
        self.column_insertion = None

    def test_copy_columns_in_a_tab_differently_sorted(self):
        self._build_test()
        self._apply_method_to_test()
        self._compare_new_columns()
        self._restore_file_state_and_save() 

    def _build_test(self):
        self.file_object_from = File('test.xlsx', dataonly=True)
        self.tab_name_from = 'sheet1'
        self.file_data_compare_1 = FileData('test.xlsx', 'Feuille5')
        self.file_data_compare_2 = FileData('test.xlsx', 'Feuille5bis')
        self.file_data_compare_1.create_file_object() 
        self.column_to_read_from = 'C'
        self.column_to_read_to = 'C'
        self.columns_to_compare = ['E', 'F', 'H']
        self.columns_to_copy = ['B','F']
        self.column_insertion = 'E'

    def _apply_method_to_test(self):
        self.file1 = self.file_data_compare_1.file_object
        self.file2 = self.file_data_compare_2.file_object
        self.controller = TwoFilesController(self.file_object_from, self.file1,
                                             self.tab_name_from, self.file_data_compare_1.tab_name,
                                             self.column_to_read_from, self.column_to_read_to)
        self.controller.copy_columns_in_a_tab_differently_sorted(self.columns_to_copy, self.column_insertion)

    def _compare_new_columns(self):
        self.assert_object = AssertIdentical(self.file_data_compare_1, self.file_data_compare_2) 
        columns_to_compare_indexes = MapIndexLetter.get_list_of_columns_indexes(self.columns_to_compare)
        for column_index in columns_to_compare_indexes:
            self.assert_object.verify_columns_identical(column_index, column_index)

    def _restore_file_state_and_save(self):
        tab_modified = self.file1.writebook[self.file_data_compare_1.tab_name]        
        tab_modified.delete_cols(column_index_from_string(self.column_insertion), len(self.columns_to_copy))
        TabUpdateFormula(ColumnDelete(['E','F'])).update_cells_formulas(tab_modified)
        self.file1.save_file()


class TestString(TestCase, String):
    
    def test_transform_string_in_binary(self):
        self.assertEqual(self.transform_string_in_binary('rrr','rr', 'rrr'), 1)
        self.assertEqual(self.transform_string_in_binary('rr','rrr'), 0)
        self.assertEqual(self.transform_string_in_binary('','rrr'), 0)

    def test_set_answer_in_group(self): 
        map_groups_to_answers = {"group1":['2','5','6'], "group2":['7','8','9'], "group3":['1','3','4'], "group4":['10']}
        map_answers_to_groups = Dictionary.reverse_dictionary(map_groups_to_answers) 
        
        self.assertEqual(self.set_answer_in_group('1', map_answers_to_groups), "group3")
        self.assertEqual(self.set_answer_in_group('9', map_answers_to_groups), "group2")

    def test_clean_string_from_spaces(self):
        strings = ['tony', 'tony ', ' tony', ' tony ', 'tony  ', '  tony']
        for string in strings:
            cleaned_string = self.clean_string_from_spaces(string)
            self.assertEqual(cleaned_string, 'tony')
            
    def test_convert_time_in_minutes(self):
        map_durations_to_minutes = {"2 jours 2 heures": '3000,0', "1 heure 25 min": '85,0', "16 min 35 s": '16,58'}
        for duration in map_durations_to_minutes.keys():
            duration_in_min = self.convert_time_in_minutes(duration)
            duration_in_min_expected = map_durations_to_minutes[duration]
            self.assertEqual(duration_in_min, duration_in_min_expected)

    def test_get_columns_from(self):
        string = "C-E,H,J-L"
        expected_list = ['C','D','E','H','J','K','L']
        self.assertListEqual(self.get_columns_from(string), expected_list)

    def test_get_range_letter(self):
        string = 'D-H'
        expected_list = ['D','E','F','G','H']
        self.assertListEqual(self.get_range_letter(string), expected_list)


class TestTabUpdate(TestCase, TabUpdateFormula):

    def test_update_formula_when_insert_one_line(self):
        modification_object = LineInsert(['2'])
        updated_formula = modification_object._update_a_cell("SI(J10+K$1+L$3)")
        self.assertEqual(updated_formula, "SI(J11+K$1+L$4)")

    def test_update_formula_when_delete_one_line(self):
        modification_object = LineDelete(['11'])
        updated_formula = modification_object._update_a_cell("SI(J12+K$1+L$3)")
        self.assertEqual(updated_formula, "SI(J11+K$1+L$3)")

    def test_update_formula_when_delete_one_line_greater_than_all_line_numbers(self):
        modification_object = LineDelete(['13'])
        updated_formula = modification_object._update_a_cell("SI(J12+K$1+L$3)")
        self.assertEqual(updated_formula, "SI(J12+K$1+L$3)")

    def test_update_formula_when_insert_one_column(self):
        modification_object = ColumnInsert(['C'])
        updated_formula = modification_object._update_a_cell("SI(J10+K$1+L$3)")
        self.assertEqual(updated_formula, "SI(K10+L$1+M$3)")

    def test_update_formula_when_delete_one_column(self):
        modification_object = ColumnDelete(['C'])
        updated_formula = modification_object._update_a_cell("SI(J10+K$1+L$3)")
        self.assertEqual(updated_formula, "SI(I10+J$1+K$3)")

    def test_update_formula_when_insert_multiple_lines(self):
        modification_object = LineInsert(['2', '5'])
        updated_formula = modification_object._update_a_cell("SI(J10+K$1+L$3)") 
        self.assertEqual(updated_formula, "SI(J12+K$1+L$4)")
 
    def test_update_formula_when_insert_multiple_columns(self):
        modification_object = ColumnInsert(['C', 'D','E', 'F'])
        updated_formula = modification_object._update_a_cell("SI(D10+E$1)") 
        self.assertEqual(updated_formula, "SI(H10+I$1)")


class AssertIdentical(TestCase):
    """Check if two files are the same, or have the same tabs or columns"""

    def __init__(self, file_data1, file_data2, *args, **kwargs):
        super().__init__(*args, **kwargs) 
        self.file_data1 = file_data1
        self.file_data2 = file_data2
        self.file_object1 = self.file_data1.file_object
        self.file_object2 = self.file_data2.file_object
        self._initialize_tabs_attributes()

    def _initialize_tabs_attributes(self): 
        try:  
            self.tab1 = self.file_object1.writebook[self.file_data1.tab_name] 
            self.tab2 = self.file_object2.writebook[self.file_data2.tab_name]  
        except KeyError: 
            self.tab1 = None
            self.tab2 = None

    def verify_files_identical(self):  
        self.assertEqual(self.file_object1.sheets_name, self.file_object2.sheets_name)

        for tab_name in self.file_object1.sheets_name: 
            self.tab1 = self.file_object1.writebook[tab_name]
            self.tab2 = self.file_object2.writebook[tab_name]
            self.verify_tabs_identical()

    def verify_tabs_identical(self):   
        self.assertEqual(self.tab1.max_row, self.tab2.max_row)
        self.assertEqual(self.tab1.max_column, self.tab2.max_column)

        for i in range(1,self.tab1.max_row+1):
            for j in range(1,self.tab1.max_column+1):
                self.assertEqual(self.tab1.cell(i,j).value,self.tab2.cell(i,j).value)

    def verify_columns_identical(self, column1, column2):  
        self.assertEqual(self.tab1.max_row, self.tab2.max_row) 
            
        for i in range(2, self.tab1.max_row + 1): 
            self.assertEqual(self.tab1.cell(i, column1).value,self.tab2.cell(i, column2).value)

    def verify_cells_identical(self, *cells):
        cells_indexes = MapIndexLetter.get_list_of_cells_coordinates(cells)
        for indexes in cells_indexes:
            self.assertEqual(self.tab1.cell(indexes[0], indexes[1]).value, self.tab2.cell(indexes[0], indexes[1]).value)

class FileData():
    """Data objects containing file informations for test"""
    def __init__(self, name_file, tab_name=None, path='fichiers_xls/'):
        self.name_file = name_file
        self.tab_name = tab_name
        self.path = path
        self.initialize_file_object()

    def initialize_file_object(self):
        try:
            self.create_file_object()
        except FileNotFoundError:
            self.file_object = None            

    def create_file_object(self):
        self.file_object = File(self.name_file, path=self.path)


class MethodData():
    """Data object containing method arguments for test"""
    def __init__(self, method_name, *args, **kwargs):
        self.method_name = method_name 
        self.args = args
        self.kwargs = kwargs
    

if __name__== "__main__":
    main()