import re
import openpyxl
import typer

from openpyxl.utils import get_column_interval,\
                           column_index_from_string,\
                           get_column_letter,\
                           coordinate_to_tuple
from copy import copy 
from time import time
from model.model_factorise import Cell


class DisplayRunningInfos():
    def __init__(self):
        self.method_name = None
        self.current_running_part = None
        self.list_of_running_parts = None
        self.start_time = None
        self.elapsed_time = 0.
        self.remaining_time = 0.
        self.completion_percentage = 0.
        self.time_title = ''  

    def display_running_infos(self):
        """ 
        Print the percentage of completion, elapsed time and remaining time when executing a method 
        """
        self._compute_time_and_completion_data()
        self._display_completion_data()
        self._display_times_data() 

    def _compute_time_and_completion_data(self):
        self.completion_percentage = round((self.list_of_running_parts.index(self.current_running_part) + 1)/len(self.list_of_running_parts) * 100,2)
        self.elapsed_time = time() - self.start_time 
        self.remaining_time = (100 - self.completion_percentage) * self.elapsed_time / self.completion_percentage 

    def _display_completion_data(self):
        print(f'\n---------------Currently running method {self.method_name}---------------\n')
        print(f'Percentage of completion : {self.completion_percentage}%')
        print(f'{self.current_running_part} is finished')

    def _display_times_data(self):
        self._update_time_title('Elapsed time')
        self._display_time_in_adapted_unit(self.elapsed_time)
        self._update_time_title('Estimated remaining time')
        self._display_time_in_adapted_unit(self.remaining_time)

    def _display_time_in_adapted_unit(self, duration):
        """
        Print a duration in sec if it is less than 60s, in minutes if it is between 60s and 3600s, in hours otherwise.
        
        Inputs:
            - duration (float) 
        """
        if duration < 60:
            print(f'{self.time_title} : {round(duration, 2)} sec')
        elif 60 <= duration < 3600: 
            duration /= 60 
            print(f'{self.time_title} : {round(duration, 2)} min')
        else: 
            duration /= 3600 
            print(f'{self.time_title} : {round(duration, 2)} h')

    def _update_time_title(self, time_title):
        self.time_title = time_title

    def _update_display_infos(self, method_name, current_running_part, list_of_running_parts):
        self.method_name = method_name
        self.current_running_part = current_running_part
        self.list_of_running_parts = list_of_running_parts


class TabsCopy():
    """Make copys from a tab to a new tab"""

    def __init__(self, tab_from=None, tab_to=None):
        self.tab_from = tab_from
        self.tab_to = tab_to

    def _choose_the_tab_to_write_in(self, tab):
        self.tab_to = tab

    def _choose_the_tab_to_read(self, tab):
        self.tab_from = tab
        
    def copy_paste_line(self, line_from, line_to):#, values_only=False):
            """
            Fonction qui prend une ligne de la feuille et qui la copie dans un autre onglet.

            Inputs :  
                - line_from : ligne de l'onglet d'origine.  
                - line_to : la ligne où il faut coller dans l'onglet à modifier.
            """

            """ 
            ON REPRENDRA CETTE VERSION QUAND J ATTAQUERAIS LES HISTOIRES DE VALUES ONLY

            # Cas où on ne copie que les valeurs, cell est un str
            if values_only:
                column_index = 1 
                for column in onglet_from.iter_cols(min_row=row_from, max_row=row_from, min_col=0, max_col=onglet_from.max_column, values_only=values_only):
                    for cell in column: 
                        if cell:
                            onglet_to.cell(row_to, column_index).value = cell
                        column_index += 1
                        
            # Cas où on copie les formules, cell est un objet
            else:
                for column in onglet_from.iter_cols(min_row=row_from, max_row=row_from, min_col=0, max_col=onglet_from.max_column, values_only=values_only):
                    for cell in column: 
                        if cell.value != "":
                            onglet_to.cell(row_to, cell.column).value = cell.value """

            for column_index in range(1, self.tab_from.max_column + 1): 
                self.copy_of_a_cell(Cell(line_from, column_index), Cell(line_to, column_index)) 

    def copy_paste_column(self, column_from, column_to):
            """
            Fonction qui prend une colonne de la feuille et qui la copie dans un autre onglet.
            """ 
            for line_index in range(1, self.tab_from.max_row + 1): 
                self.copy_of_a_cell(Cell(line_index, column_from), Cell(line_index, column_to))  

    def copy_paste_multiple_columns(self, columns_int_list):
        for column in columns_int_list:
            self.copy_paste_column(column, column)

    def copy_tag_and_values_of_a_column_at_tab_bottom(self, column_index):
        """
        Fonction qui prend une colonne de valeurs nommée C et la copie à la fin de la colonne 2 d'un onglet (à partir de la première cellule vide).
        Cette fonction écrit également dans la colonne 1 les valeurs de la première cellule de C.
        """
        max_row = self.tab_to.max_row + 1
        for line_index in range(2, self.tab_from.max_row + 1):
            self.copy_of_a_cell(Cell(1, column_index), Cell(line_index - 2 + max_row, 1)) 
            self.copy_of_a_cell(Cell(line_index, column_index), Cell(line_index - 2 + max_row, 2)) 

    def add_line_at_bottom(self, line_from):
            """
            Fonction qui copie une ligne spécifique de la feuille à la fin d'un autre onglet.

            Input : 
                - line_from : ligne de l'onglet d'origine. 
            """ 
            self.copy_paste_line(line_from, self.tab_to.max_row + 1) 

    def copy_old_file_tab_in_new_file_tab(self): 
        for i in range(1, self.tab_from.max_row + 1):
            for j in range(1, self.tab_from.max_column + 1):  
                self.copy_of_a_cell(Cell(i,j), Cell(i,j))

    def copy_old_file_tab_in_new_file_tab_at_bottom(self):
        for line in range(2, self.tab_from.max_row + 1):
            self.add_line_at_bottom(line)

    def deep_copy_of_a_cell(self, cell_from, cell_to):   
        self.tab_to.cell(cell_to.line_index, cell_to.column_index).value = self.tab_from.cell(cell_from.line_index, cell_from.column_index).value  
        self.tab_to.cell(cell_to.line_index, cell_to.column_index).fill = copy(self.tab_from.cell(cell_from.line_index, cell_from.column_index).fill)
        self.tab_to.cell(cell_to.line_index, cell_to.column_index).font = copy(self.tab_from.cell(cell_from.line_index, cell_from.column_index).font) 
        self.tab_to.cell(cell_to.line_index, cell_to.column_index).border = copy(self.tab_from.cell(cell_from.line_index, cell_from.column_index).border) 
        self.tab_to.cell(cell_to.line_index, cell_to.column_index).alignment = copy(self.tab_from.cell(cell_from.line_index, cell_from.column_index).alignment)     

    def copy_of_a_cell(self, cell_from, cell_to):   
        self.tab_to.cell(cell_to.line_index, cell_to.column_index).value = self.tab_from.cell(cell_from.line_index, cell_from.column_index).value
    
    def deep_copy_multiple_cells(self, cells_list):
        for cell in cells_list: 
            self.deep_copy_of_a_cell(Cell(cell[0],cell[1]), Cell(cell[0],cell[1])) 

    def deep_copy_of_a_tab(self):
        """
        Fonction qui copie une page sur une autre. La copie est totale : valeur, couleur, cellules fusionnées
        """
        for i in range(1, self.tab_from.max_row + 1):
            for j in range(1,self.tab_from.max_column + 1): 
                self.deep_copy_of_a_cell(Cell(i,j), Cell(i,j))  
 
        self.merge_cells_as_in_tab_from()

    def merge_cells_as_in_tab_from(self):
        for merged_range in self.tab_from.merged_cells.ranges:  
            start_column, start_row, end_column, end_row = merged_range.bounds  
            self.tab_to.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)


class MapIndexLetter():
    """Handle methods to transform cell, columns in indexes"""

    @staticmethod
    def get_list_of_cells_coordinates(cells): 
        cells_list = []
        for cell in cells: 
            cells_list.append(coordinate_to_tuple(cell)) 
        return cells_list

    @staticmethod
    def get_list_of_columns_indexes(columns_letters_list):
        columns_int_list = []
        for column in columns_letters_list: 
            columns_int_list.append(column_index_from_string(column)) 
        return columns_int_list 
    
    @staticmethod
    def get_cells_indexes_of_one_line_and_some_columns(line_index, columns_indexes):
        return  [(line_index, column_index) for column_index in columns_indexes]
    
    @staticmethod
    def get_list_of_consecutive_column_letters(column_begin_index, number_of_columns):
        return [get_column_letter(column_begin_index + i ) for i in range(number_of_columns)]
    

class RegularExpression():

    @staticmethod
    def _split_a(formula):
        """Returns a list containing expressions like C1, D$15 from a formula linking them"""
        return re.split(r'(\b[A-Za-z-$]+\d+\b)', formula)  

    @staticmethod
    def _is_string_a_cell_expression(string):
        """Verifies if the string has the shape A1 or C$5 (LetterNumber or Letter$Number)"""
        return re.fullmatch(r'\b[A-Za-z-$]+\d+\b', string)
    
    @staticmethod
    def _split_cell_expression_from(cell_expression):
        """Get list of the form ['C','15'] or ['C$', '15'] from C15 or C$15"""
        return re.split(r'(\d+)', cell_expression)[:-1]
    
    @staticmethod
    def _recover_cell_expression_from(splitted_cell_expression):
        """ Get 'C5' from ['C', '5']"""
        return ''.join(splitted_cell_expression)
    
    @staticmethod
    def _recover_cell_formula_from(cells_expression_list):
        """ Get 'C5+D$6' from ['C5', '+', 'D$6']"""
        return ''.join(cells_expression_list)
    
    @staticmethod
    def get_word_jpg_name_file(string):
        """Get Mot from Motnb_.jpg"""
        return re.sub(r'([A-Z-a-z]+)\d+_[A-Z-a-z].jpg', r'\1', string)
    
    @staticmethod
    def _is_string_a_xlsx_file(string): 
        return re.fullmatch(r'.+.xlsx', string)
    
    @staticmethod
    def map_time_unity_to_value_from(time_string):
        """Get a dictionary corresponding to a string of the form 1 jour 10min 5s or '10min 5s'. 
        Example of return value : {'jour':1, 'min':10, 's': 5] from 1 jour 10 min 5 s. """
        time_values = re.findall(r'\d+', time_string)
        unities = re.findall(r'[A-Za-z]+', time_string)
        return dict(zip(unities, time_values))
    

class TabUpdateFormula():
    """Update cells formula of a tab after columns/lines are inserted/deleted. Modification_object 
    is one of the following interfaces."""
    def __init__(self, modification_object=None): 
        self.modification_object = modification_object

    def choose_modifications_to_apply(self, modification_object):
        self.modification_object = modification_object

    def _get_cell_value(self, tab, cell):
        return tab.cell(cell.line_index, cell.column_index).value 

    def update_cells_formulas(self, tab):
        """
        Fonction qui met à jour les formules d'une feuille entière  
        """
        for line_index in range(1, tab.max_row + 1):
            for column_index in range(1, tab.max_column + 1): 
                cell_value = self._get_cell_value(tab, Cell(line_index, column_index))
                if self._is_cell_value_a_formula(cell_value): 
                    tab.cell(line_index, column_index).value  = self.modification_object._update_a_cell(cell_value)
    
    @staticmethod
    def _is_cell_value_a_formula(cell_value):
        return isinstance(cell_value, str) and cell_value.startswith('=')        
    

class ColumnInsert(RegularExpression):
    """Interface aiming at modifying a cell formula after column(s) insertion(s)"""
    def __init__(self, columns_inserted):   
        self.columns_inserted = columns_inserted

    def _update_a_cell(self, formula):
        """
        Fonction qui va mettre à jour la formule d'une cellule suite à un ajout ou une suppression de colonne/ligne. 
        A cell formula is of the form '=C5+D$6'
        """
        
        parts_of_formula = self._split_a(formula) 
        for index in range(len(parts_of_formula)): 
            if self._is_string_a_cell_expression(parts_of_formula[index]): 
                parts_of_formula[index] = self._update_one_cell_expression(parts_of_formula[index])
        return self._recover_cell_formula_from(parts_of_formula) 

    def _update_one_cell_expression(self, cell_expression):
        for column_inserted in self.columns_inserted:
            cell_expression = self._update_one_cell_expression_after_one_modification(cell_expression, column_inserted)
        return cell_expression
    
    def _update_one_cell_expression_after_one_modification(self, cell_expression, column_inserted):
        """A cell expression if of type C5 or C$5"""

        splitted_cell_expression = self._split_cell_expression_from(cell_expression)
        column_of_cell_expression = splitted_cell_expression[0]
        try: 
            if self._does_column_needs_to_be_updated(column_of_cell_expression, column_inserted):
                splitted_cell_expression[0] = get_column_letter(column_index_from_string(column_of_cell_expression) + 1)
        except ValueError:
            #The column letter may be of the form C$ in excel formula so we have to clean $ to get index and readd it after 
            if self._does_column_needs_to_be_updated(column_of_cell_expression[:-1], column_inserted):
                splitted_cell_expression[0] = get_column_letter(column_index_from_string(column_of_cell_expression[:-1]) + 1) + "$"
 
        return self._recover_cell_expression_from(splitted_cell_expression)
    
    @staticmethod
    def _does_column_needs_to_be_updated(column_of_cell_expression, column_inserted):
        return column_index_from_string(column_of_cell_expression) > column_index_from_string(column_inserted)


class ColumnDelete(RegularExpression):
    def __init__(self, columns_deleted): 
        self.columns_deleted = columns_deleted

    def _update_a_cell(self, formula):
        """
        Fonction qui va mettre à jour la formule d'une cellule suite à un ajout ou une suppression de colonne/ligne. 
        A cell formula is of the form '=C5+D$6'
        """
        
        parts_of_formula = self._split_a(formula) 
        for index in range(len(parts_of_formula)): 
            if self._is_string_a_cell_expression(parts_of_formula[index]): 
                parts_of_formula[index] = self._update_one_cell_expression(parts_of_formula[index])
        return self._recover_cell_formula_from(parts_of_formula) 

    def _update_one_cell_expression(self, cell_expression):
        for column_deleted in self.columns_deleted:
            cell_expression = self._update_one_cell_expression_after_one_modification(cell_expression, column_deleted)
        return cell_expression

    def _update_one_cell_expression_after_one_modification(self, cell_expression, column_deleted):
        """A cell expression if of type C5 or C$5"""
        splitted_cell_expression = self._split_cell_expression_from(cell_expression)
        column_of_cell_expression = splitted_cell_expression[0]
        try: 
            if self._does_column_needs_to_be_updated(column_of_cell_expression, column_deleted):
                splitted_cell_expression[0] = get_column_letter(column_index_from_string(column_of_cell_expression) - 1)
        except ValueError:
            #The column letter may be of the form C$ in excel formula so we have to clean $ to get index and readd it after 
            if self._does_column_needs_to_be_updated(column_of_cell_expression[:-1], column_deleted):
                splitted_cell_expression[0] = get_column_letter(column_index_from_string(column_of_cell_expression[:-1]) - 1) + "$"
 
        return self._recover_cell_expression_from(splitted_cell_expression)
    
    @staticmethod
    def _does_column_needs_to_be_updated(column_of_cell_expression, column_deleted):
        return column_index_from_string(column_of_cell_expression) > column_index_from_string(column_deleted)
    

class LineInsert(RegularExpression):
    def __init__(self, lines_inserted):
        self.lines_inserted = lines_inserted

    def _update_a_cell(self, formula):
        """
        Fonction qui va mettre à jour la formule d'une cellule suite à un ajout ou une suppression de colonne/ligne. 
        A cell formula is of the form '=C5+D$6'
        """
        
        parts_of_formula = self._split_a(formula) 
        for index in range(len(parts_of_formula)): 
            if self._is_string_a_cell_expression(parts_of_formula[index]): 
                parts_of_formula[index] = self._update_one_cell_expression(parts_of_formula[index])
        return self._recover_cell_formula_from(parts_of_formula) 
    
    def _update_one_cell_expression(self, cell_expression):
        for line_inserted in self.lines_inserted:
            cell_expression = self._update_one_cell_expression_after_one_modification(cell_expression, int(line_inserted))
        return cell_expression

    def _update_one_cell_expression_after_one_modification(self, cell_expression, line_inserted):
        """A cell expression if of type C5 or C$5"""
        splitted_cell_expression = self._split_cell_expression_from(cell_expression)
        line_of_cell_expression = int(splitted_cell_expression[1])
        if self._does_line_letter_needs_to_be_updated(line_of_cell_expression, line_inserted): 
            splitted_cell_expression[1] = str(line_of_cell_expression + 1)
        return self._recover_cell_expression_from(splitted_cell_expression)
    
    @staticmethod
    def _does_line_letter_needs_to_be_updated(line_letter, line_deleted):
        return line_letter > line_deleted


class LineDelete(RegularExpression):
    def __init__(self, lines_deleted):
        self.lines_deleted = lines_deleted

    def _update_a_cell(self, formula):
        """
        Fonction qui va mettre à jour la formule d'une cellule suite à un ajout ou une suppression de colonne/ligne. 
        A cell formula is of the form '=C5+D$6'
        """
        
        parts_of_formula = self._split_a(formula) 
        for index in range(len(parts_of_formula)): 
            if self._is_string_a_cell_expression(parts_of_formula[index]): 
                parts_of_formula[index] = self._update_one_cell_expression(parts_of_formula[index])
        return self._recover_cell_formula_from(parts_of_formula) 
    
    def _update_one_cell_expression(self, cell_expression):
        for line_deleted in self.lines_deleted:
            cell_expression = self._update_one_cell_expression_after_one_modification(cell_expression, int(line_deleted))
        return cell_expression

    def _update_one_cell_expression_after_one_modification(self, cell_expression, line_deleted):
        """A cell expression if of type C5 or C$5"""
        splitted_cell_expression = self._split_cell_expression_from(cell_expression)
        line_of_cell_expression = int(splitted_cell_expression[1])
        if self._does_line_letter_needs_to_be_updated(line_of_cell_expression, line_deleted): 
            splitted_cell_expression[1] = str(line_of_cell_expression - 1)
        return self._recover_cell_expression_from(splitted_cell_expression)
    
    @staticmethod
    def _does_line_letter_needs_to_be_updated(line_letter, line_deleted):
        return line_letter > line_deleted
    

class String(): 
    """Handle useful methods carrying on strings"""

    @staticmethod
    def clean_string_from_spaces(string):
        """
        Fonction qui prend une chaîne de caractère et qui élimine tous les espaces de début et de fin.
        Fonction qui nettoie également les espaces insécables \xa0 par un espace régulier. 
        """
        return string.strip().replace('\xa0', ' ')
    
    @classmethod
    def get_columns_from_several(cls, *strings):
        columns_list = []
        for string in strings: 
            columns_list.append(cls.get_columns_from(string)) 
        return columns_list
    
    @classmethod
    def get_columns_from(cls, string):
        """
        Fonction qui prend en entrée une chaîne de caractères de la forme "C-E,H,J" et qui retourne une liste de colonnes 
        ['C','D','E','H','J']. 
        """
        substrings = string.split(',')
        columns_list = []
        for substring in substrings:
            columns_list = cls._add_to_list_columns_of_substring(columns_list, substring)
        return columns_list
    
    @classmethod
    def _add_to_list_columns_of_substring(cls, columns_list, substring):
        if '-' in substring:
            columns_list += cls.get_range_letter(substring)
        else:
            columns_list.append(substring)
        return columns_list

    @staticmethod
    def get_range_letter(string):
        """
        Fonction qui prend une chaîne de la forme "D-G" et qui retourne la liste des lettres entre elles. 
        """
        L = string.split('-')
        return get_column_interval(L[0], L[-1])
    
    @staticmethod
    def transform_string_in_binary(string, *args):
        """
        Fonction qui prend un str et qui le transforme en 0 ou 1

        Inputs : args : des chaînes de caractère devant renvoyer 1 
        Outputs : bool : 0 ou 1.
        """
        binary = 0 
        if string in args:
            binary = 1
        return binary
    
    @classmethod
    def convert_time_in_minutes(cls, time_string):
        """
        Function which takes a str of the form "10 jour 5 heures" and return a string giving the conversion in unity.

        Output : str
        """
        map_time_unity_to_value = RegularExpression.map_time_unity_to_value_from(time_string)

        duration = 0
        for unity in map_time_unity_to_value.keys():
            duration = cls._add_time_value_to_duration(unity, map_time_unity_to_value, duration)
    
        conversion = str(duration).replace('.',',')
        return conversion
    
    @classmethod
    def _add_time_value_to_duration(cls, unity, map_time_unity_to_value, duration):
        time_value = float(map_time_unity_to_value[unity])
        if unity in ["jour","jours"]:
            duration += 24 * 60 * time_value
        elif unity in ['heure', 'heures']:
            duration += time_value * 60
        elif unity == 'min':
            duration += time_value
        else:
            duration += round(time_value/60, 2)
        return duration

    @staticmethod
    def set_answer_in_group(answer, map_answers_to_groups):
        if answer in map_answers_to_groups.keys():
            return map_answers_to_groups[answer]
        else:
            return ""
    
    def del_extension(string):
        """Fonction 
            - qui enlève l'extension d'un nom de fichier si le nom ne contient pas de date
            - qui ne garde que la partie avant _date_ pour un fichier nommé test_date_****-**-**.xlsx. 
            - qui sert à la sauvegarde et permet ainsi d'éviter des noms à rallonge.
        """
        position = string.find('_date_')
        if position == -1: 
           position = string.find('.xlsx')

        return string[:position]


class Dictionary():
    @staticmethod
    def reverse_dictionary(dictionary):
        """
        Function taking a dictionary of the form {'group1':['a','b'],'group2':['c','d','e']} and returning the dictionary
        {'a':'group1','b':'group1','c':'group2','d':'group2','e':'group2'}
        """
        reverse_dictionary= {}
        for key, value in dictionary.items():
            for reponse in value:
                reverse_dictionary[reponse] = key
        return reverse_dictionary
    
    
class Workbook():
    def create_empty_workbook():
        workbook = openpyxl.Workbook()
        del workbook[workbook.active.title]
        return workbook

    def get_first_tab_of_new_workbook():
        workbook = openpyxl.Workbook()
        return workbook[workbook.active.title] 
    

class InputStore():
    """Store inputs from the user when using functinos"""
    def __init__(self, args, message):
        self.message = message
        self.args = args

    def ask_argument_until_none(self):
        if not self.args:
            self.args = []
            self._store_argument_if_not_none()
        return self.args
    
    def _store_argument_if_not_none(self):
        while True:
            user_input = typer.prompt(self.message, default="")
            if not user_input:
                break
            self.args.append(user_input)


class MapStore():
    """Create a dictionary containing inputs for keys and values"""
    def __init__(self, message_key, message_value):
        self.message_key = message_key
        self.message_value = message_value 
        self.mapping = {}

    def create_mapping(self):
        while True: 
            key = typer.prompt(self.message_key, default = "")
            if not key:
                break 
            value = typer.prompt(self.message_value, default = "")
            self._assign_splitted_value_if_contains_comma(key, value)

    def _assign_splitted_value_if_contains_comma(self, key, value):
        if ',' in value:
            self.mapping[key] = value.split(',')
        else:
            self.mapping[key] = value   
    