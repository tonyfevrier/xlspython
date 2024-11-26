from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_interval, column_index_from_string, get_column_letter
from copy import copy 
from time import time, sleep 
from model_factorise import Cell

import typer 
import yagmail 
import re


class UtilsForFile():
    def copy_paste_line(self,onglet_from,row_from, onglet_to, row_to, values_only=False):
        """
        Fonction qui prend une ligne de la feuille et qui la copie dans un autre onglet.

        Inputs : 
            - onglet_from : onglet d'où on copie
            - row_from : ligne de l'onglet d'origine.
            - onglet_to : onglet où coller.
            - row_to : la ligne où il faut coller dans l'onglet à modifier.

        Exemple d'utilisation : 
      
            file = File('dataset.xlsx')
            file.copy_paste_line('onglet1', 1, 'onglet2', 1)
        """

        column_index = 1 

        # Cas où on ne copie que les valeurs, cell est un str
        if values_only:
            for column in onglet_from.iter_cols(min_row=row_from, max_row=row_from, min_col=0, max_col=onglet_from.max_column, values_only=values_only):
                for cell in column: 
                    if cell:
                        onglet_to.cell(row_to, column_index).value = cell
                    column_index += 1
                    
        # Cas où on copie les formules, cell est un objet
        else:
            for column in onglet_from.iter_cols(min_row=row_from, max_row=row_from, min_col=0, max_col=onglet_from.max_column, values_only=values_only):
                for cell in column: 
                    if cell.value != "":
                        onglet_to.cell(row_to, cell.column).value = cell.value


        #for j in range(1, onglet_from.max_column + 1): 
        #    onglet_to.cell(row_to,j).value = onglet_from.cell(row_from, j).value 
 

    def copy_paste_column(self, onglet_from, column_from, onglet_to, column_to):
        """
        Fonction qui prend une colonne de la feuille et qui la copie dans un autre onglet.
        """

        for i in range(1, onglet_from.max_row + 1): 
            onglet_to.cell(i,column_to).value = onglet_from.cell(i,column_from).value 

    def deep_copy_of_a_sheet(self, sheet_from, sheet_to):
        """
        Fonction qui copie une page sur une autre. La copie est totale : valeur, couleur, cellules fusionnées
        """
        for i in range(1,sheet_from.max_row+1):
                for j in range(1,sheet_from.max_column+1): 
                    # Récupérer les informations sur la cellule fusionnée  
                    sheet_to.cell(i,j).value = sheet_from.cell(i,j).value  
                    sheet_to.cell(i,j).fill = copy(sheet_from.cell(i,j).fill)
                    sheet_to.cell(i,j).font = copy(sheet_from.cell(i,j).font) 
                    sheet_to.cell(i,j).border = copy(sheet_from.cell(i,j).border) 
                    sheet_to.cell(i,j).alignment = copy(sheet_from.cell(i,j).alignment)  

        #On parcourt le dictionnaire des cellules fusionnées et on fusionne celles de sheet to correspondante:
        for merged_range in sheet_from.merged_cells.ranges:  
            start_column, start_row, end_column, end_row = merged_range.bounds 
            # Fusionner les cellules correspondantes dans la feuille de calcul destination
            sheet_to.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)

    def add_line_at_bottom(self, onglet_from, row_from, onglet_to, values_only=False):
        """
        Fonction qui copie une ligne spécifique de la feuille à la fin d'un autre onglet.

        Input : 
            - row_origin : ligne de l'onglet d'origine.
            - onglet : l'onglet à modifier où on copie la ligne.

        Exemple d'utilisation : 
     
            file = File('dataset.xlsx')
            file.copy_paste_line('onglet1', 1, 'onglet2')
        """ 
        self.copy_paste_line(onglet_from, row_from, onglet_to, onglet_to.max_row + 1, values_only=values_only)  

    def copy_column_tags_and_values_at_bottom(self,import_sheet, column, target_sheet):
        """
        Fonction qui prend une colonne de valeurs nommée C et la copie à la fin de la colonne 2 d'un onglet (à partir de la première cellule vide).
        Cette fonction écrit également dans la colonne 1 les valeurs de la première cellule de C.
        """
        maxrow = target_sheet.max_row + 1
        for line in range(2, import_sheet.max_row + 1):
            target_sheet.cell(line - 2 + maxrow, 1).value = import_sheet.cell(1, column).value
            target_sheet.cell(line - 2 + maxrow, 2).value = import_sheet.cell(line, column).value

    def envoi_mail(self,adresse, file, expeditor, password, object, message): 
        """
        Fonction qui envoie un mail avec une pièce jointe.

        Inputs : 
            - adresse (str) : mail à qui on envoie.
            - file (str) : fichier à joindre au mail.
            - expeditor (str) : mail de la personne qui envoie.
            - password (str) : mot de passe pour permettre d'envoyer un mail d'un serveur distant.
            - object (str) : l'objet du mail.
            - message (str) : le contenu du mail. 
        """
        yag = yagmail.SMTP(expeditor, password) 
        piece_jointe = file

        yag.send(to=adresse, subject=object, contents=message, attachments = piece_jointe)
 


    def column_security(self, sheet, column):
        """
        Fonction qui prend une colonne et regarde si la colonne est vide.
        Input : column
        Output : True si elle ne contient rien, False sinon
        """
        bool = True
        for i in range(1,sheet.max_row+1): 
            if sheet.cell(i,column).value is not None:
                bool = False
                break
        return bool 
    
    @staticmethod
    def color_line(sheet, color, row_number):
        """
        Fonction qui colore une ligne spécifique

        Input :
            - color : une couleur indiquée en haxadécimal par l'utilisateur.
            - row_number : le numéro de la ligne à colorer

        Exemple d'utilisation : 
    
            sheet = Sheet('dataset.xlsx','onglet1')
            sheet.color_line('#FF0000', 3) 
        
        """

        for j in range(1, sheet.max_column + 1):
            sheet.cell(row_number,j).fill = PatternFill(fill_type = 'solid', start_color = color)

    def create_dico_from_columns(self, sheet, column_keys:int, column_values:int, first_line, last_line):
        """
        Function returning a dictionnary whose keys are elements of a column
          if they are not empty and values are elements of an other column
        
        Inputs :
            column_keys : column whose elements are the keys of the dictionnary.
            column_values : same with values
            first_line : the line we begin to read the file

        Output : 
            dico : dictionary. 
        """

        dico = {}
        for i in range(first_line,last_line):
            key = sheet.cell(i,column_keys).value 
            
            if key != "":
                dico[key] = sheet.cell(i,column_values).value
        return dico
    
    def create_dico_to_store_multiple_answers_of_participants(self, sheet, column_read, column_store, line_beggining):
        """
        Fonction qui parcourt les identifiants d'un fichier et qui crée un dictionnaire contenant le nombre de réponses de chaque participant
        et les valeurs d'une même donnée lors des différentes réponses.

        Inputs : 
            - column_read (str) : la colonne avec les identifiants des participants.
            - column_store (str) : lettre de la colonne contenant la donnée qu'on veut stocker.
            - line_beggining (int) : ligne où débute la recherche. 
        """
        dico = {}
        for line in range(line_beggining,sheet.max_row + 1):
            identifier = sheet.cell(line,column_read).value
            value_to_store = sheet.cell(line,column_store).value 
            if identifier in dico.keys():
                dico[identifier][0] += 1
                dico[identifier][1].append(value_to_store)
            else: 
                dico[identifier] = [1, [value_to_store]]
        return dico
    
    def create_newsheet_storing_multiple_answers(self,storesheet,dico):
        """
        Fonction qui prend un dico issu de la fonction create_dico_to_store_multiple_answers_of_participants et qui crée une feuille 
        présentant les participants ayant répondu plusieurs fois et les valeurs de la donnée stockée lors des différentes réponses.

        Inputs:
            - storesheet (openpyxl sheet) : feuille de stockage.
            - dico (dict) : issu de create_dico_to_store_multiple_answers_of_participants.
        """
        storesheet.cell(1,1).value = 'Identifiers'
        
        for key,list in dico.items(): 
            if list[0] >= 2:
                firstline = storesheet.max_row + 1
                storesheet.cell(firstline, 1).value = key
                for i in range(len(list[1])):
                    storesheet.cell(firstline, i+2).value = list[1][i]

    def updateCellFormulas(self,sheet,insert, rowOrColumn, modifications):
        """
        Fonction qui met à jour les formules d'une feuille entière 

        Inputs : 
            - sheet (obj): la feuille sur laquelle on agit.
            - insert (bool) : True si les modifications sont toutes des insertions, False si ce sont toutes des suppressions.
            - rowOrColumn (str): 'row' ou 'column' suivant que la série d'opérations effectuées porte sur ligne ou colonne.
            - modifications (list[str]) : liste de str donnant les modifications. Si on a inséré 10 colonnes, ce sera la liste des 10 lettres correspondantes.

        """
        for row in range(1,sheet.max_row + 1):
            for column in range(1,sheet.max_column + 1): 
                formula = sheet.cell(row,column).value 
                if isinstance(formula, str) and formula.startswith('='):
                    sheet.cell(row,column).value  = Str.updateOneFormula(formula, insert, rowOrColumn, modifications)
        

class Str():
    def __init__(self,chaine):
        self.chaine = str(chaine)
        

    def transform_string_in_binary(self,*args):
        """
        Fonction qui prend un str et qui le transforme en 0 ou 1

        Inputs : args : des chaînes de caractère devant renvoyer 1 
        Outputs : bool : 0 ou 1.
        """
        bool = 0 
        if self.chaine in args:
            bool = 1
        return bool
    
    def set_answer_in_group(self, groups_of_response):
        """
        Function which takes a response and return a string of the group containing the response.
        
        Input : groups_of_response : dictionnary whick keys are response groups and which values are a list of responses 
        associated to this group.
        Output : the string of the group containing the response. 
        """

        """
        for group in groups_of_response.keys():
            if self.chaine in groups_of_response[group] :
                return group
        return ""
        """
        if self.chaine in groups_of_response.keys():
            return groups_of_response[self.chaine]
        else:
            return ""
        
    
    def clean_string(self):
        """
        Fonction qui prend une chaîne de caractère et qui élimine tous les espaces de début et de fin.
        Fonction qui nettoie également les espaces insécables \xa0 par un espace régulier.
        Ceci rendra une chaîne de caractère qui remplacera l'attribut chaine de la classe.  
        On pourra ainsi éviter les erreurs liées à une différence d'un seul espace.      
        """
        depart = 0
        fin = len(self.chaine)
        while self.chaine[depart] == ' ' or self.chaine[fin-1] == ' ':
            if self.chaine[depart] == ' ':
                depart += 1
            if self.chaine[fin-1] == ' ':
                fin -= 1
        self.chaine = self.chaine[depart:fin]
 
        chaine2 = self.chaine.replace('\xa0', ' ')
        self.chaine = chaine2
        return self
    
    def del_extension(self):
        """Fonction 
            - qui enlève l'extension d'un nom de fichier si le nom ne contient pas de date
            - qui ne garde que la partie avant _date_ pour un fichier nommé test_date_****-**-**.xlsx. 
            - qui sert à la sauvegarde et permet ainsi d'éviter des noms à rallonge.
        """
        position = self.chaine.find('_date_')
        if position == -1: 
           position = self.chaine.find('.xlsx')

        return self.chaine[:position]
    
    def cut_string_in_parts(self, separator):
        """
        Fonction qui prend une chaîne de caractères contenant plusieurs sous-chaînes séparées par un séparateur et qui les sépare en plusieurs sous-chaînes.

        Input : separator

        Output : Un tuple contenant les morceaux de chaînes.
        """ 

        parts = ()
        chaine = self.chaine

        debut_part = 0

        for i in range(len(chaine)):
            if chaine[i] == separator:
                parts = parts + (chaine[debut_part:i],) 
                debut_part = i+1

        parts = parts + (chaine[debut_part:],) 
        return parts
    
    def convert_time_in_minutes(self):
        """
        Function which takes a str of the form "10 jour 5 heures" and return a string giving the conversion in unity.

        Output : str
        """
        parts = self.cut_string_in_parts(" ")
        
        if parts[1] in ["jour","jours"]:
            duration = 24 * 60 * float(parts[0]) 
            if len(parts) > 2:
                if parts[3] in ['heure', 'heures']:
                    duration += float(parts[2]) * 60
                elif parts[3] == 'min':
                    duration +=  float(parts[2])
                else:
                    duration += round(float(parts[2])/60,2)
        elif parts[1] in ['heure', 'heures']:
            duration = float(parts[0]) * 60
            if len(parts) > 2:
                if parts[3] == "min":
                    duration += float(parts[2])
                else:
                    duration += round(float(parts[2])/60,2)
        elif parts[1] == "min":
            duration = float(parts[0])
            if len(parts) > 2:
                duration += round(float(parts[2])/60,2)
        else:
            duration = round(float(parts[0])/60,2)
    
        conversion = str(duration).replace('.',',')
        return conversion
    
    @classmethod
    def listFromColumnsStrings(cls,*strings):
        """
        Fonction qui prend en entrée une séquence de chaînes de caractères de la forme "C-E,H,J" et qui retourne une liste de listes (ici 
        [['C','D','E','H','J'],[autres],[autres]]).

        Input : 
            - strings (list).

        Output:
            -Liste (list) : Liste des listes de lettres.
        """
        Liste = []
        
        for string in strings: 
            Liste.append(cls.columns_from_strings(string)) 
        
        return Liste
    
    @classmethod
    def columns_from_strings(cls, string):
        """
        Fonction qui prend en entrée une chaîne de caractères de la forme "C-E,H,J" et qui retourne une liste de colonnes 
        ['C','D','E','H','J'].

        Input : 
            - string (str).

        Output:
            - Liste (list) : Liste de lettres.
        """
        substrings = string.split(',')
        column_list = []
        for substring in substrings:
            if '-' in substring:
                column_list += cls.rangeLetter(substring)
            else:
                column_list.append(substring)
        return column_list
    
    @staticmethod
    def rangeLetter(string):
        """
        Fonction qui prend une chaîne de la forme "D-G" et qui retourne la liste des lettres entre elles.
            
        Input: 
            -string : forme "D-G"

        Output:
            - list : contient les colonnes allant de la première lettre à la dernière
        """
        L = string.split('-')
        return get_column_interval(L[0],L[-1])
    
    @staticmethod
    def updateOneFormula(formula, insert, rowOrColumn, modifications):
        """
        Fonction qui va mettre à jour la formule d'une cellule suite plusieurs suppressions de colonne/ligne.

        Inputs : 
            - formula (str) : la chaîne de caractères.
            - insert (bool) : True si les modifications sont toutes des insertions, False si ce sont toutes des suppressions.
            - rowOrColumn (str): 'row' ou 'column' suivant que la série d'opérations effectuées porte sur ligne ou colonne.
            - modifications (list[str]) : liste de str donnant les modifications. Si on a inséré 10 colonnes, ce sera la liste des 10 lettres correspondantes.

        Output : 
            - formula modified (str).
        """
        for elt in modifications:
            formula = Str.updateOneFormulaForOneInsertion(formula,insert,rowOrColumn,elt)
            
        return formula


    @staticmethod
    def updateOneFormulaForOneInsertion(formula, insert, rowOrColumn, modification):
        """
        Fonction qui va mettre à jour la formule d'une cellule suite à un ajout ou une suppression de colonne/ligne.

        Inputs : 
            - formula (str) : la chaîne de caractères.
            - insert (bool) : True si on a inséré, False si on a supprimé.
            - rowOrColumn (str): 'row' ou 'column' suivant que la série d'opérations effectuées porte sur ligne ou colonne.
            - modification (str) :  str donnant la modification, soit la lettre de la colonne, soit le numéro de la ligne


        Output : 
            - formula modified (str).
        """
         
        #Isoler les cellules des formules
        L1 = re.split(r'(\b[A-Za-z-$]+\d+\b)',formula) 
        for i in range(len(L1)):
            elt = L1[i]
            #si l'élt est une cellule, on la modifie :
            if re.fullmatch(r'\b[A-Za-z-$]+\d+\b', elt):
                L2 = re.split(r'(\d+)',elt)[:-1] 
                #si on a supprimé ou inséré une ligne
                if rowOrColumn == "row":
                    if int(L2[1]) > int(modification):
                        if insert: 
                            L2[1] = str(int(L2[1])+1) 
                        else:
                            L2[1] = str(int(L2[1])-1)
                #même chose sur les colonnes
                else:
                    if '$' in L2[0]:
                        letter = L2[0][:-1]
                    else:
                        letter = L2[0] 
                    if column_index_from_string(letter) > column_index_from_string(modification):
                        if insert:
                            letter = get_column_letter(column_index_from_string(letter) + 1)
                        else:
                            letter = get_column_letter(column_index_from_string(letter) - 1)
                    if '$' in L2[0]:
                        L2[0] = letter + '$'
                    else:
                        L2[0] = letter
                L1[i] = ''.join(L2) 
        return ''.join(L1) 
                        
        
    
class UtilsForcommands():
    def askArgumentUntilNone(args, message):
        """
        Fonction qui permet de demander en ligne de commande un nombre d'arguments indéterminé à l'utilisateur. S'arrête quand l'utilisateur ne rentre rien.
        """
        if not args:
            args = []
            while True:
                user_input = typer.prompt(message, default="")
                if not user_input:
                    break
                args.append(user_input)
        return args
    
    def createDictByCmd(message1, message2):
        """
        Fonction qui demande à l'utilisateur de créer un dictionaire une clé, une valeur.
        """
        dictionary = {}
        while True: 
            key = typer.prompt(message1, default = "")
            if not key:
                break 
            value = typer.prompt(message2, default = "")
            dictionary[key] = value
        return dictionary

        
    @classmethod
    def createDictListValueByCmd(cls, message1, message2):
        """
        Fonction qui demande en ligne de commande à l'utilisateur de rentrer un par un les groupes de réponses et, pour chaque groupe, de rentrer
        les réponses qui lui sont associées.
        """ 
        answers = None
        dictionary = {} 
        while True:
            user_input = typer.prompt(message1, default = "")
            if not user_input:
                break
            else: 
                values = cls.askArgumentUntilNone(answers, message2) 
            dictionary[user_input] = values
        return dictionary
            
    def insertOrOverwrite(column_write):
        """
        Fonction qui propose à l'utilisateur d'insérer ou écraser les données de la colonne dans laquelle il écrit.
        """
        bool = typer.prompt(f"By default, we insert a new column {column_write} to write data. If you agree, press enter. If you want to overwrite the column {column_write}, write False" , default = True)
        return bool
    
class Other():
    @staticmethod
    def reverse_dico_for_set_answer_in_group(dictionary):
        """
        Function taking a dictionary of the form {'group1':['a','b'],'group2':['c','d','e']} and returning the dictionary
        {'a':'group1','b':'group1','c':'group2','d':'group2','e':'group2'}
        """
        reverse_dico = {}
        for key, value in dictionary.items():
            for reponse in value:
                reverse_dico[reponse] = key
        return reverse_dico

    @staticmethod
    def getCellNumericalValue(compiler,tab,cell):
        """
        Fonction qui prend la valeur d'une cellule et qui, si c'est une formule, retourne sa valeur numérique
        """ 

        # Compiler les formules Excel 

        formula = cell.value 
        if isinstance(formula, str) and formula.startswith('='):
            value = compiler.evaluate(tab + '!' + cell.coordinate) 
        else:
            value = formula
        return value
    
    @classmethod
    def display_running_infos(cls, method_run, current_part, list_of_method_parts, start):
        """ 
        Print the percentage of completion of a method 
        
        Inputs:
            - method (str): the method which is raun
            - name (str): represents the current run (could be a tab if we run on multiple tabs)
            - list_name (list[str]) : list of names the program must run
            - start (float): the time at the beginning of the running process
        """
        completion_percentage = round((list_of_method_parts.index(current_part) + 1)/len(list_of_method_parts) * 100,2)
        time_elapsed = time() - start
        time_remaining = (100 - completion_percentage) * time_elapsed / completion_percentage
        print(f'\n---------------Currently running method {method_run}---------------\n')
        print(f'Percentage of completion : {completion_percentage}%')
        print(f'{current_part} is finished')
        cls.display_time_in_adapted_unit(time_elapsed, 'Elapsed time')
        cls.display_time_in_adapted_unit(time_remaining, 'Estimated remaining time') 

    @staticmethod
    def display_time_in_adapted_unit(duration, time_title):
        """
        Print a duration in sec if it is less than 60s, in minutes if it is between 60s and 3600s, in hours otherwise.
        
        Inputs:
            - duration (float)
            - time_type(str): the kinf of time to print in the string
        """
        if duration < 60:
            print(f'{time_title} : {round(duration, 1)} sec')
        elif 60 <= duration < 3600: 
            duration /= 60 
            print(f'{time_title} : {round(duration, 1)} min')
        else: 
            duration /= 3600 
            print(f'{time_title} : {round(duration, 1)} h')


class DisplayRunningInfos():
    def __init__(self):
        self.method_name = None
        self.current_running_part = None
        self.list_of_running_parts = None
        self.start_time = None
        self.elapsed_time = 0.
        self.remaining_time = 0.
        self.completion_percentage = 0.
        self.time_title = ''  

    def display_running_infos(self):
        """ 
        Print the percentage of completion of a method 
        
        Inputs:
            - method (str): the method which is raun
            - name (str): represents the current run (could be a tab if we run on multiple tabs)
            - list_name (list[str]) : list of names the program must run
            - start (float): the time at the beginning of the running process
        """
        self._compute_time_and_completion_data()
        self._display_completion_data()
        self._display_times_data() 

    def _compute_time_and_completion_data(self):
        self.completion_percentage = round((self.list_of_running_parts.index(self.current_running_part) + 1)/len(self.list_of_running_parts) * 100,2)
        self.elapsed_time = time() - self.start_time 
        self.remaining_time = (100 - self.completion_percentage) * self.elapsed_time / self.completion_percentage 

    def _display_completion_data(self):
        print(f'\n---------------Currently running method {self.method_name}---------------\n')
        print(f'Percentage of completion : {self.completion_percentage}%')
        print(f'{self.current_running_part} is finished')

    def _display_times_data(self):
        self._update_time_title('Elapsed time')
        self.display_time_in_adapted_unit(self.elapsed_time)
        self._update_time_title('Estimated remaining time')
        self.display_time_in_adapted_unit(self.remaining_time)

    def display_time_in_adapted_unit(self, duration):
        """
        Print a duration in sec if it is less than 60s, in minutes if it is between 60s and 3600s, in hours otherwise.
        
        Inputs:
            - duration (float)
            - time_type(str): the kind of time to print in the string
        """
        if duration < 60:
            print(f'{self.time_title} : {round(duration, 20)} sec')
        elif 60 <= duration < 3600: 
            duration /= 60 
            print(f'{self.time_title} : {round(duration, 20)} min')
        else: 
            duration /= 3600 
            print(f'{self.time_title} : {round(duration, 20)} h')

    def _update_time_title(self, time_title):
        self.time_title = time_title


class TabsCopy():
    """Make copys from a tab to a new tab"""

    def __init__(self):
        self.tab_from = None
        self.tab_to = None

    def _choose_the_tab_to_write_in(self, tab):
        self.tab_to = tab

    def _choose_the_tab_to_read(self, tab):
        self.tab_from = tab
        
    def copy_paste_line(self, line_from, line_to):#, values_only=False):
            """
            Fonction qui prend une ligne de la feuille et qui la copie dans un autre onglet.

            Inputs : 
                - onglet_from : onglet d'où on copie
                - row_from : ligne de l'onglet d'origine.
                - onglet_to : onglet où coller.
                - row_to : la ligne où il faut coller dans l'onglet à modifier.

            Exemple d'utilisation : 
        
                file = File('dataset.xlsx')
                file.copy_paste_line('onglet1', 1, 'onglet2', 1)
            """

            """ 
            ON REPRENDRA CETTE VERSION QUAND J ATTAQUERAIS LES HISTOIRES DE VALUES ONLY

            # Cas où on ne copie que les valeurs, cell est un str
            if values_only:
                column_index = 1 
                for column in onglet_from.iter_cols(min_row=row_from, max_row=row_from, min_col=0, max_col=onglet_from.max_column, values_only=values_only):
                    for cell in column: 
                        if cell:
                            onglet_to.cell(row_to, column_index).value = cell
                        column_index += 1
                        
            # Cas où on copie les formules, cell est un objet
            else:
                for column in onglet_from.iter_cols(min_row=row_from, max_row=row_from, min_col=0, max_col=onglet_from.max_column, values_only=values_only):
                    for cell in column: 
                        if cell.value != "":
                            onglet_to.cell(row_to, cell.column).value = cell.value """

            for column_index in range(1, self.tab_from.max_column + 1): 
                self.tab_to.cell(line_to, column_index).value = self.tab_from.cell(line_from, column_index).value 

    def copy_paste_column(self, column_from, column_to):
            """
            Fonction qui prend une colonne de la feuille et qui la copie dans un autre onglet.
            """ 
            for line_index in range(1, self.tab_from.max_row + 1): 
                self.tab_to.cell(line_index, column_to).value = self.tab_from.cell(line_index, column_from).value 

    def add_line_at_bottom(self, line_from):
            """
            Fonction qui copie une ligne spécifique de la feuille à la fin d'un autre onglet.

            Input : 
                - row_origin : ligne de l'onglet d'origine.
                - onglet : l'onglet à modifier où on copie la ligne.

            Exemple d'utilisation : 
        
                file = File('dataset.xlsx')
                file.copy_paste_line('onglet1', 1, 'onglet2')
            """ 
            self.copy_paste_line(line_from, self.tab_to.max_row + 1) 

    def copy_old_file_tab_in_new_file_tab(self): 
        for i in range(1, self.tab_from.max_row + 1):
            for j in range(1, self.tab_from.max_column + 1):  
                self.deep_copy_of_a_cell(Cell(i,j), Cell(i,j))

    def deep_copy_of_a_cell(self, cell_from, cell_to):   
        self.tab_to.cell(cell_to.line_index, cell_to.column_index).value = self.tab_from.cell(cell_from.line_index, cell_from.column_index).value  
        self.tab_to.cell(cell_to.line_index, cell_to.column_index).fill = copy(self.tab_from.cell(cell_from.line_index, cell_from.column_index).fill)
        self.tab_to.cell(cell_to.line_index, cell_to.column_index).font = copy(self.tab_from.cell(cell_from.line_index, cell_from.column_index).font) 

    def copy_of_a_cell(self, cell_from, cell_to):   
        self.tab_to.cell(cell_to.line_index, cell_to.column_index).value = self.tab_from.cell(cell_from.line_index, cell_from.column_index).value
    
