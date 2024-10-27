import openpyxl 
from utils import UtilsForFile, Other, Str 
from pycel import ExcelCompiler  
from time import time
from datetime import datetime 
from copy import copy


class File(UtilsForFile): 
    def __init__(self, name_file, path = 'fichiers_xls/', dataonly = False):
        """L'utilisateur sera invité à mettre son fichier xlxx dans un dossier nommé fichiers_xls
        """ 
        self.name_file = name_file  
        self.path = path
        self.dataonly = dataonly 
        self.writebook = openpyxl.load_workbook(self.path + self.name_file, data_only = dataonly)
        self.sheets_name = self.writebook.sheetnames 

    def create_excel_compiler(self):
        return ExcelCompiler(self.path + self.name_file) 
    
    def sauvegarde(self):
        """
        Fonction qui crée une sauvegarde du fichier name_file et 
        qui l'appelle name_file_time où time est le moment d'enregistrement.

        Exemple d'utilisation : 
            file = File('dataset.xlsx')
            file.sauvegarde()

            Si on veut copier les formules et pas seulement les valeurs des cellules : 
            file = File('dataset.xlsx', dataonly = False)
            file.sauvegarde()
        """
        file_copy = openpyxl.Workbook()
        del file_copy[file_copy.active.title] #supprimer l'onglet créé

        start = time()

        for onglet in self.sheets_name: 
            new_sheet = file_copy.create_sheet(onglet)
            initial_sheet = self.writebook[onglet] 

            for i in range(1,initial_sheet.max_row+1):
                for j in range(1,initial_sheet.max_column+1): 
                    new_sheet.cell(i,j).value = initial_sheet.cell(i,j).value  
                    new_sheet.cell(i,j).fill = copy(initial_sheet.cell(i,j).fill)
                    new_sheet.cell(i,j).font = copy(initial_sheet.cell(i,j).font) 
            
            Other.display_running_infos('sauvegarde', onglet, self.sheets_name, start)
            
                    
        name_file_no_extension = Str(self.name_file).del_extension() 
        
        file_copy.save(self.path  + name_file_no_extension + '_date_' + datetime.now().strftime("%Y-%m-%d_%Hh%M") + '.xlsx') 
        return file_copy

