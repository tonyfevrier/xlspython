import os
import openpyxl
import json
import re

from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, coordinate_to_tuple, get_column_letter
from time import time
from utils.utils import Other, UtilsForFile, Str 
from copy import copy
from model import File


class FileControler(UtilsForFile):
    def __init__(self, file):
        """
        Input: file (object of class File)
        """
        self.file = file


    ## Multiple tabs methods

    def apply_method_on_some_sheets(self, onglets, method_name, *args, **kwargs):
        """ 
        Vous avez un fichier contenant plusieurs onglets et vous souhaitez appliquer une même méthode de la 
        classe Sheet sur une liste de ces onglets du fichier. On s'attend à ce que tous les onglets aient une structure identique.

        Inputs:
            - filename (str)
            - method_name (str): the name of the method to execute 
            - *args, **kwargs : arguments of the method associated with method_name
        """  
        start = time()
        for sheetname in onglets:    
            # Get the method and apply it
            method = getattr(self, method_name)
            method(sheetname, *args, **kwargs)  
            Other.display_running_infos(method_name, sheetname, onglets, start)

        self.file.writebook.save(self.file.path + self.file.name_file)



    def create_one_onglet_by_participant(self, onglet_from, column_read, newfile_name, newfile_path, first_line=2):
        """
        Fonction qui prend un onglet dont une colonne contient des chaînes de caractères comme par exemple un nom.
        Chaque chaîne de caractères peut apparaître plusieurs fois dans cette colonne (exe : quand un participant répond plusieurs fois)
        La fonction retourne un fichier contenant un onglet par chaîne de caractères.
          Chaque onglet contient toutes les lignes correspondant à cette chaîne de caractères.

        Input : 
            onglet_from : onglet de référence.
            column_read : l'étiquette de la colonne qui contient les chaînes de caractères.
            newfile_name (str): name of the newfile
            newfile_path (str): where to write/find the newfile
            first_line : ligne où commencer à parcourir.
            last_line : ligne de fin de parcours 
 
        Exemple d'utilisation : 
    
            file = File('dataset.xlsx')
            file.create_one_onglet_by_participant('onglet1', 'A') 
        """ 
        column_read = column_index_from_string(column_read)  

        sheet = self.file.writebook[onglet_from] 

        # Creation of the file aiming to contain the data if it does not already exists 
        if newfile_name not in os.listdir(newfile_path):
            new_file = openpyxl.Workbook()
        else:
            new_file = openpyxl.load_workbook(newfile_path + newfile_name)

        onglets = new_file.sheetnames

        # Create one tab by identifiant containing all its lines
        for i in range(first_line, sheet.max_row + 1):
            onglet = str(sheet.cell(i,column_read).value)

            # Prepare a new tab
            if onglet not in onglets:
                new_file.create_sheet(onglet)
                self.copy_paste_line(sheet, 1,  new_file[onglet], 1)
                onglets.append(onglet) 

            self.add_line_at_bottom(sheet, i, new_file[onglet]) 
        
        # Deletion of the first tab if the newfile was created
        if newfile_name not in os.listdir(newfile_path):
            del new_file[new_file.sheetnames[0]]
        new_file.save(newfile_path + newfile_name)
        

    def extract_column_from_all_sheets(self, column):
        """
        Fonction qui récupère une colonne dans chaque onglet pour former une nouvelle feuille
        contenant toutes les colonnes. La première cellule de chaque colonne correspond alors 
        au nom de l'onglet. Attention, en l'état, il faut que tous les onglets aient la même structure.

        Input : 
            column : str. L'étiquette de la colonne à récupérer dans chaque onglet 

        Exemple d'utilisation : 
    
            file = File('dataset.xlsx')
            file.extract_column_from_all_sheets('B') 

            Si on veut extraire les formules

            file = File('dataset.xlsx',dataonly = False)
            file.extract_column_from_all_sheets('B') 
        """ 
        column = column_index_from_string(column)
         
        new_sheet = self.file.writebook.create_sheet(f"gather_{column}")
        column_to = 1

        start = time()
        for name_onglet in self.file.sheets_name: 
            onglet = self.file.writebook[name_onglet] 
            self.copy_paste_column(onglet,column,new_sheet,column_to)
            column_to = new_sheet.max_column + 1
            new_sheet.cell(1,new_sheet.max_column).value = name_onglet 
            Other.display_running_infos('extract_column_from_all_sheets', name_onglet, self.file.sheets_name, start)

            
        self.file.writebook.save(self.file.path + self.file.name_file) 
        self.file.sheets_name = self.file.writebook.sheetnames 

    def extract_cells_from_all_sheets(self, *cells):
        """
        Vous avez un fichier avec des onglets de structure identique avec un onglet par participant. Vous souhaitez
        récupérer des cellules identiques dans tous les onglets et créer un onglet avec une ligne par participant,
        qui contient les valeurs de ces cellules. Fonction analogue à gather_multiple_answers mais ne portant pas sur une
        seule feuille.

        Inputs:
            - cells (list[str])
        """ 

        # Recover cells coordinates
        cell_list = []
        for cell in cells: 
            cell_list.append(coordinate_to_tuple(cell)) 
        
        # Create a new tab
        gathered_sheet = self.file.writebook.create_sheet('gathered_data')
        current_line = 2

        start = time()

        # Fill one line by tab
        for name_onglet in self.file.sheets_name:   
            current_onglet = self.file.writebook[name_onglet]
            gathered_sheet.cell(current_line, 1).value = name_onglet
            current_column = 2

            # Fill selected values one by one
            for tuple in cell_list:  
                gathered_sheet.cell(current_line, current_column).value = current_onglet.cell(tuple[0],tuple[1]).value
                current_column += 1
            current_line += 1
            Other.display_running_infos('extract_cells_from_all_sheets', name_onglet, self.file.sheets_name, start)


        self.file.sheets_name = self.file.writebook.sheetnames 
        self.file.writebook.save(self.file.path + self.file.name_file)
        

    def apply_column_formula_on_all_sheets(self, *column_list):
        """
        Fonction qui reproduit les formules d'une colonne ou plusieurs colonnes
          du premier onglet sur toutes les colonnes situées à la même position dans les 
          autres onglets.

        Input : 
            -column_list : int. les positions des colonnes où récupérer et coller.

        Exemples d'utilisation : 

            Bien veiller à mettre dataonly = False sinon il ne copiera pas les formules mais
            les valeurs des cellules. On peut aussi copier les valeurs des cellules : pour cela,
            enlever dataonly = False.

            Sur une colonne
                file = File('dataset.xlsx', dataonly = False)
                file.apply_column_formula_on_all_sheets(2) 

            Sur trois colonnes
                file = File('dataset.xlsx', dataonly = False)
                file.apply_column_formula_on_all_sheets(2,5,10) 

            Sur toutes les colonnes du fichier à partir de la colonne colmin jusque la colonne colmax :
                file = File('dataset.xlsx', dataonly = False)
                file.apply_column_formula_on_all_sheets(*[i for i in range(colmin,colmax + 1)]) 
        """
        column_int_list = []
        for column in column_list: 
            column_int_list.append(column_index_from_string(column))  

        start = time()

        #on applique les copies dans tous les onglets sauf le premier
        for name_onglet in self.file.sheets_name[1:]:
            for column in column_int_list:
                self.copy_paste_column(self.file.writebook[self.file.sheets_name[0]],column,self.file.writebook[name_onglet],column)
            Other.display_running_infos('apply_column_formula_on_all_sheets', name_onglet, self.file.sheets_name[1:], start)
            

        self.file.writebook.save(self.file.path + self.file.name_file)

    def apply_cells_formula_on_all_sheets(self, *cells):
        """
        Fonction qui reproduit les formules d'une cellule ou plusieurs cellules
          du premier onglet sur toutes les cellules situées à la même position dans les 
          autres onglets.

        Input : 
            -cells : string. les positions des cellule où récupérer et coller 

        Exemples d'utilisation : 

            Bien veiller à mettre dataonly = False sinon il ne copiera pas les formules mais
            les valeurs des cellules. On peut aussi copier les valeurs des cellules : pour cela,
            enlever dataonly = False.

            file = File('dataset.xlsx', dataonly = False)
            file.apply_column_formula_on_all_sheets('C5','D6')  
        """

        #obtenir les indices de la cellule 
        cell_list = []
        for cell in cells: 
            cell_list.append(coordinate_to_tuple(cell)) 

        start = time()

        #on applique les copies dans tous les onglets sauf le premier
        for name_onglet in self.file.sheets_name[1:]:   
            for tuple in cell_list: 
                self.file.writebook[name_onglet].cell(tuple[0],tuple[1]).value = self.file.writebook[self.file.sheets_name[0]].cell(tuple[0],tuple[1]).value  
                self.file.writebook[name_onglet].cell(tuple[0],tuple[1]).fill = copy(self.file.writebook[self.file.sheets_name[0]].cell(tuple[0],tuple[1]).fill)  
                self.file.writebook[name_onglet].cell(tuple[0],tuple[1]).font = copy(self.file.writebook[self.file.sheets_name[0]].cell(tuple[0],tuple[1]).font)  
                self.file.writebook[name_onglet].cell(tuple[0],tuple[1]).border = copy(self.file.writebook[self.file.sheets_name[0]].cell(tuple[0],tuple[1]).border)  
                self.file.writebook[name_onglet].cell(tuple[0],tuple[1]).alignment = copy(self.file.writebook[self.file.sheets_name[0]].cell(tuple[0],tuple[1]).alignment)    
            Other.display_running_infos('apply_cells_formula_on_all_sheets', name_onglet, self.file.sheets_name[1:], start)

        self.file.writebook.save(self.file.path + self.file.name_file)

    def gather_columns_in_one(self, onglet, *column_lists):
        """
        Vous avez des groupes de colonnes de valeurs avec une étiquette en première cellule. Pour chaque groupe, vous souhaitez former deux colonnes de valeurs : l'une qui contient
        les valeurs rassemblées en une colonne, l'autre, à sa gauche, qui indique l'étiquette de la colonne dans laquelle elle a été prise.

        Inputs : 
            - onglet (str) : nom de l'onglet d'où on importe les colonnes.
            - column_lists (list[list[str]]) : liste de groupes de colonnes. Chaque groupe est une liste de colonnes.
        """ 

        for liste in column_lists:
            tab_number = column_lists.index(liste)
            self.file.writebook.create_sheet(f"tab_column_gathered_{tab_number}")
            target_sheet = self.file.writebook[f"tab_column_gathered_{tab_number}"]
            for column in liste: 
                self.copy_column_tags_and_values_at_bottom(self.file.writebook[onglet], column_index_from_string(column), target_sheet)

        self.file.writebook.save(self.file.path + self.file.name_file) 

    def build_file_from_tab(self, tab):
        """
        Fonction qui prend un nom d'onglet dans un fichier et qui crée un fichier associé.

        Input :
            - tab (str) : the name of the tab from which we want to create the file.
        """

        sheet_from = self.file.writebook[tab]
        newfile = openpyxl.Workbook() 
        sheet_to = newfile['Sheet']  
        path = 'multifiles/' 
  
        self.deep_copy_of_a_sheet(sheet_from, sheet_to) 

        namefile = path + tab + '.xlsx'
        newfile.save(namefile) 
        return namefile
            
            
    def one_file_by_tab_sendmail(self, send = False, adressjson = "", objet = "", message = ""):
        """
        Vous souhaitez fabriquer un fichier par onglet. Chaque fichier aura le nom de l'onglet. 
        Vous souhaitez éventuellement envoyer chaque fichier à la personne associée.
        Attention, pour utiliser cette fonction, les onglets doivent être de la forme "prenom nom" sans caractère spéciaux. 

        Inputs : 
            send(optional boolean) : True si on veut envoyer le mail, False si on veut juste couper en fichiers.
            adressjson(str) : nom du fichier xlsx qui contient deux colonnes la première avec les noms des onglets, la seconde avec l'adresse mail. Ce fichier doit être mis dans le dossier fichier_xls. 
            objet(optional str) : Objet du message.
            message (optional str) : Contenu du message.
        """ 
        if adressjson != "":
            file = open(self.file.path + adressjson, 'r')
            mailinglist = json.load(file)
            file.close()

        start = time()

        for tab in self.file.sheets_name: 

            file_to_send = self.build_file_from_tab(tab)
            if send:
                if adressjson == "":
                    prenom = tab.split(" ")[0]
                    nom = tab.split(" ")[1]
                    self.envoi_mail(prenom + "." + nom + "@universite-paris-saclay.fr", file_to_send, "tony.fevrier62@gmail.com", "qkxqzhlvsgdssboh", objet, message)
                else: 
                    self.envoi_mail(mailinglist[tab], file_to_send, "tony.fevrier62@gmail.com", "qkxqzhlvsgdssboh", objet, message) 
            Other.display_running_infos('one_file_by_tab_sendmail', tab, self.file.sheets_name, start)
             

    def merge_cells_on_all_tabs(self, start_column, end_column, start_row, end_row):
        """
        Fonction qui merge les mêmes cellules sur tous les onglets d'un fichier 

        Inputs :
            - start_column (string): Letter of the column where to start the merging
            - end_column (string): Letter of the column where to end the merging
            - start_row (int): Index of the row where to start the merging
            - end_row (int): Index of the row where to start the merging

        """
        
        start_column = column_index_from_string(start_column)
        end_column = column_index_from_string(end_column)

        start = time()

        for tab in self.file.sheets_name: 
            sheet = self.file.writebook[tab] 
            sheet.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)
            Other.display_running_infos('merge_cells_on_all_tabs', tab, self.file.sheets_name, start)

        self.file.writebook.save(self.file.path + self.file.name_file)

    def check_linenumber_of_tabs(self, line_number):
        """
        Fonction qui prend un fichier et qui contrôle si tous les onglets ont un nombre de lignes égal à l'argument
        """
        wrong_tabs = []
        for tab in self.file.sheets_name:
            if self.file.writebook[tab].max_row != line_number:
                wrong_tabs.append(tab)
        return wrong_tabs
    

    ## Sheet methods

    def color_special_cases_in_column(self, sheet_name, column, chainecolor):
        """
        Fonction qui regarde pour une colonne donnée colore les cases correspondant à certaines chaînes de caractères.

        Input : 
            - column : le numéro de la colonne.
            - chainecolor (dict) : les chaînes de caractères qui vont être colorées et les couleurs qui correspondent à écrire avec la syntaxe suivante {'vrai':'couleur1','autre':couleur2}. Attention,
                la couleur doit être entrée en hexadécimal et les chaînes de caractères ne doivent pas avoir d'espace au début ou à la fin.
        
        Exemple d'utilisation : 
        
            sheet = Sheet('dataset.xlsx','onglet1')
            sheet.color_special_cases_in_column('L', {'vrai': '#FF0000','faux': '#00FF00'}) 

        """ 
        sheet = self.file.writebook[sheet_name]
        column = column_index_from_string(column) 
        
        for i in range(1, sheet.max_row + 1):
            cellule = sheet.cell(i,column) 

            if cellule.value is str:
                key = Str(cellule.value).clean_string().chaine
            else: 
                key = cellule.value

            if key in chainecolor.keys():
                cellule.fill = PatternFill(fill_type = 'solid', start_color = chainecolor[key])

        #self.file.writebook.save(self.file.path + self.file.name_file)

    def color_special_cases_in_sheet(self, sheet_name, chainecolor): 
        """
        Fonction qui colore les cases contenant à certaines chaînes de caractères d'une feuille
        
        Input : 
            - column : le numéro de la colonne.
            - chainecolor (dict) : les chaînes de caractères qui vont être colorées et les couleurs qui correspondent à écrire avec la syntaxe suivante {'vrai':'couleur1','autre':couleur2}. Attention,
                la couleur doit être entrée en hexadécimal et les chaînes de caractères ne doivent pas avoir d'espace au début ou à la fin.
            
        Exemple d'utilisation : 
    
            sheet = Sheet('dataset.xlsx','onglet1')
            sheet.color_special_cases_in_sheet({'vrai': '#FF0000','faux': '#00FF00'}) 
                
        """ 
        sheet = self.file.writebook[sheet_name]

        for j in range(1, sheet.max_column + 1):
            self.color_special_cases_in_column(sheet_name, get_column_letter(j),chainecolor)

    def add_column_in_sheet_differently_sorted(self, sheet_name, column_identifiant, column_insertion, other_sheet):
        """
        Fonction qui insère dans un onglet des colonnes d'un autre onglet de référence. 
        Les deux feuilles ont une colonne d'identifiants communs (exemple : des mails) mais qui peut être
        triés dans des ordres différents.
        La fonction récupère un ou plusieurs éléments d'une ligne déterminée par un identifiant.
        Elle recherche l'identifiant dans la seconde feuille et insère les éléments
        dans la ligne correspondante.

        Inputs :
            - column_identifiant : numéro de la colonne où sont situés les identifiants dans l'onglet qu'on souhaite modifier.
            - column_insertion : numéro de la colonne où on insère les colonnes à récupérer.
            - other_sheet : liste représentant l'onglet duquel on récupère les colonnes  ['namefile','namesheet',numéro de la colonne où sont les identifiants,[numéros des colonnes à récupérer sous forme de liste]]
                namefile doit être au format .xlsx et mis dans le dossier fichier_xls. 
            
        Exemple d'utilisation : 
    
            sheet = Sheet('dataset.xlsx','onglet1')
            sheet.add_column_in_sheet_differently_sorted('B','C',['file.xlsx', 'onglet2', 'B', ['E','F','H','AA]]) 
                  
        """
        sheet = self.file.writebook[sheet_name]
        
        file_to_copy = openpyxl.load_workbook(self.file.path + other_sheet[0], data_only=True)
        sheet_to_copy = file_to_copy[other_sheet[1]]
        columns_to_copy = other_sheet[3]
        dico = {}

        column_identifiant = column_index_from_string(column_identifiant)
        column_insertion = column_index_from_string(column_insertion)
        other_sheet[2] = column_index_from_string(other_sheet[2])
        columns_to_copy = [column_index_from_string(column) for column in columns_to_copy]

        modifications = [get_column_letter(column_insertion + i ) for i in range(len(columns_to_copy))]

        #Passage en revue les identifiants du premier fichier et création d'un dictionnaire dont les clés sont ces identifiants et les valeurs sont une liste de valeurs à récupérer.
        for i in range(1,sheet_to_copy.max_row + 1):
            value = sheet_to_copy.cell(i,other_sheet[2]).value
            dico[value] = [sheet_to_copy.cell(i,j) for j in columns_to_copy]

        sheet.insert_cols(column_insertion,len(columns_to_copy)) 

        #Passage en revue des identifiants du second fichier et insertion des valeurs si les identifiants sont dans les clés du dico
        #. 
        for i in range(1, sheet.max_row+1):
            key = sheet.cell(i,column_identifiant).value
            if key in dico.keys():
                for j in range(len(columns_to_copy)):
                    sheet.cell(i,column_insertion + j).value = dico[key][j].value
                    sheet.cell(i,column_insertion + j).fill = copy(dico[key][j].fill)
        
        self.updateCellFormulas(sheet,True,'column', modifications)         
        self.file.writebook.save(self.file.path + self.file.name_file)


    def color_lines_containing_chaines(self, sheet_name, color,*chaines):
        """
        Fonction qui colore les lignes dont une des cases contient une str particulière.

        Input : 
            - color : une couleur indiquée en haxadécimal par l'utilisateur.
            - chaines : des chaines de caractères que l'utilisateur entre et qui entraînent la coloration de la ligne.
            
        Exemple d'utilisation : 
    
            sheet = Sheet('dataset.xlsx','onglet1')
            sheet.color_lines_containing_chaine('#FF0000', 'vrai', 'hello', 'heeenri', 'ficheux') 
        
        """
        sheet = self.file.writebook[sheet_name]

        lines_to_color = []

        for i in range(1, sheet.max_row + 1):
            for j in range(1, sheet.max_column + 1):
                if str(sheet.cell(i,j).value) in chaines:
                    lines_to_color.append(i)
                    break
        
        for row in lines_to_color:
            self.color_line(sheet, color, row)
        
        #self.file.writebook.save(self.file.path + self.file.name_file)

    def column_cut_string_in_parts(self, sheet_name, column_to_cut,column_insertion,separator):
        """
        Fonction qui prend une colonne dont chaque cellule contient une grande chaîne de
          caractères. Toutes les chaînes sont composés du nombre de morceaux délimités par un séparateur,
        La fonction insère autant de colonnes que de morceaux et place un morceau par colonne dans l'ordre des morceaux.

        Inputs :
            - column_to_cut : colonne contenant les grandes str.
            - column_insertion : où insérer les colonnes
            - separator le séparateur 

        Exemple d'utilisation : 
    
            sheet = Sheet('dataset.xlsx','onglet1')
            sheet.column_cut_string_in_parts('C', 'J', ';') 
        
        """
        sheet = self.file.writebook[sheet_name]

        column_to_cut = column_index_from_string(column_to_cut) 
        column_insertion = column_index_from_string(column_insertion)
        
        for i in range(2, sheet.max_row + 1):
            value = sheet.cell(i,column_to_cut).value
            chaine_object = Str(value)
            parts = chaine_object.cut_string_in_parts(separator)
            modifications = [get_column_letter(column_insertion + i) for i in range(len(parts))]
            if i == 2:
                sheet.insert_cols(column_insertion,len(parts))
            for j in range(len(parts)):
                sheet.cell(i,column_insertion + j).value = parts[j]

        self.updateCellFormulas(sheet,True,'column', modifications)         
        #self.file.writebook.save(self.file.path + self.file.name_file) 

    def delete_columns(self, sheet_name, columns):
        """
        Prend une séquence de colonnes sous forme de lettres qu'on souhaite supprimer.

        Input : 
            - columns (str): list of column of the form 'C-J,K,L-N,Z' 
        """ 
        sheet = self.file.writebook[sheet_name]

        # Réordonner par les lettres les plus grandes pour supprimer de la droite vers la gauche dans l'excel  
        columns_to_delete = Str.columns_from_strings(columns)
        columns_to_delete.sort(reverse = True) 

        for column in columns_to_delete: 
            sheet.delete_cols(column_index_from_string(column)) 

        self.updateCellFormulas(sheet, False, 'column', columns_to_delete)
        #self.file.writebook.save(self.file.path + self.file.name_file) 

    def delete_other_columns(self, sheet_name, columns):
        """
        Prend une séquence de colonnes sous forme de lettres à conserver et supprime les autres

        Input : 
            - columns (str): list of column of the form 'C-J,K,L-N,Z'
        """
        sheet = self.file.writebook[sheet_name]

        columns_to_keep = Str.columns_from_strings(columns)
        modifications = []

        for column in range(sheet.max_column + 1, 0, -1):
            column_letter = get_column_letter(column)
            if column_letter not in columns_to_keep:
                modifications.append(column_letter)
                sheet.delete_cols(column)
       
        self.updateCellFormulas(sheet, False, 'column', modifications)
        #self.file.writebook.save(self.file.path + self.file.name_file) 
        

    def delete_lines_containing_str(self, sheet_name, column, *chaines):
        """
        Fonction qui parcourt une colonne et qui supprime la ligne si celle-ci contient une chaîne particulière.

        Inputs : 
            -column : la colonne à parcourir.
            -chaines : les chaînes de caractères qui doivent engendrer la suppression de la ligne.
        
        Exemple d'utilisation : 
    
            sheet = Sheet('dataset.xlsx','onglet1')
            sheet.delete_lines(3, 'chaine1', 'chaine2', 'chaine3', 'chaine4') 
        """
        sheet = self.file.writebook[sheet_name]

        column = column_index_from_string(column)  
        
        modifications = []
        for i in range(sheet.max_row,0,-1):
            value = Other.getCellNumericalValue(self.file.create_excel_compiler(), sheet_name, sheet.cell(i,column)) 
            if str(value) in chaines:  
                sheet.delete_rows(i) 
                modifications.append(str(i))
 
        self.updateCellFormulas(sheet,False,'row',modifications)        
        #self.file.writebook.save(self.file.path + self.file.name_file)

    def delete_doublons(self, sheet_name, column_identifiant, line_beginning = 2, color = False):
        """
        Certains participants répondent plusieurs fois. Cette fonction supprime les premières réponses
        des participants dans ce cas. Elle ne garde que leur dernière réponse. On repère les participants
        par leur identifiant unique donné dans colum_identifiant.

        Inputs:
            column_identifiant : str: lettre de la colonne qui contient les identifiants des participants.
            color : boolean : True si on veut que la ligne des participants qui ont répondu plusieurs fois soit colorée dans le datasetfinal.
        
        Exemple d'utilisation : 
    
        si on ne veut pas repérer les personnes qui étaient en doublon:
            sheet = Sheet('dataset.xlsx','onglet1')
            sheet.delete_doublons('C')

        si on veut les repérer :
            sheet = Sheet('dataset.xlsx','onglet1')
            sheet.delete_doublons('C', color = True)
        """
        sheet = self.file.writebook[sheet_name]

        column_identifiant = column_index_from_string(column_identifiant) 

        participants = {} 
        modifications = []

        #On parcourt dans le sens inverse afin d'éviter que la suppression progressive impacte la position des lignes étudiées ensuite.
        i = sheet.max_row 
        while i != line_beginning:  
            identifiant = Str(sheet.cell(i,column_identifiant).value).clean_string() 
            if identifiant.chaine in participants.keys():
                if color:
                    self.color_line(sheet, '0000a933', participants[identifiant.chaine])
                sheet.delete_rows(i)
                modifications.append(str(i))
                participants[identifiant.chaine] -= 1    
            else:
                participants[identifiant.chaine] = i 
            i -= 1

        self.updateCellFormulas(sheet,False,'row',modifications)        
        #self.file.writebook.save(self.file.path + self.file.name_file)
    
    def create_one_column_by_QCM_answer(self, sheet_name, column, column_insertion, list_string, *reponses, line_beggining = 2):
        """
        Fonction qui regarde si des réponses sont incluses dans les cellules d'une colonne.
        Chaque cellule contient l'ensemble des réponses à une question de QCM du participant sous forme de str.
        Elle regarde les cellules de column. Si une réponse est dans cette cellule, on l'indique dans la colonne
        correspondante.

        Inputs : 
            - column :  str : la colonne avec les réponses.
            - column_insertion : str : l'endroit où on insère les colonnes.
            - list_string : list : liste de deux str indiquant si la réponse est présente ou non.
            - reponses : les réponses du QCM. 
        
        Exemple : 
            sheet = Sheet('dataset.xlsx','onglet1')
            sheet.create_one_column_by_QCM_answer('C','D', ['oui', 'non'], 'reponse1', 'reponse2', 'reponse3')

        """
        sheet = self.file.writebook[sheet_name]

        column = column_index_from_string(column) 
        column_insertion = column_index_from_string(column_insertion) 
        
        modifications = [get_column_letter(column_insertion + i) for i in range(len(reponses))]

        #on crée les colonnes pour chaque réponse
        sheet.insert_cols(column_insertion,len(reponses))
        for j in range(0,len(reponses)):
            sheet.cell(1,j + column_insertion).value = reponses[j]

        #on remplit les colonnes suivant que les réponses correspondantes sont ou non dans la cellule.
        for i in range(line_beggining, sheet.max_row + 1):
            if sheet.cell(i,column).value is None:
                for j in range(0,len(reponses)):  
                        sheet.cell(i,j + column_insertion).value = list_string[1]
            else:
                for j in range(0,len(reponses)):  
                    if reponses[j] in sheet.cell(i,column).value:
                        sheet.cell(i,j + column_insertion).value = list_string[0]
                    else:
                        sheet.cell(i,j + column_insertion).value = list_string[1]

        self.updateCellFormulas(sheet,True,'column',modifications)        
        #self.file.writebook.save(self.file.path + self.file.name_file)
        
    def gather_multiple_answers(self, sheet_name, column_read, column_store, line_beggining = 2):
        """
        Dans un onglet, nous avons les réponses de participants qui ont pu répondre plusieurs fois à un questionnaire.
        Cette fonction parcourt les noms et met dans un autre onglet. La ligne du participant est alors constituée des différentes valeurs
         d'une même donnée récupérée.
        
        Inputs :
            - column_read (str) : la colonne avec les identifiants des participants.
            - column_store (str) : lettre de la colonne contenant la donnée qu'on veut stocker.
            - line_beggining (int) : ligne où débute la recherche. 
        """ 
        sheet = self.file.writebook[sheet_name]

        column_read = column_index_from_string(column_read) 
        column_store = column_index_from_string(column_store) 

        #we create a dictionary whose keys are the identifiers (of participants) and values are their number of answers and a list containing
        #the data we want to store for each answer.
        dico = self.create_dico_to_store_multiple_answers_of_participants(sheet, column_read,column_store,line_beggining)
        
        #we create the new sheet where we store participants answering multiple times and their data.
        storesheet = self.file.writebook.create_sheet('severalAnswers')
        self.create_newsheet_storing_multiple_answers(storesheet, dico)

        self.file.writebook.save(self.file.path + self.file.name_file) 
    
    def act_on_columns(function):
        """
        Décorateur qui en plus d'appliquer la fonction, transforme les lettres de colonnes en index, met à jour les formules 
        après l'insertion de la colonne et sauvegarde le fichier.
        """
        def wrapper(self, *args, **kwargs):
            """
            - args[0] (str): sheet name
            - args[1] (list[str] or str): letters of columns to read
            - args[2] (str): letter of column in which to write
            """
            # Transform all args corresponding to columns in indexes 
            modifications = [args[2]]
            if isinstance(args[1], list):
                columns_read = [column_index_from_string(column) for column in args[1]]
            else:
                columns_read = column_index_from_string(args[1]) 
            column_insertion = column_index_from_string(args[2]) 

            # Apply the function on column indexes
            sheet = self.file.writebook[args[0]]
            sheet.insert_cols(column_insertion)
            function(self, args[0], columns_read, column_insertion, *args[3:], **kwargs)

            # Update eventual formulas and save
            self.updateCellFormulas(sheet, True, 'column', modifications)         
            #self.file.writebook.save(self.file.path + self.file.name_file)
        return wrapper
    
    @act_on_columns
    def map_two_columns_to_a_third_column(self, sheet_name, columns_read, column_insertion, mapping, line_beginning=2):
        """
        Vous avez deux colonnes de lecture, suivant ce qui est écrit sur une ligne, vous voulez ou non insérer quelque chose 
        dans une nouvelle colonne.

        Inputs:
            - columns_read (list[str]): liste de deux lettres contenant les colonnes de lecture.
            - column_insertion (str): lettre de la colonne où l'insertion doit avoir lieu.
            - mapping (dict): dictionnaire dont les clés sont les chaînes à écrire. Les valeurs sont dans l'ordre les 
            str qui si elles sont présentes, entraînent l'écriture de ces chaînes.
            - line_beggining (int) : ligne où débute la recherche.
        """ 
        sheet = self.file.writebook[sheet_name]

        for i in range(line_beginning, sheet.max_row + 1):
            # Fill the new column if columns read contain some expected values
            for key, value in mapping.items():
                value1 = str(sheet.cell(i, columns_read[0]).value)
                value2 = str(sheet.cell(i, columns_read[1]).value) 
                if [value1, value2] == value:
                    sheet.cell(i, column_insertion).value = key
                    break
    
    @act_on_columns
    def column_get_part_of_str(self, sheet_name, column_read, column_insertion, separator, piece_number, line_beginning=2):
        """
        Vous avez une colonne qui contient une chaîne dont vous voulez prendre le début jusqu'à un certain séparateur.
        Ce mot est inséré dans une nouvelle colonne.
        
        Inputs:
            - column_read (str): lettre de la colonne de lecture.
            - column_insertion (str): lettre de la colonne où l'insertion doit avoir lieu.
            - separator (str): le symbole délimitant le début du mot
            - piece_number (int): l'index du morceau à prendre (début : 0)
            - line_beggining (int) : ligne où débute la recherche.
        """
        sheet = self.file.writebook[sheet_name] 

        # Fill cells of the new columns 
        for i in range(line_beginning,sheet.max_row + 1):  
            sheet.cell(i, column_insertion).value = sheet.cell(i, column_read).value.split(separator)[piece_number]  
             
    @act_on_columns
    def column_for_prime_probe_congruence(self, sheet_name, columns_read, column_insertion, line_beginning=2):
        """
        Vous avez trois colonnes l'une contient des chaines de caractères particulières qui sont prime, probe, croix de fixation ...
          Les deux autres contiennent des chaines de la forme MOTnb_.jpg où MOT peut 
        être congruent, neutre, incongruent et nb est un nombre. Vous souhaitez insérer une colonne contenant soit rien, soit prime
        suivi du MOT de la deuxième colonne si la chaîne de la première colonne est prime, soit probe suivi du MOT de la troisième 
        colonne si la chaîne de la première colonne est probe.

        Inputs:
            - columns_read (list[str]): the three columns, the two lasts contains MOTnb_.jpg and the first contains prime, probe.
            - column_insertion (str): lettre de la colonne où l'insertion doit avoir lieu. 
            - line_beggining (int) : ligne où débute la recherche.
        """ 
        sheet = self.file.writebook[sheet_name] 

        for i in range(line_beginning, sheet.max_row + 1):

            # Adjonction de la chaine de first_column à MOT
            if sheet.cell(i, columns_read[0]).value in ["prime", "Prime"]:
                mot = re.sub(r'([A-Z-a-z]+)\d+_[A-Z-a-z].jpg', r'\1', sheet.cell(i, columns_read[1]).value)
                sheet.cell(i,column_insertion).value = sheet.cell(i, columns_read[0]).value + "_" + mot 

            elif sheet.cell(i, columns_read[0]).value == ["probe", "Probe"]:
                mot = re.sub(r'([A-Z-a-z]+)\d+_[A-Z-a-z].jpg', r'\1', sheet.cell(i, columns_read[2]).value)
                sheet.cell(i,column_insertion).value = sheet.cell(i, columns_read[0]).value + "_" + mot 

    @act_on_columns
    def give_names_of_maximum(self, sheet_name, column_list, column_insertion, line_beggining = 2):
        """
        Vous avez une liste de colonnes avec des chiffres, chaque colonne a un nom dans sa première cellule. 
        Cette fonction crée une colonne dans laquelle on entre pour chaque ligne le nom de la colonne ou des colonnes qui contient le max.

        Inputs : 
            - column_insertion : 
            - columnlist :
        """ 
        sheet = self.file.writebook[sheet_name] 
        sheet.cell(1, column_insertion).value = "Colonne de(s) maximum(s)"

        #dico qui à une colonne associe le nom de la colonne
        dico = {}
        for column in column_list:
            dico[column] = sheet.cell(1,column).value
 
        for line in range(line_beggining, sheet.max_row + 1):
            #pour une ligne donnée, on récupère le nom de la colonne associé aux maximum(s).
            maximum = -1
            chaine = ""
            for column in column_list:
                cellvalue = sheet.cell(line, column).value
                if cellvalue > maximum:
                    maximum = cellvalue
                    chaine = dico[column]
                elif cellvalue == maximum:
                    chaine += "_" + dico[column]
            sheet.cell(line, column_insertion).value = chaine 

    @act_on_columns 
    def column_transform_string_in_binary(self, sheet_name, column_read, column_write,*good_answers,line_beginning = 2):
        """
        Fonction qui prend une colonne de chaîne de caractères et qui renvoie une colonne de 0 ou de 1
        L'utilisateur doit indiquer un numéro de colonne de lecture et un numéro de colonne où mettre les 0 ou 1.

        Inputs :
                column_read : l'étiquette de la colonne de lecture des réponses.
                colum_write : l'étiquette de la colonne d'écriture des 0 et 1. Par défaut, une colonne est insérée à cette position.
                good_answers : une séquence d'un nombre quelconque de bonnes réponses qui valent 1pt. Chaque mot ne doit pas contenir d'espace ni au début ni à la fin.
                line_beggining: (optionnel par défaut égaux à 2) : ligne où débute l'application de la fonction. 

        Output : rien sauf si la security est enclenchée et que l'on écrit dans une colonne déjà remplie.

        Exemple d'utilisation : 
        
            sheet = Sheet('dataset.xlsx','onglet1')
            sheet.column_transform_string_in_binary('A','B','reponse1','reponse2') 

            #Bien mettre les réponses de good_answers entre ''. 
        """  
        sheet = self.file.writebook[sheet_name] 

        for i in range(line_beginning, sheet.max_row + 1):
            chaine_object = Str(sheet.cell(i,column_read).value)   
            bool = chaine_object.clean_string().transform_string_in_binary(*good_answers) 
            sheet.cell(i,column_write).value = bool

    @act_on_columns 
    def column_convert_in_minutes(self, sheet_name, column_read,column_write,line_beginning = 2):
        """
        Fonction qui prend une colonne de chaines de caractères de la forme "10 jours 5 heures" 
        ou "5 heures 10 min" ou "10 min 5s" ou "5s" et qui renvoie le temps en minutes.
        L'utilisateur doit indiquer un numéro de colonne de lecture et un numéro de colonne à remplir.
        Input : column_read : l'étiquette de la colonne de lecture des réponses.
                colum_write : l'étiquette de la colonne d'écriture. 
                line_beggining: (optionnel par défaut égaux à 2) : ligne où débute l'application de la fonction. 

        Output : rien sauf si la security est enclenchée et que l'on écrit dans une colonne déjà remplie.
        
        Exemple d'utilisation : 
        
            sheet = Sheet('dataset.xlsx','onglet1')
            sheet.column_convert_in_minutes('A','B',line_beggining = 3) 

        """  
        sheet = self.file.writebook[sheet_name] 

        for i in range(line_beginning, sheet.max_row + 1):
            chaine_object = Str(sheet.cell(i,column_read).value) 
            if chaine_object.chaine != "None": 
                bool = chaine_object.clean_string().convert_time_in_minutes() 
                sheet.cell(i,column_write).value = bool

    @act_on_columns 
    def column_set_answer_in_group(self, sheet_name, column_read,column_write,groups_of_responses,line_beginning = 2):
        """
        Dans le cas où il y a des groupes de réponses, cette fonction qui prend une colonne de chaîne de caractères 
        et qui renvoie une colonne remplie de chaînes contenant pour chaque cellule le groupe associé.
        L'utilisateur doit indiquer un numéro de colonne de lecture et un numéro de colonne où écrire.

        Input : 
                column_read : l'étiquette de la colonne de lecture des réponses.
                colum_write : l'étiquette de la colonne d'écriture. 
                groups_of_response : dictionnary which keys are response groups and which values are a list of responses 
        associated to this group.
                line_beggining: (optionnel par défaut égaux à 2) : ligne où débute l'application de la fonction. 

        Output : rien sauf si la security est enclenchée et que l'on écrit dans une colonne déjà remplie.
        
        Exemple d'utilisation : 
        
            sheet = Sheet('dataset.xlsx','onglet1')
            sheet.column_set_answer_in_group('A', 'B', {"group1":['2','5','6'], "group2":['7','8','9'], "group3":['1','3','4'], "group4":['10']} ,line_beggining = 3) 

        """ 
        sheet = self.file.writebook[sheet_name] 
        reversed_group_of_responses = Other.reverse_dico_for_set_answer_in_group(groups_of_responses)

        for i in range(line_beginning,sheet.max_row + 1): 
            chaine_object = Str(sheet.cell(i,column_read).value)  
            group = chaine_object.clean_string().set_answer_in_group(reversed_group_of_responses) 
            sheet.cell(i,column_write).value = group 


class PathControler(FileControler):
    def __init__(self, path):
        """Input : path (object of the class Path)"""
        self.path = path

    def apply_method_on_homononymous_files(self, filename, method_name, *args, **kwargs):
        """ 
        Vous avez plusieurs dossiers contenant un fichier ayant le même nom.
        Fonction qui prend tous les fichiers d'un même nom et qui lui applique une même méthode.  

        Inputs:
            - filename (str)
            - method_name (str): the name of the method to execute 
            - *args, **kwargs : arguments of the method associated with method_name
        """
        start = time()

        # Récupérer tous les dossiers d'un dossier  
        for directory in self.path.directories:
            file = File(filename, self.path.pathname + directory + '/')
            controler = FileControler(file)
            method = getattr(controler, method_name)
            method(*args, **kwargs) 
            Other.display_running_infos(method_name, directory, self.path.directories, start)

    def apply_method_on_homononymous_sheets(self, filename, sheetname, method_name, *args, **kwargs):
        """ 
        Vous avez plusieurs dossiers contenant un fichier ayant le même nom.
        Fonction qui prend tous les fichiers d'un même nom et qui lui applique une même méthode.  

        Inputs:
            - filename (str)
            - method_name (str): the name of the method to execute 
            - *args, **kwargs : arguments of the method associated with method_name
        """
        start = time()

        # Récupérer tous les dossiers d'un dossier  
        for directory in self.path.directories: 
            file = File(filename, self.path.pathname + directory + '/')
            controler = FileControler(file) 
            method = getattr(controler, method_name)
            method(sheetname, *args, **kwargs) 
            Other.display_running_infos(method_name, directory, self.path.directories, start)
           
    def gather_files_in_different_directories(self, name_file, name_sheet, values_only=False):
        """
        Vous avez plusieurs dossiers contenant un fichier ayant le même nom. Vous souhaitez créer un seul fichier regroupant 
        toutes les lignes de ces fichiers.

        Inputs:
            - name_file(str)
            - name_sheet(str)
            - values_only(bool): to decide whether or not copying only the values and not formulas
        """
        # Récupérer tous les dossiers d'un dossier
        directories = [f for f in os.listdir(self.path.pathname) if os.path.isdir(os.path.join(self.path.pathname, f))]

        # Créer un nouveau fichier
        new_file = openpyxl.Workbook() 
        new_sheet = new_file.worksheets[0] 

        start = time()

        # Récupérer le fichier dans chacun des dossiers
        for directory in directories: 
            sheet_to_copy = File(name_file, self.path.pathname + directory + '/').writebook[name_sheet]

            # Copier une fois la première ligne
            if directory == directories[0]:
                self.copy_paste_line(sheet_to_copy, 1, new_sheet, 1, values_only=values_only)

            # Copier son contenu à la suite du fichier
            for line in range(2, sheet_to_copy.max_row + 1): 
                if line % 200 == 0:
                    print(line, sheet_to_copy.max_row + 1)
                self.add_line_at_bottom(sheet_to_copy, line, new_sheet, values_only=values_only)

            # save at the end of each directory not to use too much memory
            new_file.save(self.path.pathname  + "gathered_" + name_file)
            Other.display_running_infos('gather_files_in_different_directories', directory, directories, start)

    def create_one_onglet_by_participant(self, name_file, onglet_from, column_read, first_line=2):
        """
        VERSION ALTERNATIVE A APPLYHOMOGENEOUSFILES DOC OBSOLETE
        Fonction qui prend un onglet dont une colonne contient des chaînes de caractères comme par exemple un nom.
        Chaque chaîne de caractères peut apparaître plusieurs fois dans cette colonne (exe : quand un participant répond plusieurs fois)
        La fonction retourne un fichier contenant un onglet par chaîne de caractères.
          Chaque onglet contient toutes les lignes correspondant à cette chaîne de caractères.

        Input : 
            name_file (str): name of the file to divide
            onglet_from : onglet de référence.
            column_read : l'étiquette de la colonne qui contient les chaînes de caractères.
            first_line : ligne où commencer à parcourir.
            last_line : ligne de fin de parcours 
 
        Exemple d'utilisation : 
    
            file = File('dataset.xlsx')
            file.create_one_onglet_by_participant('onglet1', 'A') 
        """ 
        directories = [f for f in os.listdir(self.path.pathname) if os.path.isdir(os.path.join(self.path.pathname, f))]

        # Créer un nouveau fichier
        new_file = openpyxl.Workbook()  
        onglets = new_file.sheetnames
        column_read = column_index_from_string(column_read)  
        start = time()

        for directory in directories:
            file = File(name_file, self.path.pathname + directory + '/')
            sheet = file.writebook[onglet_from] 

            # Create one tab by identifiant containing all its lines
            for i in range(first_line, sheet.max_row + 1):
                onglet = str(sheet.cell(i,column_read).value)

                # Prepare a new tab
                if onglet not in onglets:
                    new_file.create_sheet(onglet)
                    self.copy_paste_line(sheet, 1,  new_file[onglet], 1)
                    onglets.append(onglet) 

                self.add_line_at_bottom(sheet, i, new_file[onglet]) 
            Other.display_running_infos('create_one_onglet_by_participant', directory, directories, start)
            
        # Deletion of the first tab 
        del new_file[new_file.sheetnames[0]]
        new_file.save(self.path.pathname + f'divided_{name_file}')