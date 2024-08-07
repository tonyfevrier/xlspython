#from xlutils.copy import copy 

import openpyxl
import json 
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, coordinate_to_tuple, get_column_letter
from copy import copy
from datetime import datetime
from utils import UtilsForFile, UtilsForSheet, Str, Other
from pycel import ExcelCompiler
import gc

class Path():
    def __init__(self,path = 'fichiers_xls/'):
        self.path = path
        
    def act_on_files(self,fonction):
        """
        Fonction qui prend tous les fichiers d'un dossier et qui applique une même action à ces fichiers.
        """
        pass 


class File(UtilsForFile): 
    def __init__(self, name_file, path = 'fichiers_xls/', dataonly = False): #True):
        """L'utilisateur sera invité à mettre son fichier xslx dans un dossier nommé fichiers_xls
        """ 
        self.name_file = name_file  
        self.path = path
        self.dataonly = dataonly 

        self.writebook = openpyxl.load_workbook(self.path + self.name_file, data_only = dataonly)
        self.sheets_name = self.writebook.sheetnames
        self.compiler = ExcelCompiler(self.path + self.name_file)  
       

    def sauvegarde(self):
        """
        Fonction qui crée une sauvegarde du fichier name_file et 
        qui l'appelle name_file_time où time est le moment d'enregistrement.

        Exemple d'utilisation : 
            file = File('dataset.xlsx')
            file.sauvegarde()

            Si on veut copier les formules et pas seulement les valeurs des cellules : 
            file = File('dataset.xlsx', dataonly = False)
            file.sauvegarde()
        """
        file_copy = openpyxl.Workbook()
        del file_copy[file_copy.active.title] #supprimer l'onglet créé

        for onglet in self.sheets_name:
            new_sheet = file_copy.create_sheet(onglet)
            initial_sheet = self.writebook[onglet] 

            for i in range(1,initial_sheet.max_row+1):
                for j in range(1,initial_sheet.max_column+1): 
                    new_sheet.cell(i,j).value = initial_sheet.cell(i,j).value  
                    new_sheet.cell(i,j).fill = copy(initial_sheet.cell(i,j).fill)
                    new_sheet.cell(i,j).font = copy(initial_sheet.cell(i,j).font) 
                    
        name_file_no_extension = Str(self.name_file).del_extension() 
        
        file_copy.save(self.path  + name_file_no_extension + '_date_' + datetime.now().strftime("%Y-%m-%d_%Hh%M") + '.xlsx') 
        return file_copy
            

    def create_one_onglet_by_participant(self, onglet_from, column_read, first_line=2, label=True):
        """
        Fonction qui prend un onglet dont une colonne contient des chaînes de caractères comme par exemple un nom.
        Chaque chaîne de caractères peut apparaître plusieurs fois dans cette colonne (exe : quand un participant répond plusieurs fois)
        La fonction retourne un fichier contenant un onglet par chaîne de caractères.
          Chaque onglet contient toutes les lignes correspondant à cette chaîne de caractères.

        Input : 
            onglet_from : onglet de référence.
            column_read : l'étiquette de la colonne qui contient les chaînes de caractères.
            first_line : ligne où commencer à parcourir.
            last_line : ligne de fin de parcours
            label : bool. Mettre sur False si on souhaite entrer les colonnes par leurs positions plutôt que leur label.
 
        Exemple d'utilisation : 
    
            file = File('dataset.xlsx')
            file.create_one_onglet_by_participant('onglet1', 'A') 
        """ 
        if label:
            column_read = column_index_from_string(column_read)  

        onglets = []
        sheet = self.writebook[onglet_from] 

        for i in range(first_line, sheet.max_row + 1):
            onglet = str(sheet.cell(i,column_read).value)
            if onglet not in onglets:
                self.writebook.create_sheet(onglet)
                self.copy_paste_line(sheet, 1,  self.writebook[onglet], 1)
                onglets.append(onglet) 
            self.add_line_at_bottom(sheet, i, self.writebook[onglet])
         
        self.writebook.save(self.path + self.name_file)
        self.sheets_name = self.writebook.sheetnames 

    def extract_column_from_all_sheets(self,column,label = True):
        """
        Fonction qui récupère une colonne dans chaque onglet pour former une nouvelle feuille
        contenant toutes les colonnes. La première cellule de chaque colonne correspond alors 
        au nom de l'onglet. Attention, en l'état, il faut que tous les onglets aient la même structure.

        Input : 
            column : str. L'étiquette de la colonne à récupérer dans chaque onglet
            label : bool. Mettre sur False si on souhaite entrer les colonnes par leurs positions plutôt que leur label.

        Exemple d'utilisation : 
    
            file = File('dataset.xlsx')
            file.extract_column_from_all_sheets('B') 

            Si on veut extraire les formules

            file = File('dataset.xlsx',dataonly = False)
            file.extract_column_from_all_sheets('B') 
        """ 
        if label:
            column = column_index_from_string(column)
         
        new_sheet = self.writebook.create_sheet(f"gather_{column}")
        column_to = 1
        for name_onglet in self.sheets_name:
            onglet = self.writebook[name_onglet] 
            self.copy_paste_column(onglet,column,new_sheet,column_to)
            column_to = new_sheet.max_column + 1
            new_sheet.cell(1,new_sheet.max_column).value = name_onglet 
            
        self.writebook.save(self.path + self.name_file) 
        self.sheets_name = self.writebook.sheetnames 

    def apply_column_formula_on_all_sheets(self, *column_list, label = True):
        """
        Fonction qui reproduit les formules d'une colonne ou plusieurs colonnes
          du premier onglet sur toutes les colonnes situées à la même position dans les 
          autres onglets.

        Input : 
            -column_list : int. les positions des colonnes où récupérer et coller.

        Exemples d'utilisation : 

            Bien veiller à mettre dataonly = False sinon il ne copiera pas les formules mais
            les valeurs des cellules. On peut aussi copier les valeurs des cellules : pour cela,
            enlever dataonly = False.

            Sur une colonne
                file = File('dataset.xlsx', dataonly = False)
                file.apply_column_formula_on_all_sheets(2) 

            Sur trois colonnes
                file = File('dataset.xlsx', dataonly = False)
                file.apply_column_formula_on_all_sheets(2,5,10) 

            Sur toutes les colonnes du fichier à partir de la colonne colmin jusque la colonne colmax :
                file = File('dataset.xlsx', dataonly = False)
                file.apply_column_formula_on_all_sheets(*[i for i in range(colmin,colmax + 1)]) 
        """
        if label:
            column_int_list = []
            for column in column_list: 
                column_int_list.append(column_index_from_string(column))  

        #on applique les copies dans tous les onglets sauf le premier
        for name_onglet in self.sheets_name[1:]:
            for column in column_int_list:
                self.copy_paste_column(self.writebook[self.sheets_name[0]],column,self.writebook[name_onglet],column)

        self.writebook.save(self.path + self.name_file)

    def apply_cells_formula_on_all_sheets(self, *cells):
        """
        Fonction qui reproduit les formules d'une cellule ou plusieurs cellules
          du premier onglet sur toutes les cellules situées à la même position dans les 
          autres onglets.

        Input : 
            -cells : string. les positions des cellule où récupérer et coller 

        Exemples d'utilisation : 

            Bien veiller à mettre dataonly = False sinon il ne copiera pas les formules mais
            les valeurs des cellules. On peut aussi copier les valeurs des cellules : pour cela,
            enlever dataonly = False.

            file = File('dataset.xlsx', dataonly = False)
            file.apply_column_formula_on_all_sheets('C5','D6')  
        """

        #obtenir les indices de la cellule 
        cell_list = []
        for cell in cells: 
            cell_list.append(coordinate_to_tuple(cell)) 

        #on applique les copies dans tous les onglets sauf le premier
        for name_onglet in self.sheets_name[1:]:  
            for tuple in cell_list: 
                self.writebook[name_onglet].cell(tuple[0],tuple[1]).value = self.writebook[self.sheets_name[0]].cell(tuple[0],tuple[1]).value  
                self.writebook[name_onglet].cell(tuple[0],tuple[1]).fill = copy(self.writebook[self.sheets_name[0]].cell(tuple[0],tuple[1]).fill)  
                self.writebook[name_onglet].cell(tuple[0],tuple[1]).font = copy(self.writebook[self.sheets_name[0]].cell(tuple[0],tuple[1]).font)  
                self.writebook[name_onglet].cell(tuple[0],tuple[1]).border = copy(self.writebook[self.sheets_name[0]].cell(tuple[0],tuple[1]).border)  
                self.writebook[name_onglet].cell(tuple[0],tuple[1]).alignment = copy(self.writebook[self.sheets_name[0]].cell(tuple[0],tuple[1]).alignment)    

        self.writebook.save(self.path + self.name_file)

    def gather_columns_in_one(self,onglet, *column_lists):
        """
        Vous avez des groupes de colonnes de valeurs avec une étiquette en première cellule. Pour chaque groupe, vous souhaitez former deux colonnes de valeurs : l'une qui contient
        les valeurs rassemblées en une colonne, l'autre, à sa gauche, qui indique l'étiquette de la colonne dans laquelle elle a été prise.

        Inputs : 
            - onglet (str) : nom de l'onglet d'où on importe les colonnes.
            - column_lists (list[list[str]]) : liste de groupes de colonnes. Chaque groupe est une liste de colonnes.
        """
         
        for list in column_lists:
            tab_number = len(self.writebook.sheetnames)
            self.writebook.create_sheet(f"onglet {tab_number}")
            target_sheet = self.writebook[f"onglet {tab_number}"]
            for column in list: 
                self.copy_column_tags_and_values_at_bottom(self.writebook[onglet], column_index_from_string(column), target_sheet)

        self.writebook.save(self.path + self.name_file) 

    def build_file_from_tab(self, tab):
        """
        Fonction qui prend un nom d'onglet dans un fichier et qui crée un fichier associé.

        Input :
            - tab (str) : the name of the tab from which we want to create the file.
        """

        sheet_from = self.writebook[tab]
        newfile = openpyxl.Workbook() 
        sheet_to = newfile['Sheet']  
        path = 'multifiles/' 
  
        self.deep_copy_of_a_sheet(sheet_from, sheet_to) 

        namefile = path + tab + '.xlsx'
        newfile.save(namefile) 
        return namefile
            
            
    def one_file_by_tab_sendmail(self, send = False, adressjson = "", objet = "", message = ""):
        """
        Vous souhaitez fabriquer un fichier par onglet. Chaque fichier aura le nom de l'onglet. 
        Vous souhaitez éventuellement envoyer chaque fichier à la personne associée.
        Attention, pour utiliser cette fonction, les onglets doivent être de la forme "prenom nom" sans caractère spéciaux. 

        Inputs : 
            send(optional boolean) : True si on veut envoyer le mail, False si on veut juste couper en fichiers.
            adressjson(str) : nom du fichier xlsx qui contient deux colonnes la première avec les noms des onglets, la seconde avec l'adresse mail. Ce fichier doit être mis dans le dossier fichier_xls. 
            objet(optional str) : Objet du message.
            message (optional str) : Contenu du message.
        """ 
        if adressjson != "":
            file = open(self.path + adressjson, 'r')
            mailinglist = json.load(file)
            file.close()

        for tab in self.sheets_name: 
            file_to_send = self.build_file_from_tab(tab)
            if send:
                if adressjson == "":
                    prenom = tab.split(" ")[0]
                    nom = tab.split(" ")[1]
                    self.envoi_mail(prenom + "." + nom + "@universite-paris-saclay.fr", file_to_send, "tony.fevrier62@gmail.com", "qkxqzhlvsgdssboh", objet, message)
                else: 
                    self.envoi_mail(mailinglist[tab], file_to_send, "tony.fevrier62@gmail.com", "qkxqzhlvsgdssboh", objet, message)  

    def merge_cells_on_all_tabs(self,start_column, end_column, start_row, end_row):
        """
        Fonction qui merge les mêmes cellules sur tous les onglets d'un fichier 

        Inputs :
            - start_column (string): Letter of the column where to start the merging
            - end_column (string): Letter of the column where to end the merging
            - start_row (int): Index of the row where to start the merging
            - end_row (int): Index of the row where to start the merging

        """
        
        start_column = column_index_from_string(start_column)
        end_column = column_index_from_string(end_column)

        for tab in self.sheets_name:
            sheet = self.writebook[tab] 
            sheet.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)

        self.writebook.save(self.path + self.name_file)

class Sheet(File,UtilsForSheet,Other): 
    def __init__(self, name_file, name_onglet,path = 'fichiers_xls/'): 
        super().__init__(name_file,path)
        self.name_onglet = name_onglet  
        self.sheet = self.writebook[self.name_onglet]
        del self.sheets_name

    def column_transform_string_in_binary(self,column_read,column_write,*good_answers,line_beginning = 2, insert = True, label = True):
        """
        Fonction qui prend une colonne de chaîne de caractères et qui renvoie une colonne de 0 ou de 1
        L'utilisateur doit indiquer un numéro de colonne de lecture et un numéro de colonne où mettre les 0 ou 1.

        Inputs :
                column_read : l'étiquette de la colonne de lecture des réponses.
                colum_write : l'étiquette de la colonne d'écriture des 0 et 1. Par défaut, une colonne est insérée à cette position.
                good_answers : une séquence d'un nombre quelconque de bonnes réponses qui valent 1pt. Chaque mot ne doit pas contenir d'espace ni au début ni à la fin.
                line_beggining: (optionnel par défaut égaux à 2) : ligne où débute l'application de la fonction.
                insert : (paramètre optionnel) le mettre à False si on ne veut pas insérer une colonne.
                label : bool. Mettre sur False si on souhaite entrer les colonnes par leurs positions plutôt que leur label.

        Output : rien sauf si la security est enclenchée et que l'on écrit dans une colonne déjà remplie.

        Exemple d'utilisation : 
        
            sheet = Sheet('dataset.xlsx','onglet1')
            sheet.column_transform_string_in_binary('A','B','reponse1','reponse2') 

            #Bien mettre les réponses de good_answers entre ''. 
        """  
        
        if label:
            column_read = column_index_from_string(column_read) 
            modifications = [column_write]
            column_write = column_index_from_string(column_write)
        else:
            modifications = [get_column_letter(column_write)]

        if insert:
            self.sheet.insert_cols(column_write)

        for i in range(line_beginning, self.sheet.max_row + 1):
            chaine_object = Str(self.sheet.cell(i,column_read).value)  
            bool = chaine_object.clean_string().transform_string_in_binary(*good_answers) 
            self.sheet.cell(i,column_write).value = bool
 
        self.updateCellFormulas(self.sheet,True,'column', modifications)         
        self.writebook.save(self.path + self.name_file)

    def column_convert_in_minutes(self,column_read,column_write,line_beginning = 2, insert = True, label = True):
        """
        Fonction qui prend une colonne de chaines de caractères de la forme "10 jours 5 heures" 
        ou "5 heures 10 min" ou "10 min 5s" ou "5s" et qui renvoie le temps en minutes.
        L'utilisateur doit indiquer un numéro de colonne de lecture et un numéro de colonne à remplir.
        Input : column_read : l'étiquette de la colonne de lecture des réponses.
                colum_write : l'étiquette de la colonne d'écriture. 
                line_beggining: (optionnel par défaut égaux à 2) : ligne où débute l'application de la fonction.
                insert : (paramètre optionnel) le mettre à False si on ne veut pas insérer une colonne.
                label : bool. Mettre sur False si on souhaite entrer les colonnes par leurs positions plutôt que leur label.

        Output : rien sauf si la security est enclenchée et que l'on écrit dans une colonne déjà remplie.
        
        Exemple d'utilisation : 
        
            sheet = Sheet('dataset.xlsx','onglet1')
            sheet.column_convert_in_minutes('A','B',line_beggining = 3) 

        """ 
        if label:
            column_read = column_index_from_string(column_read) 
            modifications = [column_write]
            column_write = column_index_from_string(column_write)
        else:
            modifications = [get_column_letter(column_write)]

        if insert:
            self.sheet.insert_cols(column_write)

        for i in range(line_beginning,self.sheet.max_row + 1):
            chaine_object = Str(self.sheet.cell(i,column_read).value) 
            if chaine_object.chaine != "None": 
                bool = chaine_object.clean_string().convert_time_in_minutes() 
                self.sheet.cell(i,column_write).value = bool
 
        self.updateCellFormulas(self.sheet,True,'column', modifications)         
        self.writebook.save(self.path + self.name_file)

    def column_set_answer_in_group(self,column_read,column_write,groups_of_responses,line_beginning = 2, insert = True, label = True):
        """
        Dans le cas où il y a des groupes de réponses, cette fonction qui prend une colonne de chaîne de caractères 
        et qui renvoie une colonne remplie de chaînes contenant pour chaque cellule le groupe associé.
        L'utilisateur doit indiquer un numéro de colonne de lecture et un numéro de colonne où écrire.

        Input : 
                column_read : l'étiquette de la colonne de lecture des réponses.
                colum_write : l'étiquette de la colonne d'écriture. 
                groups_of_response : dictionnary which keys are response groups and which values are a list of responses 
        associated to this group.
                line_beggining: (optionnel par défaut égaux à 2) : ligne où débute l'application de la fonction.
                insert : (paramètre optionnel) le mettre à False si on ne veut pas insérer une colonne.
                label : bool. Mettre sur False si on souhaite entrer les colonnes par leurs positions plutôt que leur label.

        Output : rien sauf si la security est enclenchée et que l'on écrit dans une colonne déjà remplie.
        
        Exemple d'utilisation : 
        
            sheet = Sheet('dataset.xlsx','onglet1')
            sheet.column_set_answer_in_group('A', 'B', {"group1":['2','5','6'], "group2":['7','8','9'], "group3":['1','3','4'], "group4":['10']} ,line_beggining = 3) 

        """
        if label:
            column_read = column_index_from_string(column_read) 
            modifications = [column_write]
            column_write = column_index_from_string(column_write)
        else:
            modifications = [get_column_letter(column_write)]

        if insert:
            self.sheet.insert_cols(column_write)

        reversed_group_of_responses = self.reverse_dico_for_set_answer_in_group(groups_of_responses)

        for i in range(line_beginning,self.sheet.max_row + 1): 
            chaine_object = Str(self.sheet.cell(i,column_read).value)  
            group = chaine_object.clean_string().set_answer_in_group(reversed_group_of_responses) 
            self.sheet.cell(i,column_write).value = group
            
        self.updateCellFormulas(self.sheet,True,'column', modifications)         
        self.writebook.save(self.path + self.name_file)
        
    def color_special_cases_in_column(self,column,chainecolor,label = True):
        """
        Fonction qui regarde pour une colonne donnée colore les cases correspondant à certaines chaînes de caractères.

        Input : 
            - column : le numéro de la colonne.
            - chainecolor (dict) : les chaînes de caractères qui vont être colorées et les couleurs qui correspondent à écrire avec la syntaxe suivante {'vrai':'couleur1','autre':couleur2}. Attention,
                la couleur doit être entrée en hexadécimal et les chaînes de caractères ne doivent pas avoir d'espace au début ou à la fin.
            - label : bool. Mettre sur False si on souhaite entrer les colonnes par leurs positions plutôt que leur label.
        
        Exemple d'utilisation : 
        
            sheet = Sheet('dataset.xlsx','onglet1')
            sheet.color_special_cases_in_column('L', {'vrai': '#FF0000','faux': '#00FF00'}) 

        """
        if label:
            column = column_index_from_string(column)

        for i in range(1,self.sheet.max_row + 1):
            cellule = self.sheet.cell(i,column) 

            if cellule.value is str:
                key = Str(cellule.value).clean_string().chaine
            else: 
                key = cellule.value

            if key in chainecolor.keys():
                cellule.fill = PatternFill(fill_type = 'solid', start_color = chainecolor[key])

        self.writebook.save(self.path + self.name_file)

    def color_special_cases_in_sheet(self,chainecolor): 
        """
        Fonction qui colore les cases contenant à certaines chaînes de caractères d'une feuille
        
        Input : 
            - column : le numéro de la colonne.
            - chainecolor (dict) : les chaînes de caractères qui vont être colorées et les couleurs qui correspondent à écrire avec la syntaxe suivante {'vrai':'couleur1','autre':couleur2}. Attention,
                la couleur doit être entrée en hexadécimal et les chaînes de caractères ne doivent pas avoir d'espace au début ou à la fin.
            
        Exemple d'utilisation : 
    
            sheet = Sheet('dataset.xlsx','onglet1')
            sheet.color_special_cases_in_sheet({'vrai': '#FF0000','faux': '#00FF00'}) 
                
        """

        for j in range(1, self.sheet.max_column + 1):
            self.color_special_cases_in_column(j,chainecolor,label=False)

    def add_column_in_sheet_differently_sorted(self,column_identifiant, column_insertion, other_sheet,label = True):
        """
        Fonction qui insère dans un onglet des colonnes d'un autre onglet de référence. 
        Les deux feuilles ont une colonne d'identifiants communs (exemple : des mails) mais qui peut être
        triés dans des ordres différents.
        La fonction récupère un ou plusieurs éléments d'une ligne déterminée par un identifiant.
        Elle recherche l'identifiant dans la seconde feuille et insère les éléments
        dans la ligne correspondante.

        Inputs :
            - column_identifiant : numéro de la colonne où sont situés les identifiants dans l'onglet qu'on souhaite modifier.
            - column_insertion : numéro de la colonne où on insère les colonnes à récupérer.
            - other_sheet : liste représentant l'onglet duquel on récupère les colonnes  ['namefile','namesheet',numéro de la colonne où sont les identifiants,[numéros des colonnes à récupérer sous forme de liste]]
                namefile doit être au format .xlsx et mis dans le dossier fichier_xls.
            - label : bool. Mettre sur False si on souhaite entrer les colonnes par leurs positions plutôt que leur label.
            
        Exemple d'utilisation : 
    
            sheet = Sheet('dataset.xlsx','onglet1')
            sheet.add_column_in_sheet_differently_sorted('B','C',['file.xlsx', 'onglet2', 'B', ['E','F','H','AA]]) 
                  
        """
        
        file_to_copy = openpyxl.load_workbook(self.path + other_sheet[0],data_only=True)
        sheet_to_copy = file_to_copy[other_sheet[1]]
        columns_to_copy = other_sheet[3]
        dico = {}

        if label:
            column_identifiant = column_index_from_string(column_identifiant)
            column_insertion = column_index_from_string(column_insertion)
            other_sheet[2] = column_index_from_string(other_sheet[2])
            columns_to_copy = [column_index_from_string(column) for column in columns_to_copy]

        modifications = [get_column_letter(column_insertion + i ) for i in range(len(columns_to_copy))]

        #Passage en revue les identifiants du premier fichier et création d'un dictionnaire dont les clés sont ces identifiants et les valeurs sont une liste de valeurs à récupérer.
        for i in range(1,sheet_to_copy.max_row + 1):
            value = sheet_to_copy.cell(i,other_sheet[2]).value
            dico[value] = [sheet_to_copy.cell(i,j) for j in columns_to_copy]

        self.sheet.insert_cols(column_insertion,len(columns_to_copy)) 

        #Passage en revue des identifiants du second fichier et insertion des valeurs si les identifiants sont dans les clés du dico
        #. 
        for i in range(1,self.sheet.max_row+1):
            key = self.sheet.cell(i,column_identifiant).value
            if key in dico.keys():
                for j in range(len(columns_to_copy)):
                    self.sheet.cell(i,column_insertion + j).value = dico[key][j].value
                    self.sheet.cell(i,column_insertion + j).fill = copy(dico[key][j].fill)
        
        self.updateCellFormulas(self.sheet,True,'column', modifications)         
        self.writebook.save(self.path + self.name_file)


    def color_lines_containing_chaines(self,color,*chaines):
        """
        Fonction qui colore les lignes dont une des cases contient une str particulière.

        Input : 
            - color : une couleur indiquée en haxadécimal par l'utilisateur.
            - chaines : des chaines de caractères que l'utilisateur entre et qui entraînent la coloration de la ligne.
            
        Exemple d'utilisation : 
    
            sheet = Sheet('dataset.xlsx','onglet1')
            sheet.color_lines_containing_chaine('#FF0000', 'vrai', 'hello', 'heeenri', 'ficheux') 
        
        """

        lines_to_color = []

        for i in range(1, self.sheet.max_row + 1):
            for j in range(1, self.sheet.max_column + 1):
                if str(self.sheet.cell(i,j).value) in chaines:
                    lines_to_color.append(i)
                    break
        
        for row in lines_to_color:
            self.color_line(color, row)
        
        self.writebook.save(self.path + self.name_file)

    def column_cut_string_in_parts(self,column_to_cut,column_insertion,separator, label = True):
        """
        Fonction qui prend une colonne dont chaque cellule contient une grande chaîne de
          caractères. Toutes les chaînes sont composés du nombre de morceaux délimités par un séparateur,
        La fonction insère autant de colonnes que de morceaux et place un morceau par colonne dans l'ordre des morceaux.

        Inputs :
            - column_to_cut : colonne contenant les grandes str.
            - column_insertion : où insérer les colonnes
            - separator le séparateur
            - label : bool. Mettre sur False si on souhaite entrer les colonnes par leurs positions plutôt que leur label.

        Exemple d'utilisation : 
    
            sheet = Sheet('dataset.xlsx','onglet1')
            sheet.column_cut_string_in_parts('C', 'J', ';') 
        
        """

        if label:
            column_to_cut = column_index_from_string(column_to_cut) 
            column_insertion = column_index_from_string(column_insertion)
        
        for i in range(2, self.sheet.max_row + 1):
            value = self.sheet.cell(i,column_to_cut).value
            chaine_object = Str(value)
            parts = chaine_object.cut_string_in_parts(separator)
            modifications = [get_column_letter(column_insertion + i) for i in range(len(parts))]
            if i == 2:
                self.sheet.insert_cols(column_insertion,len(parts))
            for j in range(len(parts)):
                self.sheet.cell(i,column_insertion + j).value = parts[j]

        self.updateCellFormulas(self.sheet,True,'column', modifications)         
        self.writebook.save(self.path + self.name_file) 

    def delete_lines(self,column,*chaines,label = True):
        """
        Fonction qui parcourt une colonne et qui supprime la ligne si celle-ci contient une chaîne particulière.

        Inputs : 
            -column : la colonne à parcourir.
            -chaines : les chaînes de caractères qui doivent engendrer la suppression de la ligne.
        
        Exemple d'utilisation : 
    
            sheet = Sheet('dataset.xlsx','onglet1')
            sheet.delete_lines(3, 'chaine1', 'chaine2', 'chaine3', 'chaine4') 
        """
        if label:
            column = column_index_from_string(column)  
        
        modifications = []
        for i in range(self.sheet.max_row,0,-1):
            value = self.getCellNumericalValue(self.compiler, self.name_onglet, self.sheet.cell(i,column)) 
            if str(value) in chaines: 
            #if str(self.sheet.cell(i,column).value) in chaines:  ) 
                self.sheet.delete_rows(i) 
                modifications.append(str(i))
 
        self.updateCellFormulas(self.sheet,False,'row',modifications)        
        self.writebook.save(self.path + self.name_file)

    def delete_doublons(self, column_identifiant, line_beginning = 2, color = False, label = True):
        """
        Certains participants répondent plusieurs fois. Cette fonction supprime les premières réponses
        des participants dans ce cas. Elle ne garde que leur dernière réponse. On repère les participants
        par leur identifiant unique donné dans colum_identifiant.

        Inputs:
            column_identifiant : str: lettre de la colonne qui contient les identifiants des participants.
            color : boolean : True si on veut que la ligne des participants qui ont répondu plusieurs fois soit colorée dans le datasetfinal.
        
        Exemple d'utilisation : 
    
        si on ne veut pas repérer les personnes qui étaient en doublon:
            sheet = Sheet('dataset.xlsx','onglet1')
            sheet.delete_doublons('C')

        si on veut les repérer :
            sheet = Sheet('dataset.xlsx','onglet1')
            sheet.delete_doublons('C', color = True)
        """
        if label:
            column_identifiant = column_index_from_string(column_identifiant) 

        participants = {} 
        modifications = []
        #On parcourt dans le sens inverse afin d'éviter que la suppression progressive impacte la position des lignes étudiées ensuite.
        i = self.sheet.max_row 
        while i != line_beginning:  
            identifiant = Str(self.sheet.cell(i,column_identifiant).value).clean_string() 
            if identifiant.chaine in participants.keys():
                if color:
                    self.color_line('0000a933', participants[identifiant.chaine])
                self.sheet.delete_rows(i)
                modifications.append(str(i))
                participants[identifiant.chaine] -= 1    
            else:
                participants[identifiant.chaine] = i 
            i -= 1

        self.updateCellFormulas(self.sheet,False,'row',modifications)        
        self.writebook.save(self.path + self.name_file)
    
    def create_one_column_by_QCM_answer(self, column, column_insertion, list_string, *reponses, label = True):
        """
        Fonction qui regarde si des réponses sont incluses dans les cellules d'une colonne.
        Chaque cellule contient l'ensemble des réponses à une question de QCM du participant sous forme de str.
        Elle regarde les cellules de column. Si une réponse est dans cette cellule, on l'indique dans la colonne
        correspondante.

        Inputs : 
            - column :  str : la colonne avec les réponses.
            - column_insertion : str : l'endroit où on insère les colonnes.
            - list_string : list : liste de deux str indiquant si la réponse est présente ou non.
            - reponses : les réponses du QCM.
            - label : bool : True si on entre les colonnes par leurs étiquettes, False sinon.
        
        Exemple : 
            sheet = Sheet('dataset.xlsx','onglet1')
            sheet.create_one_column_by_QCM_answer('C','D', ['oui', 'non'], 'reponse1', 'reponse2', 'reponse3')

        """
 
        if label:
            column = column_index_from_string(column) 
            column_insertion = column_index_from_string(column_insertion) 
        
        modifications = [get_column_letter(column_insertion + i) for i in range(len(reponses))]

        #on crée les colonnes pour chaque réponse
        self.sheet.insert_cols(column_insertion,len(reponses))
        for j in range(0,len(reponses)):
            self.sheet.cell(1,j + column_insertion).value = reponses[j]

        #on remplit les colonnes suivant que les réponses correspondantes sont ou non dans la cellule.
        for i in range(2, self.sheet.max_row + 1):
            if self.sheet.cell(i,column).value == None:
                for j in range(0,len(reponses)):  
                        self.sheet.cell(i,j + column_insertion).value = list_string[1]
            else:
                for j in range(0,len(reponses)):  
                    if reponses[j] in self.sheet.cell(i,column).value:
                        self.sheet.cell(i,j + column_insertion).value = list_string[0]
                    else:
                        self.sheet.cell(i,j + column_insertion).value = list_string[1]

        self.updateCellFormulas(self.sheet,True,'column',modifications)        
        self.writebook.save(self.path + self.name_file)
        
    def gather_multiple_answers(self, column_read, column_store, line_beggining = 2, label = True):
        """
        Dans un onglet, nous avons les réponses de participants qui ont pu répondre plusieurs fois à un questionnaire.
        Cette fonction parcourt les noms et met dans un autre onglet. La ligne du participant est alors constituée des différentes valeurs
         d'une même donnée récupérée.
        
        Inputs :
            - column_read (str) : la colonne avec les identifiants des participants.
            - column_store (str) : lettre de la colonne contenant la donnée qu'on veut stocker.
            - line_beggining (int) : ligne où débute la recherche.
            - label (bool) : False si on veut entrer le numéro de la colonne et pas la lettre.
        """ 
        if label:
            column_read = column_index_from_string(column_read) 
            column_store = column_index_from_string(column_store) 

        #we create a dictionary whose keys are the identifiers (of participants) and values are their number of answers and a list containing
        #the data we want to store for each answer.
        dico = self.create_dico_to_store_multiple_answers_of_participants(column_read,column_store,line_beggining)
        
        #we create the new sheet where we store participants answering multiple times and their data.
        storesheet = self.writebook.create_sheet('severalAnswers')
        self.create_newsheet_storing_multiple_answers(storesheet, dico)

        self.writebook.save(self.path + self.name_file) 

    def give_names_of_maximum(self, column_insertion, *columnlist):
        """
        Vous avez une liste de colonnes avec des chiffres, chaque colonne a un nom dans sa première cellule. Cette fonction crée une colonne dans laquelle on entre pour chaque 
        ligne le nom de la colonne ou des colonnes qui contient le max.

        Inputs : 
            - column_insertion : 
            - columnlist :
        """

        number_column_insertion = column_index_from_string(column_insertion)
        self.sheet.insert_cols(number_column_insertion)
        modifications = [number_column_insertion]
        self.sheet.cell(1, number_column_insertion).value = "Colonne de(s) maximum(s)"

        #dico qui à une colonne associe le nom de la colonne
        dico = {}
        for column in columnlist:
            dico[column] = self.sheet.cell(1,column_index_from_string(column)).value
 
        for line in range(2, self.sheet.max_row + 1):
            #pour une ligne donnée, on récupère le nom de la colonne associé aux maximum(s).
            maximum = -1
            chaine = ""
            for column in columnlist:
                cellvalue = self.sheet.cell(line, column_index_from_string(column)).value
                if cellvalue > maximum:
                    maximum = cellvalue
                    chaine = dico[column]
                elif cellvalue == maximum:
                    chaine += "_" + dico[column]
            self.sheet.cell(line, number_column_insertion).value = chaine
        
        self.updateCellFormulas(self.sheet, True, 'column', modifications)
        self.writebook.save(self.path + self.name_file) 
        

    

        






   
      


